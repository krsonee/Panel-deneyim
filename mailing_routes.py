"""Mailing modülü API rotaları — CRM, şablon, kampanya, IVR, rapor iskeleti."""

from __future__ import annotations

import csv
import html as html_lib
import io
import json
import os
import re
import secrets
import threading
import time
import urllib.parse
from contextlib import closing

from flask import Blueprint, jsonify, redirect, request

import smartico_api
from database import (
    execute,
    fetchall,
    fetchone,
    get_db,
    get_mail_setting,
    insert_returning_id,
    iso,
    safe_rollback,
    scalar,
    upsert_mail_setting,
    utcnow,
    uses_postgres,
)

IMPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mail_imports")
IMPORT_CHUNK_SIZE = 5000
# Tek seferde ~100M e-posta (yaklaşık 4-5 GB CSV) yüklenebilsin diye üst sınır.
IMPORT_MAX_BYTES = 5 * 1024 * 1024 * 1024
# Yükleme isteği yarıda kesilirse (proxy timeout, bağlantı kopması) pending işler
# panelde "hiçbir şey yok" gibi görünüyordu — bu süre sonra hata olarak işaretlenir.
IMPORT_STALE_PENDING_SECONDS = 10 * 60
IMPORT_STALE_RUNNING_SECONDS = 15 * 60

EMAIL_HEADER_ALIASES = frozenset({
    "email", "e-posta", "eposta", "e-mail", "e_mail", "mail",
})
EMAIL_COLUMN_KEYS = (
    "email", "Email", "EMAIL",
    "E-posta", "eposta", "Eposta",
    "mail", "e_mail", "e-posta",
)


def _ensure_import_dir():
    os.makedirs(IMPORT_DIR, exist_ok=True)


def _import_job_path(job_id, filename):
    ext = os.path.splitext(filename or "")[1].lower()
    if ext not in (".csv", ".xlsx", ".xlsm"):
        ext = ".csv"
    return os.path.join(IMPORT_DIR, f"job_{job_id}{ext}")


def _import_job_age_seconds(iso_str):
    if not iso_str:
        return IMPORT_STALE_PENDING_SECONDS + 1
    try:
        from datetime import datetime, timezone

        ref = datetime.fromisoformat(str(iso_str).replace("Z", "+00:00"))
        if ref.tzinfo is None:
            ref = ref.replace(tzinfo=timezone.utc)
        return (utcnow() - ref).total_seconds()
    except Exception:
        return IMPORT_STALE_PENDING_SECONDS + 1


def _normalize_header_key(key):
    return (key or "").strip().lower().replace("_", "-")


def _find_email_column_index(header):
    for i, h in enumerate(header):
        if _normalize_header_key(h) in EMAIL_HEADER_ALIASES:
            return i
    return None


def _values_look_like_email_column(rows):
    """İlk sütundaki değerlerin çoğu geçerli e-posta mı (başlıksız tek sütun listesi)."""
    total = 0
    emails = 0
    for row in rows:
        if not row:
            continue
        val = row[0] if len(row) > 0 else None
        if val is None or str(val).strip() == "":
            continue
        total += 1
        if EMAIL_RE.match(str(val).strip()):
            emails += 1
    return total > 0 and emails >= max(1, int(total * 0.8))


def _extract_email_from_row(row):
    """Satırdan e-posta çıkar — bilinen sütun adları ve tek sütunlu listeler."""
    if not row:
        return ""
    for key in EMAIL_COLUMN_KEYS:
        email = _normalize_email_candidate(row.get(key))
        if email:
            return email
    for key, val in row.items():
        if _normalize_header_key(key) in EMAIL_HEADER_ALIASES:
            email = _normalize_email_candidate(val)
            if email:
                return email
    if len(row) == 1:
        email = _normalize_email_candidate(next(iter(row.values()), ""))
        if email:
            return email
    for val in row.values():
        email = _normalize_email_candidate(val)
        if email:
            return email
    return ""


def _reconcile_stale_import_job(conn, row):
    """Takılı pending/running işleri kapatır — ASLA yeniden başlatmaz.

    Eski davranış (restart) milyon kontak importunu her status poll'da
    tekrar başlatıp Postgres'i kilitliyordu; panel tamamen donuyordu.
    """
    job = _row(row)
    status = job.get("status")
    path = _import_job_path(job["id"], job.get("filename"))
    age_sec = _import_job_age_seconds(job.get("updated_at") or job.get("created_at"))

    if status not in ("pending", "running", "cancelling"):
        return job

    # Kısa ömürlü job'lara dokunma (aktif worker henüz yazıyor olabilir)
    grace = 90 if status == "running" else 45
    if age_sec < grace:
        return job

    err = (
        "İçe aktarma sunucu yeniden başlatıldığında veya zaman aşımında durduruldu. "
        "Paneli kilitlememek için otomatik yeniden başlatılmadı — dosyayı tekrar yükleyin."
    )
    now = iso(utcnow())
    execute(
        conn,
        "UPDATE mail_import_jobs SET status = 'error', error = ?, updated_at = ? "
        "WHERE id = ? AND status IN ('pending','running','cancelling')",
        (err, now, job["id"]),
    )
    conn.commit()
    try:
        if os.path.isfile(path):
            os.remove(path)
    except Exception:
        pass
    job["status"] = "error"
    job["error"] = err
    job["updated_at"] = now
    return job


def _cancel_all_active_imports(reason="Panel koruması: aktif içe aktarma durduruldu."):
    """Açık pending/running import'ları error'a çek — DB kilidini kes."""
    try:
        with closing(get_db()) as conn:
            rows = fetchall(
                conn,
                "SELECT id, filename FROM mail_import_jobs "
                "WHERE status IN ('pending','running','cancelling')",
            ) or []
            if not rows:
                return 0
            now = iso(utcnow())
            for row in rows:
                job = _row(row)
                execute(
                    conn,
                    "UPDATE mail_import_jobs SET status = 'error', error = ?, updated_at = ? WHERE id = ?",
                    (reason, now, job["id"]),
                )
                path = _import_job_path(job["id"], job.get("filename"))
                try:
                    if os.path.isfile(path):
                        os.remove(path)
                except Exception:
                    pass
            conn.commit()
            print(f"🛑 mail import cancel: {len(rows)} job durduruldu")
            return len(rows)
    except Exception as exc:
        print(f"⚠️  mail import cancel failed: {exc}")
        return 0


MODULE_ACCESS = ("module.mailing",)
MAIL_DASH = ("mailing.dashboard",)
MAIL_CRM = ("mailing.crm",)
MAIL_REL = ("mailing.relations", "mailing.crm")  # gerçek CRM; rehber yetkisi olan da görür
MAIL_TPL = ("mailing.templates",)
MAIL_CAMP = ("mailing.campaigns",)
MAIL_IVR = ("mailing.ivr",)
MAIL_REP = ("mailing.reports",)
MAIL_SET = ("mailing.settings",)

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", re.I)
_EMAIL_FIND_RE = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.I)


def _normalize_email_candidate(raw) -> str:
    """Hücre/satırdan e-posta ayıkla — isim <mail>, mailto:, tırnak, trailing nokta."""
    s = str(raw or "").replace("\ufeff", "").strip()
    if not s:
        return ""
    s = re.sub(r"^mailto:\s*", "", s, flags=re.I)
    angle = re.search(r"<\s*([^<>\s]+@[^<>\s]+)\s*>", s)
    if angle:
        s = angle.group(1)
    s = s.strip().strip("\"'[]")
    if re.search(r"[\s,;\t]", s) or ("@" in s and not EMAIL_RE.match(s.lower())):
        m = _EMAIL_FIND_RE.search(s)
        if m:
            s = m.group(0)
    s = s.lower().strip("<>").rstrip(".,;:")
    return s if EMAIL_RE.match(s) else ""
LINK_TOKEN_RE = re.compile(r"\{\{\s*link\s*:\s*([^}]+)\s*\}\}", re.I)
HREF_LINK_TOKEN_RE = re.compile(
    r"href\s*=\s*([\"'])\s*\{\{\s*link\s*:\s*([^}]+)\s*\}\}\s*\1",
    re.I,
)
HREF_RE = re.compile(r'(<a\b[^>]*\bhref\s*=\s*["\'])(https?://[^"\']+)(["\'])', re.I)


def _row(r):
    if not r:
        return None
    return dict(r)


def _rows(rs):
    return [dict(r) for r in (rs or [])]


def _parse_tags(raw):
    if raw is None:
        return []
    if isinstance(raw, list):
        return [str(t).strip() for t in raw if str(t).strip()]
    if isinstance(raw, str):
        raw = raw.strip()
        if not raw:
            return []
        try:
            data = json.loads(raw)
            if isinstance(data, list):
                return [str(t).strip() for t in data if str(t).strip()]
        except Exception:
            pass
        return [t.strip() for t in raw.split(",") if t.strip()]
    return []


def _tags_json(tags):
    return json.dumps(_parse_tags(tags), ensure_ascii=False)


def _like_literal(value: str) -> str:
    """LIKE jokerlerini (% _) ve escape karakterini nötralize et."""
    return (
        (value or "")
        .replace("\\", "\\\\")
        .replace("%", "\\%")
        .replace("_", "\\_")
    )


def _tag_match_clause(tag, column="tags"):
    """Etiket eşleşme SQL'i — JSON text içinde `\"etiket\"` (PG + SQLite).

    jsonb cast kullanılmaz: bozuk tags satırları tüm sorguyu abort eder.
    Etiket adındaki % / _ literal kalır (örn. «%100» kampanya etiketi).
    """
    tag = (tag or "").strip()
    if not tag:
        return "1=0", ()
    # Standart JSON dizi elemanı: ..."Etiket"...
    return f"{column} LIKE ? ESCAPE '\\'", (f'%"{_like_literal(tag)}"%',)


def _parse_tag_filter_list(tag_filter):
    """tag_filter string / virgüllü liste / tag_filters[] → benzersiz etiket listesi."""
    if isinstance(tag_filter, (list, tuple)):
        raw_parts = [str(x) for x in tag_filter]
    else:
        raw = (tag_filter or "").strip()
        if not raw:
            return []
        raw_parts = re.split(r"[,|;]+", raw)
    out = []
    seen = set()
    for part in raw_parts:
        t = (part or "").strip()
        if not t or t in seen:
            continue
        seen.add(t)
        out.append(t)
    return out


def _normalize_tag_filter_storage(tag_filter):
    """DB tag_filter kolonuna yazılacak biçim: 'a, b' veya ''."""
    tags = _parse_tag_filter_list(tag_filter)
    return ", ".join(tags)


def _tag_match_any_clause(tags, column="tags"):
    """Birden fazla etiket → OR (herhangi birinde olan kontak)."""
    tags = _parse_tag_filter_list(tags)
    if not tags:
        return "1=1", ()
    if len(tags) == 1:
        return _tag_match_clause(tags[0], column)
    parts = []
    params = []
    for t in tags:
        clause, tparams = _tag_match_clause(t, column)
        parts.append(f"({clause})")
        params.extend(tparams)
    return "(" + " OR ".join(parts) + ")", tuple(params)


def _contact_out(row):
    d = _row(row) if not isinstance(row, dict) else dict(row)
    if not d:
        return None
    d["tags"] = _parse_tags(d.get("tags"))
    d["unsubscribed"] = bool(d.get("unsubscribed"))
    return d


DEFAULT_GREETING_NAME = "Değerli üye"


def _mail_logo_url():
    """Şablonlardaki __MAIL_LOGO__ — güncel renkli site logosu (navy zeminli jpg)."""
    base = _public_base()
    path = "/static/mailing/makrobet-logo-mail.png?v=20260723c"
    return (base + path) if base else path


def _bizzo_logo_url():
    base = _public_base()
    path = "/static/mailing/bizzo-logo.png?v=20260723d"
    return (base + path) if base else path


def _spam_tip_banner_html():
    """Spam tip şeridi — soft amber (engine ile aynı dil)."""
    gold = "#ffcc00"
    return (
        '<table role="presentation" width="100%" cellpadding="0" cellspacing="0">'
        '<tr><td align="center" style="padding:0 12px 12px;">'
        '<table role="presentation" width="600" cellpadding="0" cellspacing="0" bgcolor="#1a1608" '
        'style="width:100%;max-width:600px;background-color:#1a1608;'
        'border:1px solid #5a4208;border-radius:12px;">'
        '<tr><td align="center" style="padding:12px 16px;'
        f"font-family:Arial,Helvetica,sans-serif;font-size:12px;line-height:1.5;color:{gold};"
        'background-color:#1a1608;">'
        "Spam klasöründeyse "
        f'<strong style="color:{gold};">butonlar çalışmaz</strong>. '
        f'Önce <strong style="color:{gold};">Spam değil</strong> deyin, sonra tıklayın.'
        "</td></tr></table></td></tr></table>"
    )


def _ensure_spam_tip(html):
    """HTML gövdede spam şeridi yoksa <body> hemen ardına ekler (yeni şablonlar dahil)."""
    html = html or ""
    if "Spam değil" in html or "Spam olmadığını bildir" in html:
        return html
    banner = _spam_tip_banner_html()
    low = html.lower()
    body_idx = low.find("<body")
    if body_idx >= 0:
        gt = html.find(">", body_idx)
        if gt >= 0:
            return html[: gt + 1] + banner + html[gt + 1 :]
    return banner + html


def _mail_promo_img_url(name):
    base = _public_base()
    path = f"/static/mailing/promos/{name}"
    return (base + path) if base else path


def _apply_mail_assets(text):
    text = text or ""
    if "__MAIL_LOGO__" in text:
        text = text.replace("__MAIL_LOGO__", _mail_logo_url())
    if "__BIZZO_LOGO__" in text:
        text = text.replace("__BIZZO_LOGO__", _bizzo_logo_url())
    promo_map = {
        "__MB_IMG_KASA__": "kasa.jpg",
        "__MB_IMG_KAYIP__": "kayip.jpg",
        "__MB_IMG_ARKADAS__": "arkadas.jpg",
        "__MB_IMG_RACE__": "race.jpg",
    }
    for token, fname in promo_map.items():
        if token in text:
            text = text.replace(token, _mail_promo_img_url(fname))
    # HTML e-postaysa spam tip şeridini garanti et
    if "<" in text and ("<html" in text.lower() or "<body" in text.lower() or "<table" in text.lower()):
        text = _ensure_spam_tip(text)
    return text


def _render_template(text, contact):
    """{{name}} boşsa 'Merhaba ,' gibi bozuk bir selamlama çıkmasın —
    isim yoksa nazik bir varsayılan ('Değerli üye') kullanılır."""
    text = _apply_mail_assets(text or "")
    name = ((contact or {}).get("name") or "").strip()
    mapping = {
        "name": name or DEFAULT_GREETING_NAME,
        "email": (contact or {}).get("email") or "",
        "phone": (contact or {}).get("phone") or "",
    }
    for key, val in mapping.items():
        text = text.replace("{{" + key + "}}", str(val))
    return text


def _plain_to_html(text):
    """Basit yazıyı basit HTML'e çevir — satır sonları + {{link:}} korunur."""
    text = (text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    holders = {}

    def hold(m):
        key = f"__MAILINK{len(holders)}__"
        holders[key] = m.group(0)
        return key

    protected = LINK_TOKEN_RE.sub(hold, text)
    parts = []
    for block in protected.split("\n\n"):
        esc = html_lib.escape(block).replace("\n", "<br>\n")
        for key, raw in holders.items():
            esc = esc.replace(html_lib.escape(key), raw).replace(key, raw)
        parts.append(f"<p>{esc}</p>")
    return "\n".join(parts)


def _public_base():
    """Click/open/unsub absolute base — worker’da request yoksa da kırılmasın.

    Mailing servisinde her zaman mikromail host; yanlış PUBLIC_BASE_URL
    (takipmkr vs) open pixel’i 302’ye düşürüp açılma sayımını öldürür.
    """
    mode = (os.environ.get("SERVICE_MODE") or "").strip().lower()
    mailing_default = "https://mikromail.onrender.com"
    base = (os.environ.get("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    # Panel host’una işaret eden PUBLIC_BASE mailing’de override
    if mode in ("mailing", "mailing-worker"):
        if not base or "takipmkr" in base.lower() or "bizzopanel" in base.lower():
            mm = (os.environ.get("MIKROMAIL_URL") or mailing_default).strip().rstrip("/")
            return mm or mailing_default
        return base
    if base:
        return base
    try:
        root = (request.url_root or "").rstrip("/")
        if root:
            return root
    except RuntimeError:
        pass
    if mode in ("mailing", "mailing-worker"):
        return mailing_default
    return ""


def _track_url(token):
    base = _public_base() or "https://mikromail.onrender.com"
    return f"{base}/m/c/{token}"


_SMARTICO_LINK_RE = re.compile(r"^\s*sc\s*:\s*(.+)$", re.I | re.S)


def _split_smartico_marker(dest):
    """'sc:https://...' işaretini ayıkla. Dönüş: (asıl_url, is_smartico)."""
    dest = (dest or "").strip()
    m = _SMARTICO_LINK_RE.match(dest)
    if m:
        return m.group(1).strip(), True
    return dest, False


def _click_serializer():
    """İmzalı tıklama tokenı — DB silinse bile dest_url yönlendirmesi çalışır.

    Env secret yoksa sabit "dev-mikromail" string'ine düşmez (tahmin edilebilir
    imza = link/contact_id sahteciliği); mail_ops_secret ile aynı kalıcı,
    rastgele anahtarı kullanır (bkz. mail_ops._ops_secret).
    """
    from itsdangerous import URLSafeSerializer

    secret = (
        os.environ.get("MAILING_SECRET_KEY")
        or os.environ.get("SECRET_KEY")
        or ""
    ).strip()
    if not secret:
        try:
            from mail_ops import _ops_secret
            secret = _ops_secret()
        except Exception as exc:
            print(f"⚠️  click serializer secret fallback: {exc}")
            secret = "dev-mikromail"
    return URLSafeSerializer(str(secret), salt="mail-click-v2")


def _normalize_click_token(raw: str) -> str:
    import urllib.parse

    t = urllib.parse.unquote((raw or "").strip())
    # Gmail / istemci sondaki noktalama ekleyebiliyor
    t = t.strip().rstrip(".,);]>\"'")
    if t.endswith("/"):
        t = t[:-1]
    return t


def _loads_signed_click(token: str):
    """İmzalı token → dict veya None."""
    if not token:
        return None
    try:
        data = _click_serializer().loads(token)
        if isinstance(data, dict) and (data.get("d") or "").strip():
            return data
    except Exception:
        return None
    return None


def _make_click_token(conn, *, dest_url, send_id=None, contact_id=None, campaign_id=None, is_smartico=False):
    """Takip tokenı üret.

    v2: dest imzalı token içinde (kontak/send silinse bile buton çalışır).
    DB kaydı analitik için best-effort.
    """
    dest_url = (dest_url or "").strip()
    payload = {
        "d": dest_url,
        "s": int(send_id) if send_id else None,
        "c": int(contact_id) if contact_id else None,
        "g": int(campaign_id) if campaign_id else None,
        "sc": 1 if is_smartico else 0,
    }
    try:
        token = _click_serializer().dumps(payload)
    except Exception:
        token = secrets.token_urlsafe(12)
    now = iso(utcnow())

    def _insert_click(with_smartico: bool):
        if with_smartico:
            insert_returning_id(
                conn,
                """
                INSERT INTO mail_click_links
                (token, send_id, contact_id, campaign_id, dest_url, is_smartico, click_count, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 0, ?)
                """,
                (
                    token,
                    send_id,
                    contact_id,
                    campaign_id,
                    dest_url,
                    1 if is_smartico else 0,
                    now,
                ),
            )
        else:
            insert_returning_id(
                conn,
                """
                INSERT INTO mail_click_links
                (token, send_id, contact_id, campaign_id, dest_url, click_count, created_at)
                VALUES (?, ?, ?, ?, ?, 0, ?)
                """,
                (token, send_id, contact_id, campaign_id, dest_url, now),
            )

    # SAVEPOINT: insert fail olsa mail_sends satırını rollback etme
    sp = "sp_mail_click"
    try:
        if uses_postgres():
            execute(conn, f"SAVEPOINT {sp}")
        _insert_click(True)
        if uses_postgres():
            execute(conn, f"RELEASE SAVEPOINT {sp}")
    except Exception as exc:
        print(f"⚠️  mail_click_links insert: {exc}")
        try:
            if uses_postgres():
                execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
            else:
                safe_rollback(conn)
        except Exception:
            safe_rollback(conn)
        # is_smartico kolonu yoksa / FK vs. — kolon olmadan dene
        try:
            if uses_postgres():
                execute(conn, f"SAVEPOINT {sp}")
            _insert_click(False)
            if uses_postgres():
                execute(conn, f"RELEASE SAVEPOINT {sp}")
        except Exception as exc2:
            print(f"⚠️  mail_click_links insert retry: {exc2}")
            try:
                if uses_postgres():
                    execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
                else:
                    safe_rollback(conn)
            except Exception:
                safe_rollback(conn)
    return token


def _append_query_param(url, key, value):
    """URL'e query param ekle/güncelle — mevcut parametreleri korur."""
    try:
        parts = urllib.parse.urlsplit(url)
        qs = urllib.parse.parse_qsl(parts.query, keep_blank_values=True)
        qs = [(k, v) for k, v in qs if k != key]
        qs.append((key, str(value)))
        new_query = urllib.parse.urlencode(qs)
        return urllib.parse.urlunsplit((parts.scheme, parts.netloc, parts.path, new_query, parts.fragment))
    except Exception:
        sep = "&" if "?" in url else "?"
        return f"{url}{sep}{key}={value}"


def _inject_tracking(conn, body, *, send_id, contact_id=None, campaign_id=None, as_html=True):
    """{{link:url}} / {{link:sc:url}} → imzalı /m/c/ takip URL.

    Tıklanınca mikromail 302 ile siteye gider; clicked_at + raporlar çalışır.
    'sc:' Smartico — redirect anında afp1 eklenir (mail_click).
    """
    body = body or ""
    if not body:
        return body

    base = _public_base() or "https://mikromail.onrender.com"

    def absolutize_track(url):
        u = (url or "").strip()
        if not u:
            return u
        if u.startswith("/m/c/"):
            return base + u
        if u.startswith("m/c/"):
            return base + "/" + u
        return u

    def token_for(raw_dest):
        dest, is_sc = _split_smartico_marker(raw_dest)
        dest = (dest or "").strip()
        if not dest:
            return dest
        if "/m/c/" in dest:
            return absolutize_track(dest)
        if not re.match(r"^https?://", dest, re.I):
            if dest.startswith("//"):
                dest = "https:" + dest
            elif not dest.startswith("/"):
                dest = "https://" + dest.lstrip("/")
        tok = _make_click_token(
            conn,
            dest_url=dest,
            send_id=send_id,
            contact_id=contact_id,
            campaign_id=campaign_id,
            is_smartico=is_sc,
        )
        return _track_url(tok)

    def repl_token(m):
        tracked = token_for(m.group(1))
        if as_html:
            label = html_lib.escape(
                _split_smartico_marker(m.group(1))[0].strip() or tracked, quote=True
            )
            return f'<a href="{html_lib.escape(tracked, quote=True)}" target="_blank" rel="noopener">{label}</a>'
        return tracked

    if as_html:
        def repl_href_link(m):
            q = m.group(1)
            tracked = token_for(m.group(2))
            return f"href={q}{tracked}{q}"

        body = HREF_LINK_TOKEN_RE.sub(repl_href_link, body)

    body = LINK_TOKEN_RE.sub(repl_token, body)

    if as_html:
        def repl_href(m):
            # Zaten takip / unsub / pixel değilse sarmala
            raw = m.group(2) or ""
            if "/m/c/" in raw or "/m/u/" in raw or "/m/o/" in raw:
                return m.group(0)
            tracked = token_for(raw)
            return f"{m.group(1)}{tracked}{m.group(3)}"

        body = HREF_RE.sub(repl_href, body)
        body = re.sub(
            r'href=(["\'])(/m/c/[^"\']+)\1',
            lambda m: f"href={m.group(1)}{base}{m.group(2)}{m.group(1)}",
            body,
            flags=re.I,
        )

    if "{{link:" in body or "{{ link:" in body:
        body = LINK_TOKEN_RE.sub(repl_token, body)
        body = re.sub(
            r"\{\{\s*link\s*:\s*([^}]+)\s*\}\}",
            lambda m: token_for(m.group(1)) or "",
            body,
            flags=re.I,
        )
    return body


def _tag_contact(conn, contact_id, tag, now=None):
    if not contact_id or not tag:
        return
    now = now or iso(utcnow())
    row = fetchone(conn, "SELECT tags FROM mail_contacts WHERE id = ?", (contact_id,))
    if not row:
        return
    tags = _parse_tags(row["tags"])
    if tag not in tags:
        tags.append(tag)
        execute(
            conn,
            "UPDATE mail_contacts SET tags = ?, updated_at = ? WHERE id = ?",
            (_tags_json(tags), now, contact_id),
        )
    exists = scalar(conn, "SELECT COUNT(*) FROM mail_contact_tags WHERE name = ?", (tag,))
    if not exists:
        insert_returning_id(
            conn,
            "INSERT INTO mail_contact_tags (name, created_at) VALUES (?, ?)",
            (tag, now),
        )


def _untag_contact(conn, contact_id, tag, now=None):
    if not contact_id or not tag:
        return
    now = now or iso(utcnow())
    row = fetchone(conn, "SELECT tags FROM mail_contacts WHERE id = ?", (contact_id,))
    if not row:
        return
    tags = _parse_tags(row["tags"])
    if tag in tags:
        tags.remove(tag)
        execute(
            conn,
            "UPDATE mail_contacts SET tags = ?, updated_at = ? WHERE id = ?",
            (_tags_json(tags), now, contact_id),
        )


def _bulk_retag_contacts(conn, *, action, from_tag="", to_tag="", contact_ids=None, match_tag="", limit=None, tenant_id=None):
    """Toplu etiket işlemleri — eşleşen ID'leri belleğe yüklemez, parça parça işler.

    action:
      - add: to_tag ekle
      - remove: from_tag kaldır
      - move: from_tag kaldır + to_tag ekle
    Kapsam: contact_ids listesi veya match_tag ile eşleşen kontaklar.

    tenant_id verilirse (tenant login veya superadmin impersonate) hem
    contact_ids hem match_tag taraması SADECE o tenant'ın kontaklarıyla
    sınırlanır — önceden bu filtre yoktu, bir firma başka firmanın
    kontaklarını toplu etiketleyip/segmentleyebiliyordu (veri sızıntısı).
    """
    action = (action or "").strip().lower()
    from_tag = (from_tag or "").strip()
    to_tag = (to_tag or "").strip()
    match_tag = (match_tag or "").strip()
    now = iso(utcnow())

    if action not in ("add", "remove", "move"):
        raise ValueError("Geçersiz işlem. add / remove / move kullanın.")
    if action == "add" and not to_tag:
        raise ValueError("Eklenecek etiket gerekli.")
    if action == "remove" and not from_tag:
        raise ValueError("Kaldırılacak etiket gerekli.")
    if action == "move" and (not from_tag or not to_tag):
        raise ValueError("Taşıma için kaynak ve hedef etiket gerekli.")
    if action == "move" and from_tag == to_tag:
        raise ValueError("Kaynak ve hedef etiket aynı olamaz.")

    if to_tag:
        _ensure_tag(conn, to_tag, now)
    if from_tag:
        _ensure_tag(conn, from_tag, now)

    ids = []
    if contact_ids:
        for raw in contact_ids:
            try:
                ids.append(int(raw))
            except (TypeError, ValueError):
                continue
        ids = list(dict.fromkeys(ids))
    elif not match_tag:
        raise ValueError("contact_ids veya match_tag gerekli.")

    updated = 0
    matched = 0
    batch_size = 800
    max_batches = 50000
    hard_limit = int(limit) if limit else None

    def _apply_row(row):
        nonlocal updated
        row = _row(row) if not isinstance(row, dict) else row
        tags = _parse_tags(row.get("tags"))
        before = list(tags)
        if action == "add":
            if to_tag not in tags:
                tags.append(to_tag)
        elif action == "remove":
            tags = [t for t in tags if t != from_tag]
        else:  # move
            tags = [t for t in tags if t != from_tag]
            if to_tag not in tags:
                tags.append(to_tag)
        if tags == before:
            return False
        execute(
            conn,
            "UPDATE mail_contacts SET tags = ?, updated_at = ? WHERE id = ?",
            (_tags_json(tags), now, int(row["id"])),
        )
        updated += 1
        return True

    tid_clause = " AND tenant_id = ?" if tenant_id else ""
    tid_param = (int(tenant_id),) if tenant_id else ()

    if ids:
        matched = 0
        for i in range(0, len(ids), batch_size):
            part = ids[i : i + batch_size]
            placeholders = ",".join(["?"] * len(part))
            rows = fetchall(
                conn,
                f"SELECT id, tags FROM mail_contacts WHERE id IN ({placeholders}){tid_clause}",
                tuple(part) + tid_param,
            ) or []
            matched += len(rows)
            for row in rows:
                _apply_row(row)
            try:
                conn.commit()
            except Exception:
                pass
    else:
        # match_tag: tüm id'leri çekme — etiket kalktıkça eşleşenler azalır
        clause, params = _tag_match_clause(match_tag)
        clause = clause + tid_clause
        params = tuple(params) + tid_param
        for _ in range(max_batches):
            take = batch_size
            if hard_limit is not None:
                remain = hard_limit - matched
                if remain <= 0:
                    break
                take = min(batch_size, remain)
            rows = fetchall(
                conn,
                f"SELECT id, tags FROM mail_contacts WHERE {clause} ORDER BY id ASC LIMIT ?",
                tuple(params) + (take,),
            )
            if not rows:
                break
            matched += len(rows)
            changed_any = False
            for row in rows:
                if _apply_row(row):
                    changed_any = True
            try:
                conn.commit()
            except Exception:
                pass
            if not changed_any:
                break
            # add: aynı satırlar tekrar gelir — tek tur yeterli
            if action == "add":
                break

    _invalidate_mail_stats_cache()
    refresh_names = []
    if from_tag:
        refresh_names.append(from_tag)
    if to_tag:
        refresh_names.append(to_tag)
    if match_tag and match_tag not in refresh_names:
        refresh_names.append(match_tag)

    cleaned = []
    # Taşıma/kaldırma sonrası kaynak + hedef sayıları canlı güncelle
    for name in refresh_names:
        try:
            n = _recount_tag(conn, name)
            if action in ("move", "remove") and name == from_tag and int(n or 0) <= 0:
                cleaned = _cleanup_empty_tags(conn, [from_tag])
        except Exception:
            pass
    try:
        conn.commit()
    except Exception:
        pass
    _invalidate_mail_stats_cache()

    return {
        "ok": True,
        "matched": matched,
        "updated": updated,
        "action": action,
        "from_tag": from_tag,
        "to_tag": to_tag,
        "cleaned_tags": cleaned,
    }


# «Önceden mail atılmışları hariç tut» — bu etiketler / önekler muaf (test QA).
# «Makro Test» — bu etiketteki kontaklar hariç tut kutusu açık olsa da HER
# kampanyada mutlaka gönderilir (kullanıcı isteği: domain rotasyon testi için
# kalıcı test havuzu, checkbox durumundan bağımsız).
_EXCLUDE_SENT_EXEMPT_EXACT = frozenset({
    "test", "qa", "sandbox", "deneme",
    "tolgatest", "tolga test", "tolga-test", "tolga_test",
    "makrotest", "makro test", "makro-test", "makro_test",
})
_EXCLUDE_SENT_EXEMPT_PREFIXES = (
    "test-", "test_", "test:", "test/", "test ",
    "qa-", "qa_", "qa:", "qa/", "qa ",
    "sandbox-", "sandbox_", "sandbox:", "sandbox/", "sandbox ",
    "deneme-", "deneme_", "deneme:", "deneme/", "deneme ",
    "tolgatest", "tolga test", "tolga-test", "tolga_test", "tolga-",
    "makrotest", "makro test", "makro-test", "makro_test", "makro-",
)
# İsim içinde geçen sabit parçalar (örn. «tolgatest2», «xx-tolgatest»)
_EXCLUDE_SENT_EXEMPT_CONTAINS = ("tolgatest", "makrotest")


def _load_exclude_sent_exempt_custom(conn=None):
    """Ayarlar: exclude_sent_exempt_tags = 'benim-test, lab' (virgüllü)."""
    if conn is None:
        return []
    try:
        raw = (get_mail_setting(conn, "exclude_sent_exempt_tags", "") or "").strip()
    except Exception:
        return []
    return _parse_tag_filter_list(raw)


def _tag_is_exclude_sent_exempt(tag, custom_exempt=None):
    """True → bu etiket üzerinden seçilenlere 'önceden gönderilmiş' filtresi uygulanmaz."""
    t = (tag or "").strip()
    if not t:
        return False
    tl = t.lower().replace(" ", "")  # «tolga test» / «TolgaTest»
    tl_raw = t.lower()
    if tl_raw in _EXCLUDE_SENT_EXEMPT_EXACT or tl in {x.replace(" ", "") for x in _EXCLUDE_SENT_EXEMPT_EXACT}:
        return True
    for c in (custom_exempt or []):
        c = (c or "").strip()
        if c and (t == c or tl_raw == c.lower() or tl == c.lower().replace(" ", "")):
            return True
    for prefix in _EXCLUDE_SENT_EXEMPT_PREFIXES:
        if tl_raw.startswith(prefix) or tl.startswith(prefix.replace(" ", "")):
            return True
    for needle in _EXCLUDE_SENT_EXEMPT_CONTAINS:
        if needle in tl_raw or needle in tl:
            return True
    return False


def _contact_has_exclude_exempt_tag_sql(custom_exempt=None):
    """Kontak tags alanında muaf etiket var mı? (manuel seçim / karışık kampanya).

    Case-insensitive: kontak "Makro Test", "makro test" veya "MAKRO TEST"
    etiketiyle kaydedilmiş olabilir — hepsi eşleşsin diye LOWER(tags) LIKE
    LOWER(...) kullanılır (düz _tag_match_clause case-sensitive)."""
    parts = []
    params = []
    seen = set()
    for name in sorted(_EXCLUDE_SENT_EXEMPT_EXACT):
        parts.append("LOWER(tags) LIKE ? ESCAPE '\\'")
        params.append(f'%"{_like_literal(name.lower())}"%')
        seen.add(name.lower())
    for name in (custom_exempt or []):
        name = (name or "").strip()
        if not name or name.lower() in seen:
            continue
        seen.add(name.lower())
        parts.append("LOWER(tags) LIKE ? ESCAPE '\\'")
        params.append(f'%"{_like_literal(name.lower())}"%')
    for prefix in _EXCLUDE_SENT_EXEMPT_PREFIXES:
        # JSON elemanı önek ile başlar: ..."test-foo"...
        parts.append("LOWER(tags) LIKE ? ESCAPE '\\'")
        params.append(f'%"{_like_literal(prefix.lower())}%')
    for needle in _EXCLUDE_SENT_EXEMPT_CONTAINS:
        # ..."…tolgatest…"… — JSON içinde geçsin yeter
        parts.append("LOWER(tags) LIKE ? ESCAPE '\\'")
        params.append(f"%{_like_literal(needle)}%")
    if not parts:
        return "1=0", ()
    return "(" + " OR ".join(parts) + ")", tuple(params)


def _in_flight_recipient_sql():
    """Zamanlanmış / kuyruktaki / gönderilmekte olan kampanya alıcıları."""
    return """
    NOT EXISTS (
      SELECT 1 FROM mail_campaign_recipients r
      INNER JOIN mail_campaigns c ON c.id = r.campaign_id
      WHERE r.contact_id = mail_contacts.id
        AND LOWER(COALESCE(c.status, '')) IN ('draft', 'scheduled', 'queued', 'sending', 'paused')
        AND LOWER(COALESCE(r.status, '')) IN ('pending', 'sending')
    )
    """


def _campaign_selection_where(
    tag_filter, exclude_previously_sent, only_verified=False, custom_exempt=None,
    tenant_id=None,
):
    """Kampanya alıcı seçiminde kullanılan WHERE + params — hem sayım
    önizlemesinde hem gerçek eklemede aynı filtre mantığı kullanılsın diye.

    tag_filter: tek etiket, 'a, b' veya liste — birden fazla ise OR (birleşim).
    test / qa / sandbox / deneme (+ önekler) ve Ayarlar exclude_sent_exempt_tags
    etiketleri «önceden gönderilmiş» filtresinden muaf.
    exclude_previously_sent ayrıca diğer kampanyalarda bekleyen (in-flight)
    alıcıları da dışlar — sıradaki kampanya aynı kişiyi kapmasın.
    """
    clauses = [
        "unsubscribed = 0",
        "NOT EXISTS (SELECT 1 FROM mail_suppressions s WHERE s.email = LOWER(mail_contacts.email))",
    ]
    params = []
    if tenant_id:
        clauses.append("tenant_id = ?")
        params.append(int(tenant_id))
    tags = _parse_tag_filter_list(tag_filter)
    if tags:
        clause, tparams = _tag_match_any_clause(tags)
        clauses.append(clause)
        params.extend(tparams)
    if exclude_previously_sent:
        in_flight = _in_flight_recipient_sql()
        sent_sql = "NOT EXISTS (SELECT 1 FROM mail_sends s WHERE s.contact_id = mail_contacts.id)"
        # Hem geçmiş gönderim hem başka kampanyada bekleyen alıcı
        pool_block = f"(({sent_sql}) AND ({in_flight}))"
        exempt_tags = [t for t in tags if _tag_is_exclude_sent_exempt(t, custom_exempt)]
        non_exempt_tags = [t for t in tags if not _tag_is_exclude_sent_exempt(t, custom_exempt)]
        if not tags:
            clauses.append(pool_block)
        elif not non_exempt_tags:
            # Sadece test/muaf etiketler — filtre yok
            pass
        elif not exempt_tags:
            clauses.append(pool_block)
        else:
            # Karışık: muaf etiketi olanlar geçer; diğerleri elenir
            ex_clause, ex_params = _tag_match_any_clause(exempt_tags)
            clauses.append(f"(({ex_clause}) OR {pool_block})")
            params.extend(ex_params)
    if only_verified:
        # SMTP ile valid + (SMTP kapalıyken) mx_ok kabul
        clauses.append("LOWER(COALESCE(verify_status, '')) IN ('valid', 'mx_ok')")
    return " AND ".join(clauses), params


def _count_tag_campaign_match(
    conn, tag, *, exclude_previously_sent=False, only_verified=False, custom_exempt=None,
    tenant_id=None,
):
    """Tek etiket için kampanya filtreleriyle eşleşen kontak sayısı.

    Returns: (count, approx)
    """
    tag = (tag or "").strip()
    if not tag:
        return 0, False
    if custom_exempt is None:
        custom_exempt = _load_exclude_sent_exempt_custom(conn)
    # Ek filtre yoksa registry hızlı; unsub/suppression farkı için approx=True
    if not exclude_previously_sent and not only_verified and not tenant_id:
        n = _registry_tag_count(conn, tag)
        if n is not None:
            return int(n), True
    where_sql, params = _campaign_selection_where(
        tag, exclude_previously_sent, only_verified=only_verified,
        custom_exempt=custom_exempt, tenant_id=tenant_id,
    )
    try:
        n = int(scalar(
            conn,
            f"SELECT COUNT(*) FROM mail_contacts WHERE {where_sql}",
            tuple(params),
        ) or 0)
        return n, False
    except Exception:
        n = _registry_tag_count(conn, tag)
        return int(n or 0), True


def _tag_breakdown_for_campaign(
    conn, tags, *, exclude_previously_sent=False, only_verified=False, limit=25,
    custom_exempt=None, tenant_id=None,
):
    """Seçilen etiketler için etiket başına sayı.

    Büyük listelerde filtreli COUNT etiketi başına çok yavaş olabilir —
    bu yüzden varsayılan: registry (yaklaşık). exact=True istenirse
    filtreli sayım yapılır.
    """
    tags = _parse_tag_filter_list(tags)[: max(0, int(limit or 25))]
    if custom_exempt is None:
        custom_exempt = _load_exclude_sent_exempt_custom(conn)
    out = []
    # Filtreliyse bile breakdown için önce registry (UI donmasın);
    # birleşim toplamı select-preview'da ayrı exact COUNT ile gelir.
    use_exact = bool(exclude_previously_sent or only_verified or tenant_id) and len(tags) <= 3
    for tag in tags:
        if use_exact:
            count, approx = _count_tag_campaign_match(
                conn, tag,
                exclude_previously_sent=exclude_previously_sent,
                only_verified=only_verified,
                custom_exempt=custom_exempt,
                tenant_id=tenant_id,
            )
        else:
            n = _registry_tag_count(conn, tag)
            if n is None:
                count, approx = _count_tag_campaign_match(
                    conn, tag,
                    exclude_previously_sent=False,
                    only_verified=False,
                    custom_exempt=custom_exempt,
                    tenant_id=tenant_id,
                )
            else:
                count, approx = int(n), True
        out.append({
            "tag": tag,
            "count": int(count or 0),
            "approx": bool(approx),
            "exclude_sent_exempt": _tag_is_exclude_sent_exempt(tag, custom_exempt),
        })
    return out


def _insert_campaign_recipient_ids(conn, campaign_id, contact_ids, now):
    """Alıcı ekle — contact_id tekil; duplicate yutulur (UNIQUE)."""
    from database import uses_postgres

    contact_ids = [int(x) for x in (contact_ids or []) if x is not None]
    contact_ids = list(dict.fromkeys(contact_ids))
    if not contact_ids:
        return 0
    attached = 0
    chunk_size = 2000
    for i in range(0, len(contact_ids), chunk_size):
        chunk = contact_ids[i:i + chunk_size]
        if not chunk:
            continue
        if uses_postgres():
            values_sql = ",".join(["(?, ?, 'pending', ?)"] * len(chunk))
            vparams = []
            for contact_id in chunk:
                vparams += [campaign_id, contact_id, now]
            try:
                cur = execute(
                    conn,
                    f"""
                    INSERT INTO mail_campaign_recipients (campaign_id, contact_id, status, created_at)
                    VALUES {values_sql}
                    ON CONFLICT (campaign_id, contact_id) DO NOTHING
                    """,
                    tuple(vparams),
                )
                try:
                    attached += int(cur.rowcount or 0)
                except Exception:
                    attached += len(chunk)
            except Exception as exc:
                print(f"⚠️  recipient insert chunk: {exc}")
                # Fallback tek tek
                for contact_id in chunk:
                    try:
                        execute(
                            conn,
                            """
                            INSERT INTO mail_campaign_recipients (campaign_id, contact_id, status, created_at)
                            VALUES (?, ?, 'pending', ?)
                            ON CONFLICT (campaign_id, contact_id) DO NOTHING
                            """,
                            (campaign_id, contact_id, now),
                        )
                        attached += 1
                    except Exception:
                        pass
        else:
            for contact_id in chunk:
                try:
                    execute(
                        conn,
                        """
                        INSERT OR IGNORE INTO mail_campaign_recipients
                        (campaign_id, contact_id, status, created_at)
                        VALUES (?, ?, 'pending', ?)
                        """,
                        (campaign_id, contact_id, now),
                    )
                    attached += 1
                except Exception:
                    pass
    # Gerçek eklenen sayı
    try:
        real = int(
            scalar(
                conn,
                "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ?",
                (campaign_id,),
            )
            or 0
        )
        return real
    except Exception:
        return attached


def _filter_sendable_contact_ids(
    conn, ids, *, exclude_previously_sent=False, only_verified=False, custom_exempt=None,
    tenant_id=None,
):
    """Verilen ID listesini unsub/suppression/(opsiyonel) verify filtrelerinden geçir."""
    from database import _table_columns

    ids = [int(x) for x in (ids or []) if str(x).isdigit() or isinstance(x, int)]
    ids = list(dict.fromkeys(ids))
    if not ids:
        return []
    if custom_exempt is None:
        custom_exempt = _load_exclude_sent_exempt_custom(conn)
    cols = _table_columns(conn, "mail_contacts") or set()
    has_verify = "verify_status" in cols
    out = []
    for i in range(0, len(ids), 500):
        chunk = ids[i:i + 500]
        ph = ",".join(["?"] * len(chunk))
        clauses = [
            f"id IN ({ph})",
            "unsubscribed = 0",
        ]
        params = list(chunk)
        if tenant_id:
            clauses.append("tenant_id = ?")
            params.append(int(tenant_id))
        # suppression tablosu yoksa transaction öldürmesin
        try:
            from database import uses_postgres
            if uses_postgres():
                exists = scalar(
                    conn,
                    """
                    SELECT COUNT(*) FROM information_schema.tables
                    WHERE table_name = 'mail_suppressions'
                    """,
                )
            else:
                exists = scalar(
                    conn,
                    "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='mail_suppressions'",
                )
            if exists:
                clauses.append(
                    "NOT EXISTS (SELECT 1 FROM mail_suppressions s WHERE s.email = LOWER(mail_contacts.email))"
                )
        except Exception:
            pass
        if exclude_previously_sent:
            ex_sql, ex_params = _contact_has_exclude_exempt_tag_sql(custom_exempt)
            in_flight = _in_flight_recipient_sql()
            clauses.append(
                f"((NOT EXISTS (SELECT 1 FROM mail_sends s WHERE s.contact_id = mail_contacts.id) "
                f"AND ({in_flight})) OR {ex_sql})"
            )
            params.extend(ex_params)
        if only_verified and has_verify:
            clauses.append("LOWER(COALESCE(verify_status, '')) IN ('valid', 'mx_ok')")
        rows = fetchall(
            conn,
            f"SELECT id FROM mail_contacts WHERE {' AND '.join(clauses)} ORDER BY id ASC",
            tuple(params),
        ) or []
        out.extend(int(r["id"]) for r in rows)
    return out


def _ensure_contacts_from_emails(conn, emails, now, tenant_id=None):
    """Elle girilen e-postaları kontak olarak upsert eder; id listesi döner.

    Postgres'te başarısız SQL transaction'ı öldürür — try/except ile
    ikinci sorgu çalışmaz. Bu yüzden kolonları önce kontrol ediyoruz.

    Tenant scoped: Firma A'nın elle girdiği e-posta Firma B'nin kontağıyla
    eşleşip Firma B'nin kaydına yazmasın (ya da tenant'sız/başka firma
    kontağı sessizce iliştirilmesin) — eşleşme + yeni kayıt tenant_id'ye göre.
    """
    from database import _table_columns

    cleaned = []
    seen = set()
    for raw in emails or []:
        em = (raw or "").strip().lower()
        if not em or em in seen:
            continue
        if not EMAIL_RE.match(em):
            continue
        seen.add(em)
        cleaned.append(em)
    if not cleaned:
        return []
    cols = _table_columns(conn, "mail_contacts") or set()
    has_verify = "verify_status" in cols
    has_tenant = "tenant_id" in cols
    tid = int(tenant_id) if tenant_id else (1 if has_tenant else None)
    ids = []
    for em in cleaned:
        if has_tenant:
            row = fetchone(
                conn,
                "SELECT id FROM mail_contacts WHERE LOWER(email) = ? AND tenant_id = ?",
                (em, tid),
            )
        else:
            row = fetchone(conn, "SELECT id FROM mail_contacts WHERE LOWER(email) = ?", (em,))
        if row:
            ids.append(int(row["id"]))
            if has_verify:
                execute(
                    conn,
                    """
                    UPDATE mail_contacts
                    SET verify_status = CASE
                        WHEN COALESCE(verify_status, '') IN ('valid', 'mx_ok') THEN verify_status
                        ELSE 'mx_ok'
                    END
                    WHERE id = ? AND unsubscribed = 0
                    """,
                    (int(row["id"]),),
                )
            continue
        if has_verify and has_tenant:
            insert_sql = """
                INSERT INTO mail_contacts
                (email, name, tags, source, unsubscribed, notes, verify_status, created_at, updated_at, tenant_id)
                VALUES (?, '', '[]', 'manual_campaign', 0, '', 'mx_ok', ?, ?, ?)
                """
            insert_params = (em, now, now, tid)
        elif has_verify:
            insert_sql = """
                INSERT INTO mail_contacts
                (email, name, tags, source, unsubscribed, notes, verify_status, created_at, updated_at)
                VALUES (?, '', '[]', 'manual_campaign', 0, '', 'mx_ok', ?, ?)
                """
            insert_params = (em, now, now)
        elif has_tenant:
            insert_sql = """
                INSERT INTO mail_contacts (email, name, tags, source, unsubscribed, notes, created_at, updated_at, tenant_id)
                VALUES (?, '', '[]', 'manual_campaign', 0, '', ?, ?, ?)
                """
            insert_params = (em, now, now, tid)
        else:
            insert_sql = """
                INSERT INTO mail_contacts (email, name, tags, source, unsubscribed, notes, created_at, updated_at)
                VALUES (?, '', '[]', 'manual_campaign', 0, '', ?, ?)
                """
            insert_params = (em, now, now)
        # SAVEPOINT: (tenant_id, email) tenant-scoped değil de eski global UNIQUE(email)
        # hâlâ devredeyse (şema migration'ı henüz uygulanmadıysa) INSERT çakışabilir —
        # bu durumda kampanya/işlem tamamen patlamasın, mevcut (başka firmanın) kaydını
        # bul ve onu kullan. Savepoint sayesinde diğer e-postaların transaction'ı bozulmaz.
        cid = None
        sp = f"sp_ecfe_{len(ids)}"
        try:
            execute(conn, f"SAVEPOINT {sp}")
            cid = insert_returning_id(conn, insert_sql, insert_params)
            execute(conn, f"RELEASE SAVEPOINT {sp}")
        except Exception as exc:
            print(f"⚠️  _ensure_contacts_from_emails insert çakıştı ({em}): {exc}")
            try:
                execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
            except Exception:
                pass
            fallback = fetchone(conn, "SELECT id FROM mail_contacts WHERE LOWER(email) = ? ORDER BY id ASC", (em,))
            cid = int(fallback["id"]) if fallback else None
        if cid:
            ids.append(int(cid))
    return ids


def _attach_campaign_recipients(
    conn, campaign_id, *, tag_filter, max_recipients, exclude_previously_sent, now,
    only_verified=False, contact_ids=None, emails=None, tenant_id=None,
):
    """Etiket / seçili ID / elle e-posta ile kampanya alıcılarını ekler.

    tag_filter + contact_ids birlikte gelirse birleşim (OR): tam etiketler ∪ manuel seçimler.

    max_recipients kesiminde sıra:
      1) elle / seçili ID
      2) test-muaf etiketler (Tolga Test, test, qa…) — tam kapsama
      3) diğer etiketler (ORDER BY id ASC)
    Böylece büyük Davet/mail_mx_ok havuzu test adreslerini ezmez.
    """
    contact_ids = list(contact_ids or [])
    emails = list(emails or [])

    if emails:
        contact_ids = _ensure_contacts_from_emails(conn, emails, now, tenant_id=tenant_id) + contact_ids
        # Elle girilen adresler bilinçli test — verified filtresi uygulama
        only_verified = False

    merged = []
    seen = set()

    def _push(ids):
        for cid in ids:
            try:
                cid = int(cid)
            except (TypeError, ValueError):
                continue
            if cid in seen:
                continue
            seen.add(cid)
            merged.append(cid)

    def _ids_for_tags(tag_list):
        if not tag_list:
            return []
        where_sql, params = _campaign_selection_where(
            tag_list, exclude_previously_sent, only_verified=only_verified,
            custom_exempt=custom_exempt, tenant_id=tenant_id,
        )
        sql = f"SELECT id FROM mail_contacts WHERE {where_sql} ORDER BY id ASC"
        return [r["id"] for r in fetchall(conn, sql, tuple(params))]

    custom_exempt = _load_exclude_sent_exempt_custom(conn)

    if contact_ids:
        _push(_filter_sendable_contact_ids(
            conn, contact_ids,
            exclude_previously_sent=exclude_previously_sent,
            only_verified=only_verified,
            custom_exempt=custom_exempt,
            tenant_id=tenant_id,
        ))

    tags = _parse_tag_filter_list(tag_filter)
    if tags:
        exempt_tags = [t for t in tags if _tag_is_exclude_sent_exempt(t, custom_exempt)]
        non_exempt_tags = [t for t in tags if not _tag_is_exclude_sent_exempt(t, custom_exempt)]
        # Önce muaf/test etiketleri — max_recipients bunları kesmesin
        if exempt_tags:
            _push(_ids_for_tags(exempt_tags))
        if non_exempt_tags:
            _push(_ids_for_tags(non_exempt_tags))
        elif not exempt_tags:
            _push(_ids_for_tags(tags))

    if not tags and not contact_ids and not emails:
        raise ValueError(
            "Alıcı seçilmedi — etiket, seçili kontak veya elle liste zorunlu "
            "(tüm rehber tek seferde eklenmez)."
        )

    if max_recipients:
        merged = merged[: int(max_recipients)]
    return _insert_campaign_recipient_ids(conn, campaign_id, merged, now)


def _merge_contact_tag_sql(existing_tags_json, tag):
    """Mevcut tags JSON'a etiket ekle (yoksa)."""
    tag = (tag or "").strip()
    if not tag:
        return existing_tags_json or "[]"
    try:
        parsed = json.loads(existing_tags_json or "[]")
        if not isinstance(parsed, list):
            parsed = []
    except Exception:
        parsed = []
    if tag not in parsed:
        parsed.append(tag)
    return json.dumps(parsed, ensure_ascii=False)


def _bulk_upsert_contacts_fallback(conn, batch, tag, now, tenant_id=None):
    """ON CONFLICT yoksa: yeni insert + mevcut update (tenant dahil)."""
    from database import _table_columns

    if not batch:
        return 0, 0, 0
    tag = (tag or "").strip()
    cols = _table_columns(conn, "mail_contacts") or set()
    has_tenant = "tenant_id" in cols
    tid = int(tenant_id) if tenant_id else (1 if has_tenant else None)

    emails = [email for email, _ in batch]
    placeholders = ",".join(["?"] * len(emails))
    existing_rows = fetchall(
        conn,
        f"SELECT id, email, name, tags FROM mail_contacts WHERE LOWER(email) IN ({placeholders})",
        tuple(e.lower() for e in emails),
    ) or []
    by_email = {str(r["email"]).lower(): dict(r) for r in existing_rows}

    inserted = 0
    updated = 0
    for email, name in batch:
        key = email.lower()
        row = by_email.get(key)
        if row:
            new_tags = _merge_contact_tag_sql(row.get("tags"), tag)
            new_name = (row.get("name") or "").strip() or (name or "")
            if has_tenant:
                execute(
                    conn,
                    """
                    UPDATE mail_contacts
                    SET name = ?, tags = ?, updated_at = ?,
                        tenant_id = COALESCE(tenant_id, ?)
                    WHERE id = ?
                    """,
                    (new_name, new_tags, now, tid, row["id"]),
                )
            else:
                execute(
                    conn,
                    "UPDATE mail_contacts SET name = ?, tags = ?, updated_at = ? WHERE id = ?",
                    (new_name, new_tags, now, row["id"]),
                )
            updated += 1
        else:
            tag_json = json.dumps([tag], ensure_ascii=False) if tag else "[]"
            if has_tenant:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_contacts
                    (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at, tenant_id)
                    VALUES (?, '', ?, ?, 'csv', 0, '', ?, ?, ?)
                    """,
                    (email, name or "", tag_json, now, now, tid),
                )
            else:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_contacts
                    (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at)
                    VALUES (?, '', ?, ?, 'csv', 0, '', ?, ?)
                    """,
                    (email, name or "", tag_json, now, now),
                )
            inserted += 1
            by_email[key] = {
                "id": cid,
                "email": email,
                "name": name or "",
                "tags": tag_json,
            }
    return inserted + updated, inserted, updated


def _bulk_upsert_contacts(conn, batch, tag, now, tenant_id=None):
    """batch: [(email, name), ...]. Tek SQL ifadesiyle toplu insert/upsert.
    Döner: (upserted, inserted, updated)

    Postgres'te UNIQUE(email) yoksa ON CONFLICT patlar → fallback'e düşer.
    """
    if not batch:
        return 0, 0, 0
    from database import _table_columns, ensure_mail_contacts_unique_email

    try:
        ensure_mail_contacts_unique_email(conn)
    except Exception:
        pass

    tag = (tag or "").strip()
    tag_json_single = json.dumps([tag], ensure_ascii=False) if tag else "[]"
    tag_like_pattern = f'%"{tag}"%'
    emails = [email for email, _ in batch]
    placeholders = ",".join(["?"] * len(emails))
    cols = _table_columns(conn, "mail_contacts") or set()
    has_tenant = "tenant_id" in cols
    tid = int(tenant_id) if tenant_id else None
    if has_tenant and not tid:
        tid = 1
    # Mevcut/yeni sayımı da tenant'a göre yapılmalı — email artık (tenant_id, email)
    # bazında unique; başka firmanın kaydı burada "mevcut" sayılıp yanlış rapor vermesin.
    if has_tenant:
        existing_rows = fetchall(
            conn,
            f"SELECT email FROM mail_contacts WHERE LOWER(email) IN ({placeholders}) AND tenant_id = ?",
            tuple(e.lower() for e in emails) + (tid,),
        )
    else:
        existing_rows = fetchall(
            conn,
            f"SELECT email FROM mail_contacts WHERE LOWER(email) IN ({placeholders})",
            tuple(e.lower() for e in emails),
        )
    existing_set = {str(r["email"]).lower() for r in (existing_rows or [])}
    inserted = sum(1 for email, _ in batch if email.lower() not in existing_set)
    updated = len(batch) - inserted
    values_sql = []
    params = []
    if has_tenant:
        for email, name in batch:
            values_sql.append("(?, ?, ?, ?, 0, '', ?, ?, ?)")
            params += [email, name, tag_json_single, "csv", now, now, tid]
        sql = f"""
            INSERT INTO mail_contacts
            (email, name, tags, source, unsubscribed, notes, created_at, updated_at, tenant_id)
            VALUES {",".join(values_sql)}
            ON CONFLICT (tenant_id, email) DO UPDATE SET
                name = CASE WHEN mail_contacts.name = '' THEN EXCLUDED.name ELSE mail_contacts.name END,
                tags = CASE
                    WHEN ? = '' THEN mail_contacts.tags
                    WHEN mail_contacts.tags LIKE ? THEN mail_contacts.tags
                    WHEN mail_contacts.tags = '[]' THEN ?
                    ELSE substr(mail_contacts.tags, 1, length(mail_contacts.tags) - 1) || ',"' || ? || '"]'
                END,
                updated_at = EXCLUDED.updated_at
        """
    else:
        for email, name in batch:
            values_sql.append("(?, ?, ?, ?, 0, '', ?, ?)")
            params += [email, name, tag_json_single, "csv", now, now]
        sql = f"""
            INSERT INTO mail_contacts (email, name, tags, source, unsubscribed, notes, created_at, updated_at)
            VALUES {",".join(values_sql)}
            ON CONFLICT (email) DO UPDATE SET
                name = CASE WHEN mail_contacts.name = '' THEN EXCLUDED.name ELSE mail_contacts.name END,
                tags = CASE
                    WHEN ? = '' THEN mail_contacts.tags
                    WHEN mail_contacts.tags LIKE ? THEN mail_contacts.tags
                    WHEN mail_contacts.tags = '[]' THEN ?
                    ELSE substr(mail_contacts.tags, 1, length(mail_contacts.tags) - 1) || ',"' || ? || '"]'
                END,
                updated_at = EXCLUDED.updated_at
        """
    params += [tag, tag_like_pattern, tag_json_single, tag]
    try:
        cur = execute(conn, sql, tuple(params))
        upserted = cur.rowcount if cur.rowcount and cur.rowcount > 0 else len(batch)
        return upserted, inserted, updated
    except Exception as exc:
        msg = str(exc).lower()
        if "on conflict" not in msg and "unique or exclusion" not in msg:
            raise
        try:
            conn.rollback()
        except Exception:
            pass
        print(f"⚠️  bulk upsert ON CONFLICT yok — fallback: {exc}")
        return _bulk_upsert_contacts_fallback(conn, batch, tag, now, tenant_id=tenant_id)


def _detect_csv_delimiter(path):
    """Türkiye'de Excel'in varsayılan CSV export'u virgül yerine noktalı virgül
    (;) kullanır (virgül ondalık ayracı olduğu için). Header/örnek satırlara
    bakıp doğru ayracı otomatik seçiyoruz — aksi halde tüm satır tek bir
    sütun gibi okunur ve email/name sütunları hiç bulunamaz."""
    try:
        with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
            sample = f.read(65536)
    except OSError:
        return ","
    if not sample.strip():
        return ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        return dialect.delimiter
    except csv.Error:
        pass
    first_line = next((ln for ln in sample.splitlines() if ln.strip()), "")
    counts = {d: first_line.count(d) for d in (",", ";", "\t")}
    best = max(counts, key=counts.get)
    return best if counts[best] > 0 else ","


def _count_csv_rows(path):
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        return max(sum(1 for _ in f) - 1, 0)


def _iter_csv_rows(path):
    delimiter = _detect_csv_delimiter(path)
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            yield row


def _count_xlsx_rows(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    total = 0
    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            first_row = next(rows_iter, None)
            if first_row is None:
                continue
            header = [(str(h).strip() if h is not None else "") for h in first_row]
            if _find_email_column_index(header) is not None:
                total += sum(1 for _ in rows_iter)
                continue
            sample = [first_row]
            for _ in range(4):
                try:
                    sample.append(next(rows_iter))
                except StopIteration:
                    break
            if _values_look_like_email_column(sample):
                total += len(sample)
                total += sum(1 for _ in rows_iter)
    finally:
        wb.close()
    return total


def _iter_xlsx_sheet_rows(ws):
    """Tek bir worksheet'ten satır dict'leri veya başlıksız e-posta listesi üretir."""
    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    if first_row is None:
        return
    header = [(str(h).strip() if h is not None else "") for h in first_row]
    if _find_email_column_index(header) is not None:
        for values in rows_iter:
            row = {}
            for i, key in enumerate(header):
                if not key:
                    continue
                val = values[i] if i < len(values) else None
                row[key] = "" if val is None else str(val).strip()
            if any(str(v).strip() for v in row.values()):
                yield row
        return
    sample = [first_row]
    for _ in range(4):
        try:
            sample.append(next(rows_iter))
        except StopIteration:
            break
    if not _values_look_like_email_column(sample):
        return
    for values in sample:
        val = values[0] if values else None
        if val is None:
            continue
        email = str(val).strip()
        if EMAIL_RE.match(email):
            yield {"email": email}
    for values in rows_iter:
        val = values[0] if values else None
        if val is None:
            continue
        email = str(val).strip()
        if EMAIL_RE.match(email):
            yield {"email": email}


def _iter_xlsx_rows(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            yield from _iter_xlsx_sheet_rows(ws)
    finally:
        wb.close()


def _run_import_job(job_id, path, tag, tenant_id=None):
    """Arka plan thread: dosyayı satır satır okuyup IMPORT_CHUNK_SIZE'lık
    gruplar halinde bulk upsert eder — HTTP isteğinden bağımsız çalışır,
    timeout'a düşmez. CSV ve XLSX (.xlsx) destekler.

    tenant_id request context'ten thread'e taşınmalı (session thread'de yok).
    """
    now = iso(utcnow())
    is_xlsx = os.path.splitext(path)[1].lower() in (".xlsx", ".xlsm")
    iter_fn = _iter_xlsx_rows if is_xlsx else _iter_csv_rows
    last_batch_err = ""
    try:
        with closing(get_db()) as conn:
            existing = fetchone(conn, "SELECT status FROM mail_import_jobs WHERE id = ?", (job_id,))
            if existing and existing["status"] == "cancelling":
                execute(conn, "UPDATE mail_import_jobs SET status = 'cancelled', updated_at = ? WHERE id = ?", (now, job_id))
                conn.commit()
                return
            execute(conn, "UPDATE mail_import_jobs SET status = 'running', updated_at = ? WHERE id = ?", (now, job_id))
            conn.commit()

            # 20M+ satırlı dosyalarda önce tüm dosyayı saymak (count_fn) dakikalarca
            # sürüp updated_at'i donduruyordu; panel "kayboldu" sanıyordu. Satır sayımını
            # atlayıp doğrudan işlemeye başlıyoruz — total_rows iş bitince set edilir.
            processed = 0
            upserted = 0
            inserted = 0
            updated = 0
            skipped = 0
            batch = []
            cancelled = False
            # Manuel next() döngüsü kullanıyoruz ki satırı ÜRETİRKEN (örn. bozuk
            # encoding, tutarsız sütun sayısı) bir hata çıksa bile o satırı
            # geçersiz sayıp devam edebilelim — tek bozuk satır tüm job'ı
            # 'error' durumuna düşürmesin, milyonlarca satırın kalanı işlensin.
            row_iter = iter_fn(path)
            while True:
                try:
                    row = next(row_iter)
                except StopIteration:
                    break
                except Exception:
                    processed += 1
                    skipped += 1
                    continue
                processed += 1
                try:
                    email = _extract_email_from_row(row)
                    if not email:
                        skipped += 1
                        continue
                    name = (row.get("name") or row.get("Name") or "").strip()
                    batch.append((email, name))
                except Exception:
                    skipped += 1
                    continue
                if len(batch) >= IMPORT_CHUNK_SIZE:
                    try:
                        batch_upserted, batch_inserted, batch_updated = _bulk_upsert_contacts(
                            conn, batch, tag, iso(utcnow()), tenant_id=tenant_id
                        )
                        upserted += batch_upserted
                        inserted += batch_inserted
                        updated += batch_updated
                        conn.commit()
                    except Exception as batch_exc:
                        conn.rollback()
                        skipped += len(batch)
                        last_batch_err = str(batch_exc)[:300]
                        print(f"⚠️  mail import batch fail job={job_id}: {batch_exc}")
                    batch = []
                    execute(
                        conn,
                        """
                        UPDATE mail_import_jobs
                        SET processed_rows = ?, upserted_count = ?, inserted_count = ?, updated_count = ?,
                            skipped_count = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (processed, upserted, inserted, updated, skipped, iso(utcnow()), job_id),
                    )
                    conn.commit()
                    status_row = fetchone(conn, "SELECT status FROM mail_import_jobs WHERE id = ?", (job_id,))
                    if status_row and status_row["status"] == "cancelling":
                        cancelled = True
                        break
                elif processed % 5000 == 0:
                    execute(
                        conn,
                        """
                        UPDATE mail_import_jobs
                        SET processed_rows = ?, upserted_count = ?, inserted_count = ?, updated_count = ?,
                            skipped_count = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (processed, upserted, inserted, updated, skipped, iso(utcnow()), job_id),
                    )
                    conn.commit()
            if batch and not cancelled:
                try:
                    batch_upserted, batch_inserted, batch_updated = _bulk_upsert_contacts(
                        conn, batch, tag, iso(utcnow()), tenant_id=tenant_id
                    )
                    upserted += batch_upserted
                    inserted += batch_inserted
                    updated += batch_updated
                    conn.commit()
                except Exception as batch_exc:
                    conn.rollback()
                    skipped += len(batch)
                    last_batch_err = str(batch_exc)[:300]
                    print(f"⚠️  mail import batch fail job={job_id}: {batch_exc}")

            final_now = iso(utcnow())
            if cancelled:
                execute(
                    conn,
                    """
                    UPDATE mail_import_jobs
                    SET status = 'cancelled', total_rows = ?, processed_rows = ?, upserted_count = ?,
                        inserted_count = ?, updated_count = ?, skipped_count = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (processed, processed, upserted, inserted, updated, skipped, final_now, job_id),
                )
                conn.commit()
                return
            if tag:
                _ensure_tag(conn, tag, final_now)
            done_err = ""
            if upserted == 0 and skipped > 0 and last_batch_err:
                done_err = f"Kayıt yazılamadı: {last_batch_err}"
            elif upserted == 0 and skipped > 0 and processed > 0:
                done_err = (
                    "Hiç geçerli e-posta yazılamadı. Dosyada e-posta sütunu / satır formatını kontrol et "
                    "(başlık: email). Aktif tenant ile aynı firmaya yazılır."
                )
            execute(
                conn,
                """
                UPDATE mail_import_jobs
                SET status = 'done', total_rows = ?, processed_rows = ?, upserted_count = ?,
                    inserted_count = ?, updated_count = ?, skipped_count = ?, error = ?, updated_at = ?
                WHERE id = ?
                """,
                (processed, processed, upserted, inserted, updated, skipped, done_err, final_now, job_id),
            )
            conn.commit()
            _invalidate_mail_stats_cache()
            if tag:
                try:
                    _recount_tag(conn, tag)
                    conn.commit()
                except Exception:
                    pass
            if uses_postgres():
                try:
                    # reltuples / n_live_tup güncellensin — dashboard kartları dolsun
                    execute(conn, "ANALYZE mail_contacts")
                    conn.commit()
                except Exception:
                    pass
    except Exception as exc:
        try:
            with closing(get_db()) as conn:
                execute(
                    conn,
                    "UPDATE mail_import_jobs SET status = 'error', error = ?, updated_at = ? WHERE id = ?",
                    (str(exc)[:500], iso(utcnow()), job_id),
                )
                conn.commit()
        except Exception:
            pass
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


def _ensure_tag(conn, name, now=None):
    name = (name or "").strip()
    if not name:
        return False
    if now is None:
        now = iso(utcnow())
    exists = scalar(conn, "SELECT COUNT(*) FROM mail_contact_tags WHERE name = ?", (name,))
    if not exists:
        insert_returning_id(
            conn,
            "INSERT INTO mail_contact_tags (name, created_at) VALUES (?, ?)",
            (name, now),
        )
        return True
    return False


def _tag_usage_count(conn, name, *, live=True, tenant_id=None):
    """Etiket kullanım sayısı. live=True: gerçek COUNT (LIKE); False: yalnız registry.

    tenant_id verilirse SADECE o tenant'ın kontakları sayılır (registry global
    olduğu için tenant_id verildiğinde live=False cache'i kullanılamaz).
    """
    name = (name or "").strip()
    if not name:
        return 0
    if not live and not tenant_id:
        cached = _registry_tag_count(conn, name)
        if cached is not None:
            return int(cached)
        return 0
    clause, params = _tag_match_clause(name)
    if tenant_id:
        clause += " AND tenant_id = ?"
        params = tuple(params) + (int(tenant_id),)
    try:
        return int(scalar(conn, f"SELECT COUNT(*) FROM mail_contacts WHERE {clause}", params) or 0)
    except Exception:
        return 0


def _delete_tag(conn, name, *, force=False):
    """Etiketi registry'den sil. Kontak varsa force=True ile önce kontaktan kaldırır."""
    name = (name or "").strip()
    if not name:
        raise ValueError("Etiket adı gerekli.")
    usage = _tag_usage_count(conn, name)
    stripped = 0
    if usage > 0:
        if not force:
            raise ValueError(
                f"«{name}» etiketinde {usage} kontak var. "
                "Önce taşı/kaldır veya zorla sil (kontaklardan da silinir)."
            )
        result = _bulk_retag_contacts(
            conn, action="remove", from_tag=name, match_tag=name
        )
        stripped = int(result.get("updated") or 0)
    execute(conn, "DELETE FROM mail_contact_tags WHERE name = ?", (name,))
    return {"deleted": name, "stripped": stripped, "had_contacts": usage}


def _cleanup_empty_tags(conn, names=None):
    """0 kontak kalan etiketleri registry'den sil — çöp birikmesin."""
    if names is None:
        rows = fetchall(conn, "SELECT name FROM mail_contact_tags")
        names = [(r["name"] or "").strip() for r in rows]
    deleted = []
    for name in names:
        name = (name or "").strip()
        if not name:
            continue
        # Önce registry contact_count; 0 ise doğrula, >0 ise canlı say
        cached = None
        try:
            row = fetchone(conn, "SELECT contact_count FROM mail_contact_tags WHERE name = ?", (name,))
            if row is not None and "contact_count" in (row.keys() if hasattr(row, "keys") else []):
                cached = int(row["contact_count"] or 0)
        except Exception:
            cached = None
        usage = cached if cached is not None else _tag_usage_count(conn, name)
        if usage == 0:
            # Güvenlik: registry 0 diyorsa bir kez canlı doğrula
            if cached == 0:
                usage = _tag_usage_count(conn, name)
            if usage == 0:
                execute(conn, "DELETE FROM mail_contact_tags WHERE name = ?", (name,))
                deleted.append(name)
    if deleted:
        try:
            conn.commit()
        except Exception:
            pass
    return deleted


_STATS_CACHE = {"ts": 0.0, "payload": None}
_STATS_LOCK = threading.Lock()
_TAG_COUNT_CACHE = {"ts": 0.0, "rows": None}
_TAG_SYNC_STATE = {"ts": 0.0, "running": False, "last_added": 0}
# Tenant-scoped varyantlar — global cache'ler tüm platform içindir, tenant login
# olduğunda / superadmin impersonate ettiğinde bunlar kullanılır (bkz. veri
# sızıntısı fix'i: contacts/stats artık tenant_id'ye göre ayrı cache'lenir).
_TENANT_STATS_CACHE = {}
_TENANT_TAG_COUNT_CACHE = {}


def _approx_contact_total(conn, tenant_id=None):
    """Kontak toplamı — önce canlı istatistik, 0 ise exact COUNT (import sonrası kartlar boş kalmasın).

    tenant_id verilirse yaklaşık pg istatistiği (tüm tablo) atlanır — SADECE o
    tenant'ın satırları exact COUNT ile sayılır (önceden tenant_id hiç
    kullanılmıyordu, her firma platform genelindeki kontak sayısını görüyordu).
    """
    if tenant_id:
        try:
            return int(scalar(
                conn, "SELECT COUNT(*) FROM mail_contacts WHERE tenant_id = ?", (int(tenant_id),)
            ) or 0), False
        except Exception:
            return 0, False
    if uses_postgres():
        for sql in (
            "SELECT n_live_tup::bigint FROM pg_stat_user_tables WHERE relname = 'mail_contacts'",
            "SELECT reltuples::bigint FROM pg_class WHERE relname = 'mail_contacts'",
        ):
            try:
                n = scalar(conn, sql)
                if n is not None and int(n) > 0:
                    return int(n), True
            except Exception:
                pass
        # Import sonrası reltuples sık 0 kalır — COUNT şart
        try:
            return int(scalar(conn, "SELECT COUNT(*) FROM mail_contacts") or 0), False
        except Exception:
            return 0, True
    try:
        return int(scalar(conn, "SELECT COUNT(*) FROM mail_contacts") or 0), False
    except Exception:
        return 0, False


def _approx_mailed_contacts(conn, total, tenant_id=None):
    """En az 1 mail gitmiş kontak sayısı — timeout ile; olmazsa None.

    tenant_id verilirse SADECE o tenant'ın gönderimleri sayılır.
    """
    tid_clause = " AND tenant_id = ?" if tenant_id else ""
    tid_params = (int(tenant_id),) if tenant_id else ()
    try:
        if uses_postgres():
            try:
                execute(conn, "SET LOCAL statement_timeout = '4000ms'")
            except Exception:
                pass
            n = scalar(
                conn,
                f"SELECT COUNT(DISTINCT contact_id) FROM mail_sends WHERE contact_id IS NOT NULL{tid_clause}",
                tid_params,
            )
            return int(n or 0)
        return int(scalar(
            conn,
            f"SELECT COUNT(DISTINCT contact_id) FROM mail_sends WHERE contact_id IS NOT NULL{tid_clause}",
            tid_params,
        ) or 0)
    except Exception:
        return None
    finally:
        if uses_postgres():
            try:
                execute(conn, "SET LOCAL statement_timeout = 0")
            except Exception:
                pass


def _registry_tag_count(conn, name):
    name = (name or "").strip()
    if not name:
        return None
    try:
        row = fetchone(
            conn,
            "SELECT contact_count FROM mail_contact_tags WHERE name = ?",
            (name,),
        )
        if row is None:
            return None
        if hasattr(row, "keys") and "contact_count" in row.keys():
            return int(row["contact_count"] or 0)
    except Exception:
        pass
    return None


def _harvest_tags_into_registry(conn, tag_names, now=None):
    """Görünen/keşfedilen etiketleri registry'ye ekle (ucuz)."""
    if now is None:
        now = iso(utcnow())
    added = 0
    for name in tag_names or []:
        name = (name or "").strip()
        if not name:
            continue
        if _ensure_tag(conn, name, now):
            added += 1
    return added


def _sync_missing_tags_from_contacts(conn, *, max_rows=250000, batch_size=2000):
    """Kontak tags JSON'unda olup registry'de olmayan etiketleri bul/ekle.

    Tam tablo taraması pahalı — en yeni id'lerden geriye batch; max_rows ile sınırlı.
    Dönüş: eklenen etiket sayısı.
    """
    now = iso(utcnow())
    existing = {
        (r["name"] or "").strip()
        for r in (fetchall(conn, "SELECT name FROM mail_contact_tags") or [])
        if (r["name"] or "").strip()
    }
    discovered = set()
    cursor_id = int(scalar(conn, "SELECT COALESCE(MAX(id), 0) FROM mail_contacts") or 0) + 1
    scanned = 0
    while scanned < max_rows and cursor_id > 1:
        take = min(batch_size, max_rows - scanned)
        rows = fetchall(
            conn,
            "SELECT id, tags FROM mail_contacts WHERE id < ? ORDER BY id DESC LIMIT ?",
            (cursor_id, take),
        )
        if not rows:
            break
        for row in rows:
            row = _row(row)
            cursor_id = min(cursor_id, int(row["id"]))
            for t in _parse_tags(row.get("tags")):
                if t and t not in existing:
                    discovered.add(t)
        scanned += len(rows)
        if len(rows) < take:
            break
    added = 0
    for name in sorted(discovered):
        if _ensure_tag(conn, name, now):
            added += 1
            existing.add(name)
    if added:
        try:
            conn.commit()
        except Exception:
            pass
        _invalidate_mail_stats_cache()
    return {"scanned": scanned, "added": added, "discovered": len(discovered)}


def _maybe_sync_missing_tags_async(*, force=False, interval_sec=21600):
    """Devre dışı — milyon satır taraması paneli kilitliyordu."""
    return


def _set_registry_tag_count(conn, name, n):
    """Registry contact_count güncelle (COUNT atmadan)."""
    name = (name or "").strip()
    if not name:
        return
    try:
        from database import _table_columns
        cols = _table_columns(conn, "mail_contact_tags") or set()
    except Exception:
        cols = set()
    if cols and "contact_count" in cols:
        execute(
            conn,
            "UPDATE mail_contact_tags SET contact_count = ? WHERE name = ?",
            (int(n or 0), name),
        )


def _delete_contacts_by_ids(conn, ids):
    """Kontakları FK bağımlılıklarıyla sil (kampanya / send / click / ivr).

    Postgres'te ON DELETE CASCADE yok; düz DELETE IntegrityError verir.

    ÖNEMLİ: mail_click_links SATIRLARI SİLİNMEZ — sadece FK nullenir.
    Aksi halde eski maillerdeki /m/c/<token> butonları 'link bulunamadı' olur.
    """
    ids = [int(x) for x in (ids or []) if x is not None]
    ids = list(dict.fromkeys(ids))
    if not ids:
        return 0
    ph = ",".join(["?"] * len(ids))
    params = tuple(ids)

    # Tıklama linklerini koru (dest_url + token kalsın) — sadece FK kopar
    try:
        execute(
            conn,
            f"""
            UPDATE mail_click_links
            SET contact_id = NULL,
                send_id = NULL
            WHERE contact_id IN ({ph})
               OR send_id IN (SELECT id FROM mail_sends WHERE contact_id IN ({ph}))
            """,
            params + params,
        )
    except Exception:
        try:
            execute(
                conn,
                f"UPDATE mail_click_links SET contact_id = NULL WHERE contact_id IN ({ph})",
                params,
            )
        except Exception:
            pass

    try:
        execute(
            conn,
            f"DELETE FROM mail_ivr_events WHERE send_id IN (SELECT id FROM mail_sends WHERE contact_id IN ({ph}))",
            params,
        )
    except Exception:
        pass

    for table in (
        "mail_ivr_events",
        "mail_campaign_recipients",
        "mail_sends",
    ):
        try:
            execute(conn, f"DELETE FROM {table} WHERE contact_id IN ({ph})", params)
        except Exception:
            pass

    execute(conn, f"DELETE FROM mail_contacts WHERE id IN ({ph})", params)
    return len(ids)


def _recount_tag(conn, name):
    """Tek etiketin kontak sayısını DB'ye yazar (hızlı CRM için)."""
    name = (name or "").strip()
    if not name:
        return 0
    n = _tag_usage_count(conn, name)
    _set_registry_tag_count(conn, name, n)
    return n


def _recount_all_tags(conn, *, limit=300):
    """Registry'deki etiketleri canlı COUNT ile günceller."""
    rows = fetchall(
        conn,
        "SELECT name FROM mail_contact_tags ORDER BY name ASC LIMIT ?",
        (max(1, min(int(limit or 300), 500)),),
    ) or []
    out = []
    for r in rows:
        name = (r["name"] or "").strip()
        if not name:
            continue
        n = _recount_tag(conn, name)
        out.append({"name": name, "count": int(n or 0)})
    try:
        conn.commit()
    except Exception:
        pass
    return out


def _contact_tag_counts(conn, *, force=False, live=False, tenant_id=None):
    """Etiket sayıları — registry; live=True ise önce recount.

    tenant_id verilirse registry (global, tüm tenant'lar paylaşır) kullanılmaz —
    her etiket için SADECE o tenant'ın kontakları canlı COUNT ile sayılır ve
    ayrı bir tenant-scoped cache'te tutulur (önceden tenant filtresi yoktu,
    bir firma tüm platformun etiket dağılımını görebiliyordu).
    """
    import time

    if tenant_id:
        tkey = int(tenant_id)
        now = time.time()
        cached = _TENANT_TAG_COUNT_CACHE.get(tkey)
        if not force and not live and cached and (now - cached["ts"]) < 180:
            return cached["rows"]
        names = [
            (r["name"] or "").strip()
            for r in (fetchall(conn, "SELECT name FROM mail_contact_tags ORDER BY name ASC") or [])
            if (r["name"] or "").strip()
        ]
        rows = []
        for name in names:
            n = _tag_usage_count(conn, name, live=True, tenant_id=tkey)
            if n:
                rows.append({"name": name, "count": int(n)})
        rows.sort(key=lambda item: (-item["count"], item["name"].lower()))
        _TENANT_TAG_COUNT_CACHE[tkey] = {"ts": now, "rows": rows}
        return rows

    if live:
        _recount_all_tags(conn)
        force = True

    now = time.time()
    if not force and _TAG_COUNT_CACHE["rows"] is not None and (now - _TAG_COUNT_CACHE["ts"]) < 180:
        return _TAG_COUNT_CACHE["rows"]

    registry_rows = fetchall(conn, "SELECT * FROM mail_contact_tags ORDER BY name ASC")
    counts = {}
    for r in registry_rows or []:
        name = (r["name"] or "").strip()
        if not name:
            continue
        keys = r.keys() if hasattr(r, "keys") else []
        if "contact_count" in keys:
            counts[name] = int(r["contact_count"] or 0)
        else:
            counts[name] = 0

    rows = sorted(
        [{"name": name, "count": int(counts[name] or 0)} for name in counts],
        key=lambda item: (-item["count"], item["name"].lower()),
    )
    _TAG_COUNT_CACHE["ts"] = time.time()
    _TAG_COUNT_CACHE["rows"] = rows
    return rows


def _invalidate_mail_stats_cache():
    _STATS_CACHE["ts"] = 0.0
    _STATS_CACHE["payload"] = None
    _TAG_COUNT_CACHE["ts"] = 0.0
    _TAG_COUNT_CACHE["rows"] = None
    _TENANT_STATS_CACHE.clear()
    _TENANT_TAG_COUNT_CACHE.clear()


_TAG_RECOUNT_STATE = {"running": False, "queued": set()}


def _refresh_tag_counts_async(tag_names=None):
    """Etiket sayılarını arka planda canlı yenile (küçük parti)."""
    names = [str(n).strip() for n in (tag_names or []) if str(n).strip()]
    if not names:
        return

    def _run():
        try:
            with closing(get_db()) as conn:
                for name in names[:80]:
                    try:
                        _recount_tag(conn, name)
                    except Exception:
                        pass
                try:
                    conn.commit()
                except Exception:
                    pass
                _invalidate_mail_stats_cache()
        except Exception:
            pass

    threading.Thread(target=_run, daemon=True, name="mail-tag-recount").start()


def _stub_send(conn, *, channel, to_email, subject, contact=None, campaign_id=None,
               contact_id=None, template_id=None, domain_id=None, to_phone="",
               html_body="", text_body=""):
    """Geriye uyumlu sarmalayıcı — gerçek gönderim mail_delivery.deliver_mail."""
    from mail_delivery import deliver_mail

    send_id, status, _err = deliver_mail(
        conn,
        channel=channel,
        to_email=to_email,
        subject=subject,
        contact=contact,
        campaign_id=campaign_id,
        contact_id=contact_id,
        template_id=template_id,
        domain_id=domain_id,
        to_phone=to_phone,
        html_body=html_body,
        text_body=text_body,
        inject_tracking=_inject_tracking,
    )
    return send_id, status


def create_mailing_click_blueprint():
    """Public click / open / unsubscribe — auth yok."""
    bp = Blueprint("mailing_click", __name__)

    @bp.route("/m/c/<path:token>", methods=["GET"])
    def mail_click(token):
        token = _normalize_click_token(token)
        now = iso(utcnow())

        def _dead_click_page():
            # Marka / fallback link YOK — operatör sıfırdan kuracak
            html = (
                "<!doctype html><meta charset=utf-8><title>Link</title>"
                "<meta name=viewport content='width=device-width,initial-scale=1'>"
                "<body style='font-family:sans-serif;padding:1.5rem;background:#0f172a;color:#e2e8f0'>"
                "<h2 style='margin:0 0 0.5rem'>Link geçersiz</h2>"
                "<p style='color:#94a3b8;line-height:1.5'>Bu takip linki artık aktif değil.</p>"
                "</body>"
            )
            return html, 404, {"Content-Type": "text/html; charset=utf-8"}

        def _finalize_dest(dest, *, is_sc=False, contact_id=None):
            dest = (dest or "").strip()
            dest, sc2 = _split_smartico_marker(dest)
            is_sc = bool(is_sc) or sc2
            if not dest or dest.startswith("/"):
                return None
            if not re.match(r"^https?://", dest, re.I):
                dest = "https://" + dest.lstrip("/")
            if is_sc and contact_id:
                try:
                    with closing(get_db()) as c2:
                        subid_param = (get_mail_setting(c2, "smartico_subid_param", "afp1") or "afp1").strip() or "afp1"
                except Exception:
                    subid_param = "afp1"
                dest = _append_query_param(dest, subid_param, contact_id)
            return dest

        # 1) İmzalı v2 token — DB şart değil
        signed = _loads_signed_click(token)
        if signed:
            dest = _finalize_dest(
                signed.get("d"),
                is_sc=bool(signed.get("sc")),
                contact_id=signed.get("c"),
            )
            if not dest:
                return _dead_click_page()
            try:
                with closing(get_db()) as conn:
                    row = fetchone(conn, "SELECT * FROM mail_click_links WHERE token = ?", (token,))
                    if row:
                        first = row["first_clicked_at"] or now
                        execute(
                            conn,
                            """
                            UPDATE mail_click_links SET
                                click_count = COALESCE(click_count, 0) + 1,
                                first_clicked_at = ?,
                                last_clicked_at = ?
                            WHERE id = ?
                            """,
                            (first, now, row["id"]),
                        )
                        if row["send_id"]:
                            execute(
                                conn,
                                "UPDATE mail_sends SET clicked_at = COALESCE(clicked_at, ?) WHERE id = ?",
                                (now, row["send_id"]),
                            )
                        if row["contact_id"]:
                            opened = False
                            if row.get("send_id"):
                                srow = fetchone(
                                    conn,
                                    "SELECT opened_at FROM mail_sends WHERE id = ?",
                                    (row["send_id"],),
                                )
                                opened = bool(srow and srow.get("opened_at"))
                            try:
                                from mail_ops import tag_click_outcome
                                tag_click_outcome(conn, row["contact_id"], opened=opened, now=now)
                            except Exception:
                                _tag_contact(conn, row["contact_id"], "mail_tiklayan", now)
                    elif signed.get("c"):
                        try:
                            _tag_contact(conn, int(signed["c"]), "mail_tiklayan", now)
                        except Exception:
                            pass
                    conn.commit()
            except Exception as exc:
                print(f"⚠️  mail_click signed analytics: {exc}")
            return redirect(dest, code=302)

        # 2) Eski rastgele token — DB (bot/mail-client hit edebilir; analitik hatası
        # yönlendirmeyi asla bloklamasın — imzalı v2 dalıyla aynı davranış)
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_click_links WHERE token = ?", (token,))
            if not row:
                return _dead_click_page()
            row = _row(row)
            dest = _finalize_dest(
                row.get("dest_url"),
                is_sc=bool(row.get("is_smartico")),
                contact_id=row.get("contact_id"),
            )
            if not dest:
                return _dead_click_page()
            try:
                first = row.get("first_clicked_at") or now
                execute(
                    conn,
                    """
                    UPDATE mail_click_links SET
                        click_count = COALESCE(click_count, 0) + 1,
                        first_clicked_at = ?,
                        last_clicked_at = ?
                    WHERE id = ?
                    """,
                    (first, now, row["id"]),
                )
                if row.get("send_id"):
                    execute(
                        conn,
                        "UPDATE mail_sends SET clicked_at = COALESCE(clicked_at, ?) WHERE id = ?",
                        (now, row["send_id"]),
                    )
                if row.get("contact_id"):
                    opened = False
                    if row.get("send_id"):
                        srow = _row(fetchone(
                            conn,
                            "SELECT opened_at FROM mail_sends WHERE id = ?",
                            (row["send_id"],),
                        ))
                        opened = bool(srow and srow.get("opened_at"))
                    try:
                        from mail_ops import tag_click_outcome
                        tag_click_outcome(conn, row["contact_id"], opened=opened, now=now)
                    except Exception:
                        _tag_contact(conn, row["contact_id"], "mail_tiklayan", now)
                conn.commit()
            except Exception as exc:
                print(f"⚠️  mail_click legacy analytics: {exc}")
                try:
                    conn.rollback()
                except Exception:
                    pass
        return redirect(dest, code=302)

    @bp.route("/m/o/<int:send_id>/<sig>", methods=["GET"])
    def mail_open_pixel(send_id, sig):
        from mail_ops import record_open, verify_open_sig
        # 1x1 GIF
        pixel = (
            b"GIF89a\x01\x00\x01\x00\x80\x00\x00\xff\xff\xff\x00\x00\x00!\xf9\x04"
            b"\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;"
        )
        try:
            with closing(get_db()) as conn:
                if verify_open_sig(conn, send_id, sig):
                    record_open(conn, send_id)
                    conn.commit()
                else:
                    # Eski webhook_secret ile imzalanmış mailler — mail_ops_secret ile dene
                    print(f"⚠️  open sig mismatch send_id={send_id}")
        except Exception as exc:
            print(f"⚠️  mail_open_pixel: {exc}")
        return pixel, 200, {
            "Content-Type": "image/gif",
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
        }

    @bp.route("/m/u/<token>", methods=["GET", "POST"])
    def mail_unsubscribe(token):
        from mail_ops import apply_unsubscribe
        try:
            with closing(get_db()) as conn:
                ok, info = apply_unsubscribe(conn, token)
                try:
                    conn.commit()
                except Exception:
                    pass
        except Exception as exc:
            print(f"⚠️  mail_unsubscribe: {exc}")
            ok, info = False, ""
        if not ok:
            return (
                "<!doctype html><meta charset=utf-8><title>Hata</title>"
                "<body style='font-family:sans-serif;padding:2rem'>"
                "<h2>Geçersiz veya süresi dolmuş bağlantı</h2></body>",
                404,
            )
        return (
            "<!doctype html><meta charset=utf-8><title>Abonelik iptal</title>"
            "<body style='font-family:sans-serif;padding:2rem;max-width:520px'>"
            "<h2>Abonelikten çıktınız</h2>"
            "<p>Bu adres artık mailing listelerimize eklenmeyecek.</p>"
            f"<p class='muted' style='color:#6b7280'>{html_lib.escape(str(info or ''))}</p>"
            "</body>",
            200,
        )

    return bp


def _purge_all_mail_click_links_once():
    """Bir kerelik: tüm takip yönlendirme kayıtlarını sil (operatör sıfırdan kuracak)."""
    flag = "purge_all_click_links_v20260726a"
    try:
        with closing(get_db()) as conn:
            if (get_mail_setting(conn, flag, "") or "").strip() == "1":
                return 0
            before = 0
            try:
                before = int(scalar(conn, "SELECT COUNT(*) FROM mail_click_links") or 0)
            except Exception:
                before = -1
            if uses_postgres():
                try:
                    execute(conn, "TRUNCATE TABLE mail_click_links RESTART IDENTITY CASCADE")
                except Exception:
                    execute(conn, "DELETE FROM mail_click_links")
            else:
                execute(conn, "DELETE FROM mail_click_links")
            upsert_mail_setting(conn, flag, "1")
            conn.commit()
            print(f"✉️  purged mail_click_links: {before}")
            return before
    except Exception as exc:
        print(f"⚠️  purge mail_click_links: {exc}")
        return 0


def _purge_all_mail_contacts_once():
    """Bir kerelik: tüm mail kontakları + bağlı gönderim/tıklama kayıtlarını sil.

    Panel kasmasını bitirmek için deploy'da çalışır; mail_settings ile tekrarlanmaz.
    """
    flag = "purge_all_contacts_v20260713a"
    try:
        with closing(get_db()) as conn:
            if (get_mail_setting(conn, flag, "") or "").strip() == "1":
                return 0
            before = 0
            try:
                before = int(scalar(conn, "SELECT COUNT(*) FROM mail_contacts") or 0)
            except Exception:
                before = -1
            if uses_postgres():
                execute(
                    conn,
                    """
                    TRUNCATE TABLE
                      mail_campaign_recipients,
                      mail_click_links,
                      mail_ivr_events,
                      mail_sends,
                      mail_contacts
                    RESTART IDENTITY CASCADE
                    """,
                )
                try:
                    execute(conn, "TRUNCATE TABLE mail_import_jobs RESTART IDENTITY CASCADE")
                except Exception:
                    execute(conn, "DELETE FROM mail_import_jobs")
                try:
                    execute(conn, "UPDATE mail_contact_tags SET contact_count = 0")
                except Exception:
                    pass
            else:
                for table in (
                    "mail_campaign_recipients",
                    "mail_click_links",
                    "mail_ivr_events",
                    "mail_sends",
                    "mail_contacts",
                    "mail_import_jobs",
                ):
                    try:
                        execute(conn, f"DELETE FROM {table}")
                    except Exception:
                        pass
                try:
                    execute(conn, "UPDATE mail_contact_tags SET contact_count = 0")
                except Exception:
                    pass
            upsert_mail_setting(conn, flag, "1")
            conn.commit()
            _invalidate_mail_stats_cache()
            print(f"🧹 mail contacts purged once (before≈{before})")
            return before
    except Exception as exc:
        print(f"⚠️  mail contacts purge failed: {exc}")
        return -1


def _delivery_health_snapshot(conn, tenant_id=None):
    """Gerçek SMTP mi, stub/simüle mi — son 7 gün özeti (şifre yazmaz).

    tenant_id verilirse SADECE o tenant'ın gönderimleri sayılır — önceden bu
    filtre yoktu, her firma platformun toplam success-rate'ini görebiliyordu.
    """
    provider = (get_mail_setting(conn, "provider_mode", "stub") or "stub").strip().lower()
    smtp_host = (get_mail_setting(conn, "smtp_host", "") or "").strip()
    smtp_user = (get_mail_setting(conn, "smtp_user", "") or "").strip()
    has_settings_pw = bool((get_mail_setting(conn, "smtp_password", "") or "").strip())
    from datetime import timedelta
    since = (utcnow() - timedelta(days=7)).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _count(status=None, stub_msgid=False):
        clauses = ["CAST(created_at AS TEXT) >= ?"]
        params = [since]
        if status:
            clauses.append("status = ?")
            params.append(status)
        if stub_msgid:
            clauses.append("provider_msg_id LIKE 'stub-%'")
        if tenant_id:
            clauses.append("tenant_id = ?")
            params.append(int(tenant_id))
        try:
            return int(scalar(
                conn,
                f"SELECT COUNT(*) FROM mail_sends WHERE {' AND '.join(clauses)}",
                tuple(params),
            ) or 0)
        except Exception:
            return 0

    real_7 = _count("sent")
    sim_7 = _count("simulated")
    fail_7 = _count("failed")
    bounced_7 = _count("bounced")
    skip_7 = _count("skipped")
    queued_7 = _count("queued")
    stub_msgid_7 = _count(stub_msgid=True)

    # Alibaba hesap yöneticisinin gördüğü "success rate" — sent / (sent + fail).
    # Limit artırımı için >=%90, sürdürme için >=%80 şart (WhatsApp: 20.08.2026 görüşme).
    since_24 = (utcnow() - timedelta(hours=24)).strftime("%Y-%m-%dT%H:%M:%SZ")

    def _count_since(since_ts, status=None):
        clauses = ["CAST(created_at AS TEXT) >= ?"]
        params = [since_ts]
        if status:
            clauses.append("status = ?")
            params.append(status)
        if tenant_id:
            clauses.append("tenant_id = ?")
            params.append(int(tenant_id))
        try:
            return int(scalar(
                conn,
                f"SELECT COUNT(*) FROM mail_sends WHERE {' AND '.join(clauses)}",
                tuple(params),
            ) or 0)
        except Exception:
            return 0

    def _success_rate(since_ts):
        accepted = _count_since(since_ts, "sent") + _count_since(since_ts, "simulated")
        rejected = _count_since(since_ts, "failed") + _count_since(since_ts, "bounced")
        total = accepted + rejected
        rate = round(100.0 * accepted / total, 2) if total else None
        if rate is None:
            tier = "no_data"
        elif rate >= 90:
            tier = "good"
        elif rate >= 80:
            tier = "warn"
        else:
            tier = "danger"
        return {
            "accepted": accepted,
            "rejected": rejected,
            "total": total,
            "rate": rate,
            "tier": tier,
        }

    success_rate = {
        "last_24h": _success_rate(since_24),
        "last_7d": _success_rate(since),
        "target_increase": 90.0,
        "target_maintain": 80.0,
        "note": "Alibaba: limit artışı için ≥%90, mevcut limiti sürdürmek için ≥%80 gerekli.",
    }
    samples = []
    try:
        # to_email göstermek başka firmanın alıcı adreslerini sızdırabilir —
        # tenant_id verilmişse SADECE o firmanın gönderim örnekleri gelir.
        sample_where = "WHERE tenant_id = ?" if tenant_id else ""
        sample_params = (int(tenant_id),) if tenant_id else ()
        rows = fetchall(
            conn,
            f"""
            SELECT id, to_email, status, provider_msg_id, error, created_at, channel
            FROM mail_sends
            {sample_where}
            ORDER BY id DESC LIMIT 12
            """,
            sample_params,
        ) or []
        for r in rows:
            d = dict(r) if not isinstance(r, dict) else r
            mid = (d.get("provider_msg_id") or "").strip()
            samples.append({
                "id": d.get("id"),
                "to_email": d.get("to_email"),
                "status": d.get("status"),
                "channel": d.get("channel"),
                "created_at": d.get("created_at"),
                "error": ((d.get("error") or "")[:160] or None),
                "msgid_kind": (
                    "stub" if mid.startswith("stub-")
                    else ("real" if mid else "empty")
                ),
                "has_msgid": bool(mid),
            })
    except Exception as exc:
        samples = [{"error": str(exc)[:120]}]

    domains_pw = 0
    domains_total = 0
    try:
        if tenant_id:
            domains_total = int(scalar(
                conn,
                "SELECT COUNT(*) FROM mail_domain_allocations WHERE tenant_id = ?",
                (int(tenant_id),),
            ) or 0)
            domains_pw = int(scalar(
                conn,
                """
                SELECT COUNT(*) FROM mail_domains d
                JOIN mail_domain_allocations a ON a.domain_id = d.id
                WHERE a.tenant_id = ?
                  AND (COALESCE(d.smtp_password,'') != '' OR COALESCE(d.smtp_password_enc,'') != '')
                """,
                (int(tenant_id),),
            ) or 0)
        else:
            domains_total = int(scalar(conn, "SELECT COUNT(*) FROM mail_domains") or 0)
            domains_pw = int(scalar(
                conn,
                """
                SELECT COUNT(*) FROM mail_domains
                WHERE (COALESCE(smtp_password,'') != '' OR COALESCE(smtp_password_enc,'') != '')
                """,
            ) or 0)
    except Exception:
        pass

    is_stub = provider != "smtp"
    ghost = (not is_stub) and sim_7 > 0 and real_7 == 0
    ok = (not is_stub) and real_7 > 0 and not ghost
    if is_stub:
        verdict = "stub_mode"
        message = (
            "Sağlayıcı STUB — kampanyalar simüle edilir, gerçek SMTP’ye çıkılmaz. "
            "Ayarlar → Gönderim sağlayıcı → SMTP / DirectMail."
        )
    elif real_7 == 0 and (sim_7 > 0 or stub_msgid_7 > 0):
        verdict = "simulated_only"
        message = (
            f"Son 7 günde gerçek sent yok; simüle={sim_7}. "
            "Panel ‘iletilen’ göstermiş olabilir ama kutu boş kalır."
        )
    elif real_7 == 0 and fail_7 > 0:
        verdict = "all_failed"
        message = f"Son 7 günde {fail_7} fail, 0 gerçek gönderim — SMTP/şifre/domain kontrol et."
    elif real_7 == 0:
        verdict = "no_recent_sends"
        message = "Son 7 günde başarılı gönderim kaydı yok."
    else:
        verdict = "smtp_ok"
        message = f"Son 7 günde {real_7} gerçek SMTP gönderim kaydı var."

    return {
        "provider_mode": provider,
        "smtp_configured": bool(smtp_host and (has_settings_pw or domains_pw)),
        "smtp_host_set": bool(smtp_host),
        "smtp_user_set": bool(smtp_user),
        "settings_password_set": has_settings_pw,
        "domains_with_password": domains_pw,
        "domains_total": domains_total,
        "last_7d": {
            "sent": real_7,
            "simulated": sim_7,
            "failed": fail_7,
            "bounced": bounced_7,
            "skipped": skip_7,
            "queued": queued_7,
            "stub_msgid": stub_msgid_7,
        },
        "success_rate": success_rate,
        "ok": ok,
        "verdict": verdict,
        "message": message,
        "samples": samples,
    }


def create_mailing_blueprint(permission_required):
    from mail_campaign_worker import ensure_campaign_scheduler
    from mail_crm import ensure_mail_crm_schema
    from mail_ops import ensure_mail_ops_schema
    from mail_scrub import cancel_active_scrub_jobs, ensure_mail_scrub_schema

    external_worker = (os.environ.get("MAILING_WORKER_EXTERNAL") or "").strip().lower() in (
        "1", "true", "yes", "on",
    )
    if external_worker:
        print("✉️  mailing: external worker mode — in-process scheduler kapalı")
    else:
        ensure_campaign_scheduler()
    try:
        with closing(get_db()) as conn:
            ensure_mail_ops_schema(conn)
            ensure_mail_scrub_schema(conn)
            ensure_mail_crm_schema(conn)
            try:
                from mail_tenant import ensure_tenant_schema
                ensure_tenant_schema(conn)
                conn.commit()
            except Exception as ten_exc:
                print(f"⚠️  mail tenant schema: {ten_exc}")
            try:
                from mail_account_quota import ensure_quota_defaults
                ensure_quota_defaults(conn)
                conn.commit()
            except Exception as q_exc:
                print(f"⚠️  account quota defaults: {q_exc}")
            try:
                from mail_credit import ensure_credit_defaults
                ensure_credit_defaults(conn)
                conn.commit()
            except Exception as c_exc:
                print(f"⚠️  mail credit defaults: {c_exc}")
            try:
                from mail_template_wipe import ensure_templates_wiped_once
                wiped = ensure_templates_wiped_once(conn)
                if wiped:
                    print(f"✉️  wiped mail templates: deleted={wiped.get('deleted')}")
            except Exception as wipe_exc:
                print(f"⚠️  mail template wipe: {wipe_exc}")
            try:
                from mail_template_seeds_v2026 import seed_makrobet_2026_templates
                _v26 = seed_makrobet_2026_templates(conn, overwrite=True)
                if _v26.get("skipped"):
                    print("✉️  Makrobet 2026 seed skipped (wipe auto-seed off)")
                else:
                    print(
                        f"✉️  seeded Makrobet 2026 templates: "
                        f"added={_v26.get('added')} updated={_v26.get('updated')}"
                    )
            except Exception as seed_exc:
                print(f"⚠️  mail template seed v2026: {seed_exc}")
            try:
                from mail_template_seeds_bizzo import seed_bizzo_mail_templates
                # Wipe sonrası auto-seed kapalıysa Bizzo da geri gelmesin
                _bz = seed_bizzo_mail_templates(conn, overwrite=True)
                if _bz.get("skipped"):
                    print("✉️  Bizzo 2026 seed skipped (wipe auto-seed off)")
                else:
                    print(
                        f"✉️  seeded Bizzo 2026 templates: "
                        f"added={_bz.get('added')} updated={_bz.get('updated')}"
                    )
            except Exception as seed_exc:
                print(f"⚠️  mail template seed bizzo 2026: {seed_exc}")
            try:
                from mail_template_seeds import seed_makrobet_mail_templates
                from mail_template_seeds_bizzo import seed_bizzo_mail_templates
                _seed_res = seed_makrobet_mail_templates(conn)
                _bizzo_res = seed_bizzo_mail_templates(conn)
                n = (_seed_res.get("added", 0) + _seed_res.get("updated", 0)) if isinstance(_seed_res, dict) else int(_seed_res or 0)
                nb = (_bizzo_res.get("added", 0) + _bizzo_res.get("updated", 0)) if isinstance(_bizzo_res, dict) else int(_bizzo_res or 0)
                if n:
                    print(f"✉️  seeded {n} Makrobet mail templates")
                if nb:
                    print(f"✉️  seeded {nb} Bizzo mail templates")
            except Exception as seed_exc:
                print(f"⚠️  mail template seed: {seed_exc}")
            try:
                from mail_weekly_maintenance import ensure_sunday_maintenance
                _wm = ensure_sunday_maintenance(conn)
                if _wm and not _wm.get("skipped") and _wm.get("ok"):
                    print(
                        f"✉️  Sunday weekly maintenance ran: "
                        f"week={_wm.get('week_key')} actions={len(_wm.get('actions') or [])}"
                    )
                elif _wm and _wm.get("skipped"):
                    print(f"✉️  Sunday weekly maintenance already done ({_wm.get('week_key')})")
            except Exception as wm_exc:
                print(f"⚠️  weekly maintenance: {wm_exc}")
            try:
                from mail_template_cta_repair import repair_mail_cta_links
                _cta = repair_mail_cta_links(conn)
                print(
                    f"✉️  CTA repair (templates only): updated={_cta.get('templates_updated')} "
                    f"bizzo={_cta.get('bizzo_cta')} makro={_cta.get('makro_cta')}"
                )
            except Exception as cta_exc:
                print(f"⚠️  CTA repair: {cta_exc}")
    except Exception as exc:
        print(f"⚠️  mail_ops/scrub ensure: {exc}")
    try:
        _purge_all_mail_click_links_once()
    except Exception as exc:
        print(f"⚠️  startup click-links purge: {exc}")
    try:
        _cancel_all_active_imports()
    except Exception as exc:
        print(f"⚠️  startup import cancel: {exc}")
    # External worker modunda scrub/campaign worker'da yürür — web restart iptal etmesin
    if not external_worker:
        try:
            cancel_active_scrub_jobs()
        except Exception as exc:
            print(f"⚠️  startup scrub cancel: {exc}")
    # Standalone Mikromail'de kontak purge kapalı (satış verisi)
    if not external_worker and (os.environ.get("SERVICE_MODE") or "").strip().lower() != "mailing":
        try:
            _purge_all_mail_contacts_once()
        except Exception as exc:
            print(f"⚠️  startup contacts purge: {exc}")
    bp = Blueprint("mailing", __name__, url_prefix="/api/mailing")

    def mail_perm(*keys):
        """Yetki + (embedded) Makro allowlist — standalone mail session muaf."""
        decorated = permission_required(*keys)

        def decorator(view):
            from functools import wraps

            from flask import jsonify, session

            from panel_config import can_access_mailing

            inner = decorated(view)

            @wraps(view)
            def wrapped(*args, **kwargs):
                if session.get("mail_logged_in"):
                    return inner(*args, **kwargs)
                if (os.environ.get("SERVICE_MODE") or "").strip().lower() == "mailing":
                    return inner(*args, **kwargs)
                if not can_access_mailing(session.get("admin_username")):
                    return jsonify({"error": "Mailing yalnızca yetkili hesap içindir."}), 403
                return inner(*args, **kwargs)

            return wrapped

        return decorator

    # ── Dashboard ──────────────────────────────────────────────
    # (delivery health helper is module-level — see _delivery_health_snapshot)

    def _dashboard_kpi_for_tenant(conn, tid):
        """Bir tenant (None = platform genel) için dashboard KPI seti.

        tid verilirse TÜM sorgular tenant_id ile filtrelenir — önceden hiçbir
        filtre yoktu, her firma platformun toplam rakamlarını görüyordu.
        """
        tid_clause = " WHERE tenant_id = ?" if tid else ""
        tid_and = " AND tenant_id = ?" if tid else ""
        tid_params = (int(tid),) if tid else ()

        contacts, contacts_approx = _approx_contact_total(conn, tenant_id=tid)
        active_contacts = contacts
        templates = scalar(conn, f"SELECT COUNT(*) FROM mail_templates{tid_clause}", tid_params) or 0
        campaigns = scalar(conn, f"SELECT COUNT(*) FROM mail_campaigns{tid_clause}", tid_params) or 0
        if tid:
            sends_total = scalar(conn, "SELECT COUNT(*) FROM mail_sends WHERE tenant_id = ?", tid_params) or 0
        elif uses_postgres():
            try:
                sends_total = int(scalar(
                    conn,
                    "SELECT reltuples::bigint FROM pg_class WHERE relname = 'mail_sends'",
                ) or 0)
                if sends_total < 0:
                    sends_total = 0
            except Exception:
                sends_total = scalar(conn, "SELECT COUNT(*) FROM mail_sends") or 0
        else:
            sends_total = scalar(conn, "SELECT COUNT(*) FROM mail_sends") or 0
        sends_real = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE status = 'sent'{tid_and}", tid_params
        ) or 0)
        sends_simulated = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE status = 'simulated'{tid_and}", tid_params
        ) or 0)
        sends_sim = sends_real + sends_simulated
        sends_queued = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE status = 'queued'{tid_and}", tid_params
        ) or 0)
        sends_failed = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE status = 'failed'{tid_and}", tid_params
        ) or 0)
        delivery_health = _delivery_health_snapshot(conn, tenant_id=tid)
        opened = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE opened_at IS NOT NULL{tid_and}", tid_params
        ) or 0)
        clicked = int(scalar(
            conn, f"SELECT COUNT(*) FROM mail_sends WHERE clicked_at IS NOT NULL{tid_and}", tid_params
        ) or 0)
        try:
            if tid:
                link_clicked = int(scalar(
                    conn,
                    """
                    SELECT COUNT(DISTINCT cl.send_id) FROM mail_click_links cl
                    JOIN mail_sends s ON s.id = cl.send_id
                    WHERE cl.send_id IS NOT NULL AND COALESCE(cl.click_count, 0) > 0
                      AND s.tenant_id = ?
                    """,
                    tid_params,
                ) or 0)
            else:
                link_clicked = int(scalar(
                    conn,
                    """
                    SELECT COUNT(DISTINCT send_id) FROM mail_click_links
                    WHERE send_id IS NOT NULL AND COALESCE(click_count, 0) > 0
                    """,
                ) or 0)
            if link_clicked > clicked:
                clicked = link_clicked
        except Exception:
            pass
        ivr_events = scalar(
            conn, f"SELECT COUNT(*) FROM mail_ivr_events{tid_clause}", tid_params
        ) or 0
        try:
            suppressed = int(scalar(
                conn, f"SELECT COUNT(*) FROM mail_suppressions{tid_clause}", tid_params
            ) or 0)
        except Exception:
            suppressed = 0
        return {
            "contacts": contacts,
            "contacts_approx": bool(contacts_approx),
            "active_contacts": active_contacts,
            "templates": templates,
            "campaigns": campaigns,
            "sends_total": sends_total,
            "sends_delivered": sends_sim,
            "sends_real": sends_real,
            "sends_simulated": sends_simulated,
            "sends_queued": sends_queued,
            "sends_failed": sends_failed,
            "opened": opened,
            "clicked": clicked,
            "ivr_events": ivr_events,
            "suppressed": suppressed,
        }, delivery_health

    @bp.route("/dashboard", methods=["GET"])
    @mail_perm(*MAIL_DASH)
    def dashboard():
        from flask import session as _sess
        from mail_tenant import current_tenant_id

        _tid = current_tenant_id()
        with closing(get_db()) as conn:
            kpi, delivery_health = _dashboard_kpi_for_tenant(conn, _tid)
            try:
                from mail_tenant import enrich_domain_public, heal_ready_domains, list_allocated_domains

                heal_ready_domains(conn)
            except Exception:
                pass
            # Domainler: tenant seçiliyse SADECE o firmaya atanmış domainler —
            # önceden filtre yoktu, her firma platformdaki TÜM domainleri
            # (başka firmalara ait olanlar dahil) görebiliyordu.
            if _tid:
                from mail_tenant import list_allocated_domains
                raw_domains = list_allocated_domains(conn, int(_tid)) or []
            else:
                raw_domains = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC") or []
            try:
                from mail_tenant import enrich_domain_public

                domains = [enrich_domain_public(r) for r in raw_domains]
            except Exception:
                domains = _rows(raw_domains)
            provider = get_mail_setting(conn, "provider_mode", "stub")
            from mail_ops import smartico_dashboard_summary

            sc = smartico_dashboard_summary(conn)

            by_tenant = None
            if not _tid and _sess.get("mail_is_superadmin"):
                # "Tümü" görünümü — superadmin genel toplamların yanında
                # firma firma (isim isim) ayrı kartlar da görür.
                tenants = fetchall(
                    conn,
                    "SELECT id, slug, name, status FROM mail_tenants "
                    "WHERE status != 'deleted' ORDER BY name ASC",
                ) or []
                by_tenant = []
                for t in tenants:
                    t = _row(t)
                    t_kpi, t_health = _dashboard_kpi_for_tenant(conn, int(t["id"]))
                    by_tenant.append({
                        "tenant_id": t["id"],
                        "slug": t.get("slug"),
                        "name": t.get("name"),
                        "status": t.get("status"),
                        "kpi": t_kpi,
                        "success_rate": t_health.get("success_rate"),
                    })
        return jsonify({
            "kpi": {
                **kpi,
                "sc_register": sc.get("register"),
                "sc_deposit_total": sc.get("deposit_total"),
                "sc_ftd_count": sc.get("ftd_count"),
                "sc_ftd_total": sc.get("ftd_total"),
                "sc_withdraw_total": sc.get("withdraw_total"),
                "sc_bonus_total": sc.get("bonus_total"),
                "sc_currency": sc.get("currency") or "",
            },
            "smartico": sc,
            "domains": domains,
            "provider_mode": provider,
            "delivery_health": delivery_health,
            "by_tenant": by_tenant,
            "view_scope": "tenant" if _tid else "all",
            "note": (
                "SMTP (DirectMail) aktif — domainler gönderime hazır."
                if (provider or "").strip().lower() == "smtp"
                else "⚠️ STUB MOD — mailler gerçek gitmiyor; panel simüle ediyor. Ayarlar → SMTP."
            ),
        })

    @bp.route("/delivery-health", methods=["GET"])
    @mail_perm(*MAIL_DASH, *MAIL_REP)
    def delivery_health():
        """Son gönderimlerin gerçek SMTP mi yoksa stub/simüle mi olduğunu özetler."""
        from mail_tenant import current_tenant_id
        with closing(get_db()) as conn:
            snap = _delivery_health_snapshot(conn, tenant_id=current_tenant_id())
        return jsonify(snap)

    # ── Contacts / CRM ─────────────────────────────────────────
    @bp.route("/contacts/stats", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def contact_stats():
        """CRM özet — hızlı path. ?refresh=1 ile etiket sayıları canlı yenilenir.

        tenant_id set ise (tenant login veya superadmin impersonate) SADECE o
        tenant'ın kontakları/gönderimleri/etiketleri sayılır ve ayrı bir
        tenant-scoped cache kullanılır — önceden bu filtre yoktu, her firma
        platformun toplam kontak/etiket sayısını görebiliyordu.
        """
        import time
        from mail_tenant import current_tenant_id

        refresh = (request.args.get("refresh") or "").strip() in ("1", "true", "yes")
        sync_tags = (request.args.get("sync_tags") or "").strip() in ("1", "true", "yes")
        _tid = current_tenant_id()
        now = time.time()

        if _tid:
            cached = _TENANT_STATS_CACHE.get(int(_tid))
            if not refresh and not sync_tags and cached and (now - cached["ts"]) < 60:
                return jsonify(cached["payload"])
        else:
            if not refresh and not sync_tags and _STATS_CACHE["payload"] and (now - _STATS_CACHE["ts"]) < 60:
                # Cache hit — arka plan sync/recount YOK (milyonluk LIKE taramaları paneli kitler)
                return jsonify(_STATS_CACHE["payload"])

        with closing(get_db()) as conn:
            total, total_approx = _approx_contact_total(conn, tenant_id=_tid)
            mailed = _approx_mailed_contacts(conn, total, tenant_id=_tid)
            if mailed is None:
                mailed = 0
                never_mailed = total
                mailed_approx = True
            else:
                mailed = min(int(mailed), int(total or 0))
                never_mailed = max(int(total or 0) - mailed, 0)
                mailed_approx = False
            if sync_tags and not _tid:
                _maybe_sync_missing_tags_async(force=True)
            by_tag = _contact_tag_counts(conn, live=False, tenant_id=_tid)
            pending_recount = []
            # Registry 0 kaldıysa (import sonrası) — küçük/orta listelerde canlı say
            need_live = False
            if by_tag and int(total or 0) > 0 and not _tid:
                zeros = sum(1 for t in by_tag if int(t.get("count") or 0) == 0)
                if zeros > 0 and int(total or 0) <= 200000:
                    need_live = True
                if refresh and int(total or 0) <= 500000:
                    need_live = True
            if need_live:
                by_tag = _contact_tag_counts(conn, force=True, live=True, tenant_id=_tid)
            elif _tid and refresh:
                by_tag = _contact_tag_counts(conn, force=True, live=False, tenant_id=_tid)
        payload = {
            "total": total,
            "total_approx": bool(total_approx),
            "mailed": mailed,
            "mailed_approx": mailed_approx,
            "never_mailed": never_mailed,
            "by_tag": by_tag,
            "tag_count": len(by_tag),
            "pending_tag_recount": pending_recount,
            "cached": not refresh,
        }
        with _STATS_LOCK:
            if _tid:
                _TENANT_STATS_CACHE[int(_tid)] = {"ts": time.time(), "payload": payload}
            else:
                _STATS_CACHE["ts"] = time.time()
                _STATS_CACHE["payload"] = payload
        return jsonify(payload)

    @bp.route("/contacts", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def list_contacts():
        q = (request.args.get("q") or "").strip().lower()
        tag = (request.args.get("tag") or "").strip()
        tags_raw = (request.args.get("tags") or "").strip()
        tag_list = _parse_tag_filter_list(tags_raw) if tags_raw else ([tag] if tag else [])
        limit = min(int(request.args.get("limit") or 200), 1000)
        try:
            offset = max(0, int(request.args.get("offset") or 0))
        except (TypeError, ValueError):
            offset = 0
        with closing(get_db()) as conn:
            clauses = []
            params = []
            if tag_list:
                clause, tparams = _tag_match_any_clause(tag_list)
                clauses.append(clause)
                params.extend(tparams)
            if q:
                clauses.append("(LOWER(email) LIKE ? OR LOWER(name) LIKE ? OR LOWER(phone) LIKE ?)")
                like = f"%{q}%"
                params.extend([like, like, like])
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid:
                    clauses.append("tenant_id = ?")
                    params.append(int(_tid))
            except Exception:
                pass
            where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

            # 3M satırda COUNT(*) + LIKE paneli kilitler / 15s abort
            total = None
            total_approx = False
            if len(tag_list) == 1 and not q:
                reg = _registry_tag_count(conn, tag_list[0])
                if reg is not None:
                    total = reg
                else:
                    # Registry yoksa sayma — sayfa sonucu yeter
                    total = None
                    total_approx = True
            elif not tag_list and not q:
                total, total_approx = _approx_contact_total(conn)
            else:
                # q / çoklu etiket: exact count pahalı — atla
                total = None
                total_approx = True

            rows = _rows(fetchall(
                conn,
                f"SELECT * FROM mail_contacts{where} ORDER BY id DESC LIMIT ? OFFSET ?",
                tuple(params) + (limit, offset),
            ))
            # Görünen etiketleri registry'ye ekle (ucuz, eksik dropdown düzeltir)
            page_tags = set()
            for r in rows:
                page_tags.update(_parse_tags(r.get("tags") if isinstance(r, dict) else r["tags"]))
            if page_tags:
                added = _harvest_tags_into_registry(conn, page_tags)
                if added:
                    try:
                        conn.commit()
                    except Exception:
                        pass
                    _invalidate_mail_stats_cache()

            if total is None:
                # En az sayfa boyutu kadar göster
                total = len(rows) + offset
                if len(rows) >= limit:
                    total_approx = True

        out = [_contact_out(r) for r in rows]
        return jsonify({
            "contacts": out,
            "count": len(out),
            "total": int(total),
            "total_approx": bool(total_approx),
            "limit": limit,
            "offset": offset,
            "has_more": len(out) >= limit,
            "tags": tag_list,
        })

    @bp.route("/contacts/bulk-delete", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def bulk_delete_contacts():
        """Seçili ID'leri veya bir etiketteki kontakları sil (rehberden)."""
        data = request.get_json(silent=True) or {}
        raw_ids = data.get("contact_ids") or []
        tag = (data.get("tag") or "").strip()
        confirm_tag = (data.get("confirm_tag") or "").strip()
        try:
            limit = min(int(data.get("limit") or 5000), 20000)
        except (TypeError, ValueError):
            limit = 5000

        ids = []
        if isinstance(raw_ids, list):
            for x in raw_ids:
                try:
                    ids.append(int(x))
                except (TypeError, ValueError):
                    continue
        ids = list(dict.fromkeys(ids))[:5000]

        with closing(get_db()) as conn:
            deleted = 0
            try:
                if ids:
                    # Tenant scope: firma sadece kendi kontaklarını toplu silebilir
                    try:
                        from mail_tenant import current_tenant_id
                        _tid = current_tenant_id()
                    except Exception:
                        _tid = None
                    if _tid:
                        ph = ",".join(["?"] * len(ids))
                        owned = fetchall(
                            conn,
                            f"SELECT id FROM mail_contacts WHERE id IN ({ph}) AND tenant_id = ?",
                            tuple(ids) + (int(_tid),),
                        )
                        ids = [int(r["id"]) for r in (owned or [])]
                    for i in range(0, len(ids), 200):
                        chunk = ids[i : i + 200]
                        deleted += _delete_contacts_by_ids(conn, chunk)
                        try:
                            conn.commit()
                        except Exception:
                            pass
                elif tag:
                    if confirm_tag != tag:
                        return jsonify({
                            "error": f"Onay için confirm_tag olarak tam etiket adını gönder: «{tag}»",
                        }), 400
                    clause, tparams = _tag_match_clause(tag)
                    # Tenant scope
                    extra = ""
                    params = list(tparams)
                    try:
                        from mail_tenant import current_tenant_id
                        _tid = current_tenant_id()
                        if _tid:
                            extra = " AND tenant_id = ?"
                            params.append(int(_tid))
                    except Exception:
                        pass
                    # Parça parça sil — bellek şişmesin + FK bağımlılıkları temizlensin
                    exhausted = False
                    while deleted < limit:
                        rows = fetchall(
                            conn,
                            f"SELECT id FROM mail_contacts WHERE {clause}{extra} ORDER BY id ASC LIMIT 200",
                            tuple(params),
                        ) or []
                        if not rows:
                            exhausted = True
                            break
                        chunk_ids = [int(r["id"]) for r in rows]
                        deleted += _delete_contacts_by_ids(conn, chunk_ids)
                        try:
                            conn.commit()
                        except Exception:
                            pass
                        if len(chunk_ids) < 200:
                            exhausted = True
                            break
                    # Tamamen bittiyse COUNT atmadan 0; limit kesildiyse hafif recount
                    try:
                        if exhausted:
                            _set_registry_tag_count(conn, tag, 0)
                        else:
                            _recount_tag(conn, tag)
                    except Exception:
                        try:
                            _set_registry_tag_count(conn, tag, 0)
                        except Exception:
                            pass
                else:
                    return jsonify({"error": "contact_ids veya tag gerekli."}), 400

                _invalidate_mail_stats_cache()
                try:
                    from mail_ops import audit
                    audit(
                        conn,
                        request.headers.get("X-Admin-User") or "admin",
                        "contacts_bulk_delete",
                        f"deleted={deleted} tag={tag or '-'} ids={len(ids)}",
                    )
                except Exception:
                    pass
                conn.commit()
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                return jsonify({
                    "error": f"Silme başarısız: {exc}",
                    "deleted": deleted,
                }), 500
        return jsonify({
            "ok": True,
            "deleted": deleted,
            "tag": tag or None,
            "message": f"{deleted} kontak silindi.",
        })

    @bp.route("/contacts", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def create_contact():
        data = request.get_json(silent=True) or {}
        email = (data.get("email") or "").strip().lower()
        if not email or not EMAIL_RE.match(email):
            return jsonify({"error": "Geçerli bir e-posta girin."}), 400
        now = iso(utcnow())
        tags = _tags_json(data.get("tags"))
        with closing(get_db()) as conn:
            _tid = None
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if _tid:
                existing = fetchone(
                    conn,
                    "SELECT id FROM mail_contacts WHERE LOWER(email) = ? AND tenant_id = ?",
                    (email, int(_tid)),
                )
            else:
                existing = fetchone(conn, "SELECT id FROM mail_contacts WHERE LOWER(email) = ?", (email,))
            if existing:
                return jsonify({"error": "Bu e-posta zaten kayıtlı.", "id": existing["id"]}), 409
            if _tid:
                # SAVEPOINT: eski global UNIQUE(email) hâlâ devredeyse (şema migration'ı
                # henüz uygulanmadıysa) başka firmanın kaydıyla çakışabilir — 500 patlatmak
                # yerine net bir mesajla dön.
                sp = "sp_create_contact"
                try:
                    execute(conn, f"SAVEPOINT {sp}")
                    cid = insert_returning_id(
                        conn,
                        """
                        INSERT INTO mail_contacts
                        (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at, tenant_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            email,
                            (data.get("phone") or "").strip(),
                            (data.get("name") or "").strip(),
                            tags,
                            (data.get("source") or "manual").strip() or "manual",
                            1 if data.get("unsubscribed") in (True, 1, "1", "true", "yes", "on") else 0,
                            (data.get("notes") or "").strip(),
                            now,
                            now,
                            int(_tid),
                        ),
                    )
                    execute(conn, f"RELEASE SAVEPOINT {sp}")
                except Exception as exc:
                    print(f"⚠️  create_contact insert çakıştı ({email}): {exc}")
                    try:
                        execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
                    except Exception:
                        pass
                    conn.commit()
                    other = fetchone(conn, "SELECT id FROM mail_contacts WHERE LOWER(email) = ? LIMIT 1", (email,))
                    return jsonify({
                        "error": "Bu e-posta başka bir firmada kayıtlı, buraya eklenemedi.",
                        "id": other["id"] if other else None,
                    }), 409
                conn.commit()
                row = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (cid,))
                return jsonify({"contact": _contact_out(row)}), 201
            cid = insert_returning_id(
                conn,
                """
                INSERT INTO mail_contacts
                (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    email,
                    (data.get("phone") or "").strip(),
                    (data.get("name") or "").strip(),
                    tags,
                    (data.get("source") or "manual").strip() or "manual",
                    1 if data.get("unsubscribed") in (True, 1, "1", "true", "yes", "on") else 0,
                    (data.get("notes") or "").strip(),
                    now,
                    now,
                ),
            )
            from database import _table_columns
            cols = _table_columns(conn, "mail_contacts") or set()
            extra_sets = []
            extra_params = []
            if "lifecycle" in cols and data.get("lifecycle"):
                extra_sets.append("lifecycle = ?")
                extra_params.append((data.get("lifecycle") or "lead").strip().lower())
            if "crm_owner" in cols and "crm_owner" in data:
                extra_sets.append("crm_owner = ?")
                extra_params.append((data.get("crm_owner") or "").strip()[:120])
            if "verify_status" in cols and "verify_status" in data:
                extra_sets.append("verify_status = ?")
                extra_params.append((data.get("verify_status") or "").strip()[:40])
            if extra_sets:
                execute(
                    conn,
                    f"UPDATE mail_contacts SET {', '.join(extra_sets)} WHERE id = ?",
                    tuple(extra_params) + (cid,),
                )
            for t in _parse_tags(data.get("tags")):
                _ensure_tag(conn, t, now)
            _invalidate_mail_stats_cache()
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (cid,))
        return jsonify({"contact": _contact_out(row)}), 201

    @bp.route("/contacts/<int:contact_id>", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def get_contact(contact_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (contact_id,))
            if not row:
                return jsonify({"error": "Kontak bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kontak başka firmaya ait."}), 403
            except Exception:
                pass
        return jsonify({"contact": _contact_out(row)})

    @bp.route("/contacts/<int:contact_id>", methods=["PATCH"])
    @mail_perm(*MAIL_CRM)
    def update_contact(contact_id):
        data = request.get_json(silent=True) or {}
        now = iso(utcnow())
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (contact_id,))
            if not row:
                return jsonify({"error": "Kontak bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kontak başka firmaya ait."}), 403
            except Exception:
                pass
            email = (data.get("email") if "email" in data else row["email"] or "").strip().lower()
            if not email or not EMAIL_RE.match(email):
                return jsonify({"error": "Geçerli bir e-posta girin."}), 400
            # (tenant_id, email) composite unique — çakışma kontrolü de aynı tenant
            # kapsamında yapılmalı, yoksa başka firmanın kaydı burada engel olur.
            row_tid = row.get("tenant_id") if isinstance(row, dict) else None
            if row_tid:
                other = fetchone(
                    conn,
                    "SELECT id FROM mail_contacts WHERE LOWER(email) = ? AND id != ? AND tenant_id = ?",
                    (email, contact_id, int(row_tid)),
                )
            else:
                other = fetchone(
                    conn,
                    "SELECT id FROM mail_contacts WHERE LOWER(email) = ? AND id != ?",
                    (email, contact_id),
                )
            if other:
                return jsonify({"error": "Bu e-posta başka bir kontakta kayıtlı."}), 409
            tags = _tags_json(data.get("tags")) if "tags" in data else row["tags"]
            if "unsubscribed" in data:
                unsub = 1 if data.get("unsubscribed") in (True, 1, "1", "true", "yes", "on") else 0
            else:
                unsub = 1 if row["unsubscribed"] else 0

            # Opsiyonel CRM / verify kolonları (yoksa UPDATE'te atlanır)
            from database import _table_columns
            cols = _table_columns(conn, "mail_contacts") or set()
            sets = [
                "email = ?", "phone = ?", "name = ?", "tags = ?", "source = ?",
                "unsubscribed = ?", "notes = ?", "updated_at = ?",
            ]
            params = [
                email,
                (data.get("phone") if "phone" in data else row["phone"] or "").strip(),
                (data.get("name") if "name" in data else row["name"] or "").strip(),
                tags,
                (data.get("source") if "source" in data else row["source"] or "manual").strip() or "manual",
                unsub,
                (data.get("notes") if "notes" in data else row["notes"] or "").strip(),
                now,
            ]
            if "lifecycle" in cols and "lifecycle" in data:
                life = (data.get("lifecycle") or "lead").strip().lower() or "lead"
                sets.append("lifecycle = ?")
                params.append(life)
            if "crm_owner" in cols and "crm_owner" in data:
                sets.append("crm_owner = ?")
                params.append((data.get("crm_owner") or "").strip()[:120])
            if "verify_status" in cols and "verify_status" in data:
                sets.append("verify_status = ?")
                params.append((data.get("verify_status") or "").strip()[:40])
            if "verify_detail" in cols and "verify_detail" in data:
                sets.append("verify_detail = ?")
                params.append((data.get("verify_detail") or "").strip()[:240])
            params.append(contact_id)
            execute(
                conn,
                f"UPDATE mail_contacts SET {', '.join(sets)} WHERE id = ?",
                tuple(params),
            )
            if "tags" in data:
                for t in _parse_tags(data.get("tags")):
                    _ensure_tag(conn, t, now)
            _invalidate_mail_stats_cache()
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (contact_id,))
        return jsonify({"contact": _contact_out(row)})

    @bp.route("/contacts/<int:contact_id>", methods=["DELETE"])
    @mail_perm(*MAIL_CRM)
    def delete_contact(contact_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT id, tenant_id FROM mail_contacts WHERE id = ?", (contact_id,))
            if not row:
                return jsonify({"error": "Kontak bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kontak başka firmaya ait."}), 403
            except Exception:
                pass
            try:
                _delete_contacts_by_ids(conn, [contact_id])
                _invalidate_mail_stats_cache()
                conn.commit()
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                return jsonify({"error": f"Silme başarısız: {exc}"}), 500
        return jsonify({"ok": True})

    @bp.route("/contacts/import", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def import_contacts():
        data = request.get_json(silent=True) or {}
        raw_csv = data.get("csv") or ""
        default_tag = (data.get("tag") or "").strip()
        if not raw_csv.strip():
            return jsonify({"error": "CSV içeriği boş."}), 400
        reader = csv.DictReader(io.StringIO(raw_csv))
        now = iso(utcnow())
        created = 0
        updated = 0
        skipped = 0
        with closing(get_db()) as conn:
            from database import _table_columns
            cols = _table_columns(conn, "mail_contacts") or set()
            has_tenant = "tenant_id" in cols
            tid = None
            if has_tenant:
                try:
                    from mail_tenant import current_tenant_id
                    tid = current_tenant_id() or 1
                except Exception:
                    tid = 1
            for row in reader:
                email = _extract_email_from_row(row)
                if not email:
                    skipped += 1
                    continue
                name = (row.get("name") or row.get("Name") or "").strip()
                phone = (row.get("phone") or row.get("Phone") or row.get("tel") or "").strip()
                tags = _parse_tags(row.get("tags") or row.get("tag") or "")
                if default_tag and default_tag not in tags:
                    tags.append(default_tag)
                if has_tenant:
                    existing = fetchone(
                        conn,
                        "SELECT id, tags FROM mail_contacts WHERE LOWER(email) = ? AND tenant_id = ?",
                        (email, tid),
                    )
                else:
                    existing = fetchone(conn, "SELECT id, tags FROM mail_contacts WHERE LOWER(email) = ?", (email,))
                if existing:
                    merged = list(dict.fromkeys(_parse_tags(existing["tags"]) + tags))
                    execute(
                        conn,
                        """
                        UPDATE mail_contacts SET name = COALESCE(NULLIF(?, ''), name),
                            phone = COALESCE(NULLIF(?, ''), phone),
                            tags = ?, updated_at = ?
                        WHERE id = ?
                        """,
                        (name, phone, _tags_json(merged), now, existing["id"]),
                    )
                    updated += 1
                elif has_tenant:
                    # SAVEPOINT: eski global UNIQUE(email) hâlâ devredeyse başka firmanın
                    # kaydıyla çakışabilir — tüm CSV işlemi patlamasın, satırı güncelleme
                    # olarak say.
                    sp = "sp_csv_contact"
                    try:
                        execute(conn, f"SAVEPOINT {sp}")
                        insert_returning_id(
                            conn,
                            """
                            INSERT INTO mail_contacts
                            (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at, tenant_id)
                            VALUES (?, ?, ?, ?, ?, 0, '', ?, ?, ?)
                            """,
                            (email, phone, name, _tags_json(tags), "csv", now, now, tid),
                        )
                        execute(conn, f"RELEASE SAVEPOINT {sp}")
                        created += 1
                    except Exception as exc:
                        print(f"⚠️  csv contact insert çakıştı ({email}): {exc}")
                        try:
                            execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
                        except Exception:
                            pass
                        updated += 1
                else:
                    insert_returning_id(
                        conn,
                        """
                        INSERT INTO mail_contacts
                        (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, 0, '', ?, ?)
                        """,
                        (email, phone, name, _tags_json(tags), "csv", now, now),
                    )
                    created += 1
                for t in tags:
                    _ensure_tag(conn, t, now)
            conn.commit()
        return jsonify({"created": created, "updated": updated, "skipped": skipped})

    @bp.route("/contacts/import/start", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def start_import_job():
        """Büyük liste (yüz binler / milyonlar) için dosya yükleyip arka planda işler."""
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "CSV veya XLSX dosyası seç."}), 400
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".csv", ".xlsx", ".xlsm"):
            return jsonify({"error": "Sadece .csv veya .xlsx dosyası yükleyebilirsin."}), 400
        if request.content_length and request.content_length > IMPORT_MAX_BYTES:
            return jsonify({"error": "Dosya çok büyük (5GB üstü). Bölüp tekrar dene."}), 400
        tag = (request.form.get("tag") or "").strip()
        _ensure_import_dir()
        now = iso(utcnow())
        try:
            from mail_tenant import current_tenant_id as _cur_tid_early
            _tid_job = _cur_tid_early()
        except Exception:
            _tid_job = None
        with closing(get_db()) as conn:
            from database import _table_columns
            job_cols = _table_columns(conn, "mail_import_jobs") or set()
            if "tenant_id" in job_cols:
                job_id = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_import_jobs
                    (filename, tag, status, total_rows, processed_rows, upserted_count, inserted_count, updated_count, skipped_count, error, created_at, updated_at, tenant_id)
                    VALUES (?, ?, 'pending', 0, 0, 0, 0, 0, 0, '', ?, ?, ?)
                    """,
                    (file.filename, tag, now, now, _tid_job),
                )
            else:
                job_id = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_import_jobs
                    (filename, tag, status, total_rows, processed_rows, upserted_count, inserted_count, updated_count, skipped_count, error, created_at, updated_at)
                    VALUES (?, ?, 'pending', 0, 0, 0, 0, 0, 0, '', ?, ?)
                    """,
                    (file.filename, tag, now, now),
                )
            conn.commit()
        path = _import_job_path(job_id, file.filename)
        try:
            file.save(path)
        except Exception as exc:
            err = f"Dosya kaydedilemedi: {str(exc)[:300]}"
            with closing(get_db()) as conn:
                execute(
                    conn,
                    "UPDATE mail_import_jobs SET status = 'error', error = ?, updated_at = ? WHERE id = ?",
                    (err, iso(utcnow()), job_id),
                )
                conn.commit()
            try:
                if os.path.isfile(path):
                    os.remove(path)
            except Exception:
                pass
            return jsonify({"error": err}), 500
        threading.Thread(
            target=_run_import_job,
            args=(job_id, path, tag, _tid_job),
            daemon=True,
        ).start()
        return jsonify({"job_id": job_id, "status": "pending", "tenant_id": _tid_job}), 202

    @bp.route("/contacts/import/emails", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def import_emails_paste():
        """Yapıştırılan e-posta listesi — File/CSV olmadan doğrudan upsert (küçük-orta listeler)."""
        data = request.get_json(silent=True) or {}
        raw_list = data.get("emails")
        if isinstance(raw_list, str):
            raw_list = re.split(r"[\s,;]+", raw_list)
        if not isinstance(raw_list, list):
            return jsonify({"error": "emails listesi gerekli."}), 400
        tag = (data.get("tag") or "").strip()
        cleaned = []
        seen = set()
        invalid = 0
        for raw in raw_list:
            em = _normalize_email_candidate(raw)
            if not em:
                if str(raw or "").strip():
                    invalid += 1
                continue
            if em in seen:
                continue
            seen.add(em)
            cleaned.append(em)
        if not cleaned:
            return jsonify({
                "error": "Geçerli e-posta bulunamadı.",
                "created": 0,
                "updated": 0,
                "skipped": invalid,
            }), 400
        # Çok büyük yapıştırmada async job'a düş (dosyasız — temp csv)
        if len(cleaned) > 20000:
            return jsonify({
                "error": "20.000 üzeri için CSV dosyası yükle (daha stabil).",
                "count": len(cleaned),
            }), 400
        now = iso(utcnow())
        try:
            from mail_tenant import current_tenant_id

            tid = current_tenant_id()
        except Exception:
            tid = None
        batch = [(em, "") for em in cleaned]
        with closing(get_db()) as conn:
            try:
                upserted, inserted, updated = _bulk_upsert_contacts(
                    conn, batch, tag, now, tenant_id=tid
                )
                if tag:
                    _ensure_tag(conn, tag, now)
                conn.commit()
                _invalidate_mail_stats_cache()
            except Exception as exc:
                try:
                    conn.rollback()
                except Exception:
                    pass
                return jsonify({"error": f"Kayıt hatası: {str(exc)[:300]}"}), 500
        return jsonify({
            "ok": True,
            "created": inserted,
            "updated": updated,
            "upserted": upserted,
            "skipped": invalid,
            "tenant_id": tid,
            "message": f"{inserted} yeni · {updated} güncellendi"
            + (f" · {invalid} atlandı" if invalid else ""),
        })

    @bp.route("/contacts/import/status/<int:job_id>", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def import_job_status(job_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_import_jobs WHERE id = ?", (job_id,))
            if not row:
                return jsonify({"error": "İş bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu iş başka firmaya ait."}), 403
            except Exception:
                pass
            job = _reconcile_stale_import_job(conn, row)
        return jsonify({"job": job})

    @bp.route("/contacts/import/jobs", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def list_import_jobs():
        with closing(get_db()) as conn:
            try:
                from database import _table_columns
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                job_cols = _table_columns(conn, "mail_import_jobs") or set()
            except Exception:
                _tid = None
                job_cols = set()
            if _tid and "tenant_id" in job_cols:
                raw = fetchall(
                    conn,
                    "SELECT * FROM mail_import_jobs WHERE tenant_id = ? ORDER BY id DESC LIMIT 30",
                    (int(_tid),),
                )
            else:
                raw = fetchall(conn, "SELECT * FROM mail_import_jobs ORDER BY id DESC LIMIT 30")
            jobs = []
            for row in raw:
                jobs.append(_reconcile_stale_import_job(conn, row))
        return jsonify({"jobs": jobs})

    @bp.route("/contacts/import/cancel/<int:job_id>", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def cancel_import_job(job_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_import_jobs WHERE id = ?", (job_id,))
            if not row:
                return jsonify({"error": "İş bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu iş başka firmaya ait."}), 403
            except Exception:
                pass
            status = row["status"]
            if status in ("done", "error", "cancelled", "cancelling"):
                return jsonify({"error": f"İş zaten sonlanmış ya da iptal ediliyor ({status})."}), 400
            execute(
                conn,
                "UPDATE mail_import_jobs SET status = 'cancelling', updated_at = ? WHERE id = ?",
                (iso(utcnow()), job_id),
            )
            conn.commit()
        return jsonify({"ok": True, "status": "cancelling"})

    @bp.route("/contacts/scrub/start", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def start_scrub():
        """Liste temizleme: syntax/MX/SMTP ping → invalid’leri suppression’a al."""
        from mail_scrub import job_public, start_scrub_job

        data = request.get_json(silent=True) or {}
        # tag_filters[] veya virgüllü tag_filter — birden fazla etiket = birleşim (OR)
        if data.get("tag_filters") is not None:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filters"))
        else:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filter"))
        contact_ids = data.get("contact_ids") or []
        if not isinstance(contact_ids, list):
            contact_ids = []
        try:
            contact_ids = [int(x) for x in contact_ids if x is not None and str(x).strip() != ""]
        except (TypeError, ValueError):
            return jsonify({"error": "contact_ids geçersiz."}), 400
        scope = "selected" if contact_ids else ("filter" if tag_filter else "all")
        _tid = None
        try:
            from mail_tenant import current_tenant_id
            _tid = current_tenant_id()
        except Exception:
            _tid = None
        try:
            job_id = start_scrub_job(
                tag_filter=tag_filter,
                contact_ids=contact_ids,
                scope=scope,
                tenant_id=_tid,
            )
        except RuntimeError as exc:
            return jsonify({"error": str(exc)}), 409
        except Exception as exc:
            return jsonify({"error": f"Temizlik başlatılamadı: {exc}"}), 500
        with closing(get_db()) as conn:
            from mail_ops import audit
            audit(conn, request.headers.get("X-Admin-User") or "admin", "scrub_start", f"job={job_id} scope={scope}")
            try:
                conn.commit()
            except Exception:
                pass
            row = fetchone(conn, "SELECT * FROM mail_scrub_jobs WHERE id = ?", (job_id,))
        return jsonify({"ok": True, "job": job_public(row)}), 201

    @bp.route("/contacts/scrub/status/<int:job_id>", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def scrub_status(job_id):
        from mail_scrub import job_public

        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_scrub_jobs WHERE id = ?", (job_id,))
            if not row:
                return jsonify({"error": "İş bulunamadı."}), 404
            row = _row(row)
            try:
                from database import _table_columns
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and "tenant_id" in (_table_columns(conn, "mail_scrub_jobs") or set()) and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu iş başka firmaya ait."}), 403
            except Exception:
                pass
        return jsonify({"job": job_public(row)})

    @bp.route("/contacts/scrub/latest", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def scrub_latest():
        from mail_scrub import job_public

        with closing(get_db()) as conn:
            try:
                from database import _table_columns
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                cols = _table_columns(conn, "mail_scrub_jobs") or set()
            except Exception:
                _tid = None
                cols = set()
            if _tid and "tenant_id" in cols:
                row = fetchone(
                    conn,
                    "SELECT * FROM mail_scrub_jobs WHERE tenant_id = ? ORDER BY id DESC LIMIT 1",
                    (int(_tid),),
                )
            else:
                row = fetchone(conn, "SELECT * FROM mail_scrub_jobs ORDER BY id DESC LIMIT 1")
        return jsonify({"job": job_public(row) if row else None})

    @bp.route("/contacts/scrub/cancel/<int:job_id>", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def cancel_scrub(job_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_scrub_jobs WHERE id = ?", (job_id,))
            if not row:
                return jsonify({"error": "İş bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu iş başka firmaya ait."}), 403
            except Exception:
                pass
            status = row["status"]
            if status in ("done", "error", "cancelled", "cancelling"):
                return jsonify({"error": f"İş zaten sonlanmış ({status})."}), 400
            execute(
                conn,
                "UPDATE mail_scrub_jobs SET status = 'cancelling', updated_at = ? WHERE id = ?",
                (iso(utcnow()), job_id),
            )
            conn.commit()
        return jsonify({"ok": True, "status": "cancelling"})

    @bp.route("/contacts/scrub/force-reset", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def scrub_force_reset():
        """Takılı pending/running scrub işlerini iptal et — operatör kurtarma."""
        from mail_scrub import ensure_mail_scrub_schema

        with closing(get_db()) as conn:
            ensure_mail_scrub_schema(conn)
            rows = fetchall(
                conn,
                "SELECT id FROM mail_scrub_jobs WHERE status IN ('pending', 'running', 'cancelling')",
            ) or []
            now = iso(utcnow())
            for row in rows:
                execute(
                    conn,
                    "UPDATE mail_scrub_jobs SET status = 'cancelled', error = ?, updated_at = ? WHERE id = ?",
                    ("Operatör force-reset", now, int(row["id"])),
                )
            conn.commit()
        return jsonify({"ok": True, "cancelled": len(rows)})

    @bp.route("/tags", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def list_tags():
        with closing(get_db()) as conn:
            rows = _rows(fetchall(conn, "SELECT * FROM mail_contact_tags ORDER BY name ASC"))
        return jsonify({"tags": rows})

    # ── Gerçek CRM (ilişki) — Mail Rehber'den ayrı ─────────────
    @bp.route("/relations/overview", methods=["GET"])
    @mail_perm(*MAIL_REL)
    def relations_overview():
        from mail_crm import crm_overview, ensure_mail_crm_schema

        with closing(get_db()) as conn:
            ensure_mail_crm_schema(conn)
            data = crm_overview(conn)
        return jsonify(data)

    @bp.route("/relations/pipeline", methods=["GET"])
    @mail_perm(*MAIL_REL)
    def relations_pipeline():
        from mail_crm import list_crm_pipeline

        lifecycle = (request.args.get("lifecycle") or "").strip()
        q = (request.args.get("q") or "").strip()
        limit = request.args.get("limit") or 80
        with closing(get_db()) as conn:
            rows = list_crm_pipeline(conn, lifecycle=lifecycle, q=q, limit=limit)
        return jsonify({"contacts": rows, "count": len(rows)})

    @bp.route("/relations/contacts/<int:contact_id>", methods=["GET"])
    @mail_perm(*MAIL_REL)
    def relations_contact(contact_id):
        from mail_crm import get_contact_crm, refresh_contact_crm

        with closing(get_db()) as conn:
            if (request.args.get("refresh") or "") in ("1", "true", "yes"):
                refresh_contact_crm(conn, contact_id, apply_lifecycle=True)
                try:
                    conn.commit()
                except Exception:
                    pass
            data = get_contact_crm(conn, contact_id)
            if not data:
                return jsonify({"error": "Kontak bulunamadı."}), 404
        return jsonify({"contact": data})

    @bp.route("/relations/contacts/<int:contact_id>/lifecycle", methods=["PATCH"])
    @mail_perm(*MAIL_REL)
    def relations_set_lifecycle(contact_id):
        from mail_crm import get_contact_crm, set_lifecycle

        data = request.get_json(silent=True) or {}
        try:
            with closing(get_db()) as conn:
                set_lifecycle(
                    conn, contact_id,
                    data.get("lifecycle"),
                    owner=data.get("crm_owner") if "crm_owner" in data else None,
                )
                conn.commit()
                contact = get_contact_crm(conn, contact_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"ok": True, "contact": contact})

    @bp.route("/relations/contacts/<int:contact_id>/notes", methods=["POST"])
    @mail_perm(*MAIL_REL)
    def relations_add_note(contact_id):
        from mail_crm import add_note, get_contact_crm

        data = request.get_json(silent=True) or {}
        author = request.headers.get("X-Admin-User") or data.get("author") or "admin"
        try:
            with closing(get_db()) as conn:
                nid = add_note(conn, contact_id, data.get("body"), author=author)
                conn.commit()
                contact = get_contact_crm(conn, contact_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"ok": True, "note_id": nid, "contact": contact}), 201

    @bp.route("/relations/contacts/<int:contact_id>/tasks", methods=["POST"])
    @mail_perm(*MAIL_REL)
    def relations_add_task(contact_id):
        from mail_crm import add_task, get_contact_crm

        data = request.get_json(silent=True) or {}
        author = request.headers.get("X-Admin-User") or data.get("author") or "admin"
        try:
            with closing(get_db()) as conn:
                tid = add_task(
                    conn, contact_id,
                    data.get("title"),
                    due_at=data.get("due_at"),
                    author=author,
                )
                conn.commit()
                contact = get_contact_crm(conn, contact_id)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"ok": True, "task_id": tid, "contact": contact}), 201

    @bp.route("/relations/tasks/<int:task_id>", methods=["PATCH"])
    @mail_perm(*MAIL_REL)
    def relations_patch_task(task_id):
        from mail_crm import set_task_status

        data = request.get_json(silent=True) or {}
        try:
            with closing(get_db()) as conn:
                set_task_status(conn, task_id, data.get("status") or "done")
                conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"ok": True})

    @bp.route("/relations/recompute", methods=["POST"])
    @mail_perm(*MAIL_REL)
    def relations_recompute():
        from mail_crm import recompute_scores_batch

        data = request.get_json(silent=True) or {}
        limit = data.get("limit") or 300
        with closing(get_db()) as conn:
            n = recompute_scores_batch(conn, limit=limit)
        return jsonify({"ok": True, "updated": n, "message": f"{n} kontak skoru güncellendi"})

    @bp.route("/crm/sync-smartico", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def sync_smartico_segments():
        """Smartico'daki kayıt/FTD verisini afp1 (sub-id) üzerinden contact_id'ye
        eşleştirip CRM'i otomatik etiketler: uye_oldu / ftd_yapti / ftd_yok.
        """
        now = iso(utcnow())
        with closing(get_db()) as conn:
            affiliate_id = (get_mail_setting(conn, "smartico_affiliate_id", "") or "").strip()
            subid_param = (get_mail_setting(conn, "smartico_subid_param", "afp1") or "afp1").strip() or "afp1"
            if not affiliate_id:
                return jsonify({"error": "Önce Ayarlar'dan Smartico Affiliate ID gir."}), 400
            if not smartico_api.is_configured(conn):
                return jsonify({"error": "Smartico API anahtarı tanımlı değil (Ayarlar → Smartico CRM eşleştirme)."}), 400

            result = smartico_api.fetch_subid_conversions(conn, affiliate_id, subid_param)
            if result.get("error"):
                return jsonify({"error": result["error"]}), 400

            matched = 0
            unmatched = 0
            tagged_uye = 0
            tagged_ftd = 0
            tagged_no_ftd = 0
            for row in result["rows"]:
                subid = row.get("subid") or ""
                try:
                    contact_id = int(subid)
                except (TypeError, ValueError):
                    unmatched += 1
                    continue
                contact = fetchone(conn, "SELECT id FROM mail_contacts WHERE id = ?", (contact_id,))
                if not contact:
                    unmatched += 1
                    continue
                matched += 1
                if row.get("registration_count", 0) > 0:
                    _tag_contact(conn, contact_id, "uye_oldu", now)
                    tagged_uye += 1
                    if row.get("ftd_count", 0) > 0:
                        _tag_contact(conn, contact_id, "ftd_yapti", now)
                        _untag_contact(conn, contact_id, "ftd_yok", now)
                        tagged_ftd += 1
                    else:
                        _tag_contact(conn, contact_id, "ftd_yok", now)
                        tagged_no_ftd += 1
            conn.commit()
        return jsonify({
            "ok": True,
            "matched": matched,
            "unmatched": unmatched,
            "tagged_uye_oldu": tagged_uye,
            "tagged_ftd_yapti": tagged_ftd,
            "tagged_ftd_yok": tagged_no_ftd,
            "message": f"{matched} contact eşleşti · {tagged_uye} üye oldu · {tagged_ftd} FTD yaptı · {tagged_no_ftd} FTD yok",
        })

    @bp.route("/crm/smartico-players", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def smartico_players_report():
        """Mail {{link:sc:…}} tıklayıp kayıt olan üyelerin Smartico detayları.

        Smartico: username, registration_id, FTD/yatırım/çekim/bonus.
        Mail Rehber: e-posta, ad, telefon, etiketler (afp1 = contact_id eşleşmesi).
        """
        period = (request.args.get("period") or "30days").strip() or "30days"
        force = (request.args.get("force") or "").strip() in ("1", "true", "yes")
        q = (request.args.get("q") or "").strip().lower()
        only_matched = (request.args.get("matched") or "").strip() in ("1", "true", "yes")
        only_ftd = (request.args.get("ftd") or "").strip() in ("1", "true", "yes")

        with closing(get_db()) as conn:
            affiliate_id = (get_mail_setting(conn, "smartico_affiliate_id", "") or "").strip()
            subid_param = (get_mail_setting(conn, "smartico_subid_param", "afp1") or "afp1").strip() or "afp1"
            if not affiliate_id:
                return jsonify({
                    "error": "affiliate_id_missing",
                    "message": "Önce Ayarlar → Smartico CRM eşleştirme: Affiliate ID gir.",
                    "rows": [],
                    "summary": {},
                }), 400
            if not smartico_api.is_configured(conn):
                return jsonify({
                    "error": "not_configured",
                    "message": "Smartico API anahtarı yok (Ayarlar → Smartico CRM eşleştirme).",
                    "rows": [],
                    "summary": {},
                }), 400

            result = smartico_api.fetch_mailing_players(
                conn, affiliate_id, subid_param, period=period, force=force,
            )
            if result.get("error") and not result.get("rows"):
                return jsonify({
                    "error": result["error"],
                    "message": result["error"],
                    "rows": [],
                    "summary": result.get("summary") or {},
                    "currency": result.get("currency") or "",
                    "source": result.get("source"),
                }), 400

            rows = result.get("rows") or []
            contact_ids = []
            for row in rows:
                try:
                    cid = int(str(row.get("subid") or "").strip())
                except (TypeError, ValueError):
                    continue
                if cid > 0:
                    contact_ids.append(cid)
            contact_ids = list(dict.fromkeys(contact_ids))
            contacts = {}
            # Postgres/SQLite IN — parçalı
            chunk = 400
            for i in range(0, len(contact_ids), chunk):
                part = contact_ids[i:i + chunk]
                placeholders = ",".join("?" * len(part))
                found = fetchall(
                    conn,
                    f"SELECT id, email, name, phone, tags FROM mail_contacts WHERE id IN ({placeholders})",
                    tuple(part),
                )
                for c in found or []:
                    contacts[int(c["id"])] = dict(c)

            enriched = []
            for row in rows:
                out = dict(row)
                contact = None
                try:
                    cid = int(str(row.get("subid") or "").strip())
                    contact = contacts.get(cid)
                except (TypeError, ValueError):
                    cid = None
                if contact:
                    tags_raw = contact.get("tags") or "[]"
                    try:
                        tags = json.loads(tags_raw) if isinstance(tags_raw, str) else (tags_raw or [])
                    except (TypeError, ValueError, json.JSONDecodeError):
                        tags = []
                    full_name = (contact.get("name") or "").strip()
                    parts = full_name.split(None, 1) if full_name else []
                    out["contact_id"] = contact["id"]
                    out["email"] = contact.get("email") or ""
                    out["name"] = full_name
                    out["first_name"] = parts[0] if parts else ""
                    out["last_name"] = parts[1] if len(parts) > 1 else ""
                    out["phone"] = contact.get("phone") or ""
                    out["tags"] = tags
                    out["matched"] = True
                else:
                    out["contact_id"] = cid
                    out["email"] = ""
                    out["name"] = ""
                    out["first_name"] = ""
                    out["last_name"] = ""
                    out["phone"] = ""
                    out["tags"] = []
                    out["matched"] = False
                enriched.append(out)

            if only_matched:
                enriched = [r for r in enriched if r.get("matched")]
            if only_ftd:
                enriched = [r for r in enriched if (r.get("ftd_count") or 0) > 0]
            if q:
                def _hit(r):
                    blob = " ".join([
                        str(r.get("username") or ""),
                        str(r.get("email") or ""),
                        str(r.get("name") or ""),
                        str(r.get("subid") or ""),
                        str(r.get("ext_customer_id") or ""),
                        str(r.get("registration_id") or ""),
                        str(r.get("phone") or ""),
                    ]).lower()
                    return q in blob
                enriched = [r for r in enriched if _hit(r)]

            matched_n = sum(1 for r in enriched if r.get("matched"))
            return jsonify({
                "ok": True,
                "rows": enriched,
                "summary": result.get("summary") or {},
                "currency": result.get("currency") or "",
                "source": result.get("source"),
                "group_by": result.get("group_by"),
                "affiliate_id": affiliate_id,
                "subid_param": subid_param,
                "period": period,
                "matched_in_crm": matched_n,
                "total": len(enriched),
                "error": result.get("error"),
            })

    @bp.route("/tags", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def create_tag():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Etiket adı gerekli."}), 400
        now = iso(utcnow())
        with closing(get_db()) as conn:
            _ensure_tag(conn, name, now)
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_contact_tags WHERE name = ?", (name,))
        _invalidate_mail_stats_cache()
        return jsonify({"tag": _row(row)}), 201

    @bp.route("/tags/delete", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def delete_tag():
        """Etiketi sil. force=true ise kontaktaki etiketleri de temizler."""
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        force = bool(data.get("force"))
        if not name:
            return jsonify({"error": "Etiket adı gerekli."}), 400
        try:
            with closing(get_db()) as conn:
                result = _delete_tag(conn, name, force=force)
                conn.commit()
        except ValueError as exc:
            return jsonify({"error": str(exc), "needs_force": True}), 400
        except Exception as exc:
            return jsonify({"error": f"Etiket silinemedi: {exc}"}), 400
        _invalidate_mail_stats_cache()
        msg = f"«{name}» silindi"
        if result.get("stripped"):
            msg += f" · {result['stripped']} kontaktan kaldırıldı"
        result["message"] = msg
        return jsonify(result)

    @bp.route("/tags/cleanup", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def cleanup_tags():
        """0 kontak kalan tüm etiketleri sil."""
        with closing(get_db()) as conn:
            deleted = _cleanup_empty_tags(conn)
            conn.commit()
        _invalidate_mail_stats_cache()
        return jsonify({
            "ok": True,
            "deleted": deleted,
            "count": len(deleted),
            "message": (
                f"{len(deleted)} boş etiket silindi: {', '.join(deleted)}"
                if deleted else "Silinecek boş etiket yok"
            ),
        })

    @bp.route("/contacts/tags/bulk", methods=["POST"])
    @mail_perm(*MAIL_CRM)
    def bulk_contact_tags():
        """Toplu etiket ekle / kaldır / taşı (segment kaydırma)."""
        data = request.get_json(silent=True) or {}
        try:
            from mail_tenant import current_tenant_id
            _tid = current_tenant_id()
            with closing(get_db()) as conn:
                result = _bulk_retag_contacts(
                    conn,
                    action=data.get("action"),
                    from_tag=data.get("from_tag") or "",
                    to_tag=data.get("to_tag") or "",
                    contact_ids=data.get("contact_ids"),
                    match_tag=data.get("match_tag") or "",
                    limit=data.get("limit"),
                    tenant_id=_tid,
                )
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        except Exception as exc:
            return jsonify({"error": f"Toplu etiket işlemi başarısız: {exc}"}), 400
        action = result.get("action")
        msg = f"{result.get('updated', 0)} kontak güncellendi"
        if action == "move":
            msg = f"{result.get('updated', 0)} kontak «{result.get('from_tag')}» → «{result.get('to_tag')}» taşındı"
        elif action == "add":
            msg = f"{result.get('updated', 0)} kontağa «{result.get('to_tag')}» eklendi"
        elif action == "remove":
            msg = f"{result.get('updated', 0)} kontaktan «{result.get('from_tag')}» kaldırıldı"
        cleaned = result.get("cleaned_tags") or []
        if cleaned:
            msg += f" · boş etiket silindi: {', '.join(cleaned)}"
        result["message"] = msg
        return jsonify(result)

    # ── Templates ──────────────────────────────────────────────
    @bp.route("/templates", methods=["GET"])
    @mail_perm(*MAIL_TPL)
    def list_templates():
        with closing(get_db()) as conn:
            _tid = None
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if _tid:
                rows = _rows(fetchall(
                    conn,
                    "SELECT * FROM mail_templates WHERE tenant_id = ? ORDER BY id DESC",
                    (int(_tid),),
                ))
            else:
                rows = _rows(fetchall(conn, "SELECT * FROM mail_templates ORDER BY id DESC"))
        return jsonify({"templates": rows})

    @bp.route("/templates/reseed", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def reseed_templates():
        """Makrobet + Bizzo HTML şablonlarını günceller / eksikleri ekler."""
        from mail_template_seeds import seed_makrobet_mail_templates
        from mail_template_seeds_bizzo import seed_bizzo_mail_templates

        with closing(get_db()) as conn:
            result = seed_makrobet_mail_templates(
                conn, force_missing=True, overwrite=True, allow_when_skipped=True
            )
            bizzo = seed_bizzo_mail_templates(
                conn, force_missing=True, overwrite=True, allow_when_skipped=True
            )
        if not isinstance(result, dict):
            result = {"added": int(result or 0), "updated": 0}
        if not isinstance(bizzo, dict):
            bizzo = {"added": int(bizzo or 0), "updated": 0}
        added = int(result.get("added") or 0) + int(bizzo.get("added") or 0)
        updated = int(result.get("updated") or 0) + int(bizzo.get("updated") or 0)
        if added or updated:
            msg = f"{updated} HTML güncellendi · {added} yeni (Makrobet + Bizzo)"
        else:
            msg = "Değişiklik yok"
        return jsonify({"ok": True, "added": added, "updated": updated, "bizzo": bizzo, "message": msg})

    @bp.route("/templates/wipe-all", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def wipe_all_templates():
        """Tüm şablonları siler; otomatik seed tekrar eklemez."""
        from mail_template_wipe import wipe_all_mail_templates

        with closing(get_db()) as conn:
            try:
                result = wipe_all_mail_templates(conn)
            except Exception as exc:
                return jsonify({"error": str(exc)}), 500
        return jsonify({"ok": True, **result, "message": f"{result.get('deleted', 0)} şablon silindi"})

    @bp.route("/templates/seed-davet-deneme", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def seed_davet_deneme_template():
        """Tek Makro davet şablonu: 3.000 TL deneme + %100 kayıp (wipe skip açıkken de)."""
        from mail_template_seeds_v2026 import seed_davet_deneme_kayip_template

        with closing(get_db()) as conn:
            result = seed_davet_deneme_kayip_template(conn, overwrite=True)
        action = result.get("action") or "kept"
        if action == "added":
            msg = f"{result.get('name')} eklendi"
        elif action == "updated":
            msg = f"{result.get('name')} güncellendi"
        else:
            msg = f"{result.get('name')} zaten mevcut"
        return jsonify({**result, "message": msg})

    @bp.route("/templates/seed-steril-ayricaliklar", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def seed_steril_ayricaliklar_route():
        """Steril Makro özellik şablonu (Betroz tarzı kart grid, wipe skip açıkken de)."""
        from mail_template_seeds_v2026 import seed_steril_ayricaliklar_template

        with closing(get_db()) as conn:
            result = seed_steril_ayricaliklar_template(conn, overwrite=True)
        action = result.get("action") or "kept"
        if action == "added":
            msg = f"{result.get('name')} eklendi"
        elif action == "updated":
            msg = f"{result.get('name')} güncellendi"
        else:
            msg = f"{result.get('name')} zaten mevcut"
        return jsonify({**result, "message": msg})

    @bp.route("/templates/seed-gorselsiz-ayricaliklar", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def seed_gorselsiz_ayricaliklar_route():
        """Görselsiz MakroVip davet (hero yok; wipe skip açıkken de)."""
        from mail_template_seeds_v2026 import seed_gorselsiz_ayricaliklar_template

        with closing(get_db()) as conn:
            result = seed_gorselsiz_ayricaliklar_template(conn, overwrite=True)
        action = result.get("action") or "kept"
        if action == "added":
            msg = f"{result.get('name')} eklendi"
        elif action == "updated":
            msg = f"{result.get('name')} güncellendi"
        else:
            msg = f"{result.get('name')} zaten mevcut"
        return jsonify({**result, "message": msg})

    @bp.route("/templates/seed-bizzo-davet-1x", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def seed_bizzo_davet_1x_route():
        """Bizzo davet: 1x / sınırsız çekim (wipe skip açıkken de)."""
        from mail_template_seeds_bizzo import seed_bizzo_davet_1x_sinirsiz_template

        with closing(get_db()) as conn:
            result = seed_bizzo_davet_1x_sinirsiz_template(conn, overwrite=True)
        action = result.get("action") or "kept"
        if action == "added":
            msg = f"{result.get('name')} eklendi"
        elif action == "updated":
            msg = f"{result.get('name')} güncellendi"
        else:
            msg = f"{result.get('name')} zaten mevcut"
        return jsonify({**result, "message": msg})

    @bp.route("/templates", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def create_template():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Şablon adı gerekli."}), 400
        now = iso(utcnow())
        text_body = data.get("text_body") or ""
        html_body = data.get("html_body") or ""
        if not html_body.strip() and text_body.strip():
            html_body = _plain_to_html(text_body)
        with closing(get_db()) as conn:
            from database import _table_columns
            _tid = None
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            tpl_cols = _table_columns(conn, "mail_templates") or set()
            if _tid and "tenant_id" in tpl_cols:
                tid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_templates
                    (name, subject, html_body, text_body, created_at, updated_at, tenant_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        name,
                        (data.get("subject") or "").strip(),
                        html_body,
                        text_body,
                        now,
                        now,
                        int(_tid),
                    ),
                )
            else:
                tid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_templates (name, subject, html_body, text_body, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        name,
                        (data.get("subject") or "").strip(),
                        html_body,
                        text_body,
                        now,
                        now,
                    ),
                )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_templates WHERE id = ?", (tid,))
        return jsonify({"template": _row(row)}), 201

    @bp.route("/templates/<int:template_id>", methods=["PATCH"])
    @mail_perm(*MAIL_TPL)
    def update_template(template_id):
        data = request.get_json(silent=True) or {}
        now = iso(utcnow())
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_templates WHERE id = ?", (template_id,))
            if not row:
                return jsonify({"error": "Şablon bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu şablon başka firmaya ait."}), 403
            except Exception:
                pass
            text_body = data.get("text_body") if "text_body" in data else row["text_body"] or ""
            html_body = data.get("html_body") if "html_body" in data else row["html_body"] or ""
            # Basit yazı kaydı: html boşsa veya sync_html istenirse üret
            if data.get("sync_html_from_text") or (not (html_body or "").strip() and (text_body or "").strip()):
                html_body = _plain_to_html(text_body)
            execute(
                conn,
                """
                UPDATE mail_templates SET name = ?, subject = ?, html_body = ?, text_body = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    (data.get("name") if "name" in data else row["name"]).strip(),
                    (data.get("subject") if "subject" in data else row["subject"] or "").strip(),
                    html_body,
                    text_body,
                    now,
                    template_id,
                ),
            )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_templates WHERE id = ?", (template_id,))
        return jsonify({"template": _row(row)})

    @bp.route("/templates/<int:template_id>/test-send", methods=["POST"])
    @mail_perm(*MAIL_TPL)
    def test_send_template(template_id):
        data = request.get_json(silent=True) or {}
        email = (data.get("email") or "").strip().lower()
        if not email or not EMAIL_RE.match(email):
            return jsonify({"error": "Geçerli test e-postası girin."}), 400
        domain_id = data.get("domain_id")
        with closing(get_db()) as conn:
            tpl = fetchone(conn, "SELECT * FROM mail_templates WHERE id = ?", (template_id,))
            if not tpl:
                return jsonify({"error": "Şablon bulunamadı."}), 404
            if not domain_id:
                raw = get_mail_setting(conn, "default_domain_id", "") or ""
                try:
                    domain_id = int(raw) if raw else None
                except ValueError:
                    domain_id = None
            if not domain_id:
                first = fetchone(conn, "SELECT id FROM mail_domains ORDER BY id ASC LIMIT 1")
                domain_id = first["id"] if first else None
            contact = {"name": data.get("name") or "Test", "email": email, "phone": ""}
            # Upsert test contact lightly
            existing = fetchone(conn, "SELECT id FROM mail_contacts WHERE LOWER(email) = ?", (email,))
            now = iso(utcnow())
            if existing:
                contact_id = existing["id"]
            else:
                contact_id = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_contacts
                    (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at)
                    VALUES (?, '', ?, ?, 'test', 0, '', ?, ?)
                    """,
                    (email, contact["name"], _tags_json(["test"]), now, now),
                )
            subject = _render_template(tpl["subject"], contact)
            html_body = _render_template(tpl["html_body"] or _plain_to_html(tpl["text_body"] or ""), contact)
            text_body = _render_template(tpl["text_body"] or "", contact)
            from mail_delivery import deliver_mail
            mode = (get_mail_setting(conn, "provider_mode", "stub") or "stub").strip().lower()
            send_id, status, err = deliver_mail(
                conn,
                channel="test",
                to_email=email,
                subject=subject,
                contact=contact,
                contact_id=contact_id,
                template_id=template_id,
                domain_id=domain_id,
                html_body=html_body,
                text_body=text_body,
                inject_tracking=_inject_tracking,
            )
            links = _rows(fetchall(
                conn,
                "SELECT token, dest_url FROM mail_click_links WHERE send_id = ?",
                (send_id,),
            ))
            for L in links:
                L["track_url"] = _track_url(L["token"])
            conn.commit()
        msg = (
            "Test mail gönderildi (SMTP)."
            if mode == "smtp" and status == "sent"
            else ("Test simüle edildi (stub)." if status == "simulated" else f"Durum: {status}" + (f" — {err}" if err else ""))
        )
        return jsonify({
            "ok": status in ("sent", "simulated"),
            "send_id": send_id,
            "status": status,
            "mode": mode,
            "error": err or "",
            "tracked_links": links,
            "message": msg,
        })

    @bp.route("/templates/<int:template_id>", methods=["DELETE"])
    @mail_perm(*MAIL_TPL)
    def delete_template(template_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT id, tenant_id FROM mail_templates WHERE id = ?", (template_id,))
            if not row:
                return jsonify({"error": "Şablon bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu şablon başka firmaya ait."}), 403
            except Exception:
                pass
            execute(conn, "DELETE FROM mail_templates WHERE id = ?", (template_id,))
            conn.commit()
        return jsonify({"ok": True})

    # ── Campaigns ──────────────────────────────────────────────
    @bp.route("/campaigns", methods=["GET"])
    @mail_perm(*MAIL_CAMP)
    def list_campaigns():
        from mail_campaign_worker import is_campaign_running, reconcile_campaign_counts

        with closing(get_db()) as conn:
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if _tid:
                rows = _rows(fetchall(
                    conn,
                    "SELECT * FROM mail_campaigns WHERE tenant_id = ? ORDER BY id DESC LIMIT 100",
                    (int(_tid),),
                ))
            else:
                rows = _rows(fetchall(conn, "SELECT * FROM mail_campaigns ORDER BY id DESC LIMIT 100"))
            repaired = False
            for r in rows:
                r["recipient_count"] = r.get("total_count") or scalar(
                    conn,
                    "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ?",
                    (r["id"],),
                ) or 0
                r["pending_count"] = scalar(
                    conn,
                    "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ? AND status = 'pending'",
                    (r["id"],),
                ) or 0
                r["is_running"] = is_campaign_running(r["id"])
                # Biten kampanyada sayaç alıcıdan geride kaldıysa düzelt (eski race bug)
                try:
                    processed = (
                        int(r.get("sent_count") or 0)
                        + int(r.get("failed_count") or 0)
                        + int(r.get("skipped_count") or 0)
                    )
                    total = int(r.get("total_count") or r.get("recipient_count") or 0)
                    if (
                        r.get("status") in ("done", "cancelled", "error", "paused", "stopped")
                        and total > 0
                        and processed < total
                        and int(r.get("pending_count") or 0) == 0
                    ):
                        fixed = reconcile_campaign_counts(conn, r["id"])
                        r["sent_count"] = fixed["sent"]
                        r["failed_count"] = fixed["failed"]
                        r["skipped_count"] = fixed["skipped"]
                        if fixed.get("total"):
                            r["total_count"] = fixed["total"]
                            r["recipient_count"] = fixed["total"]
                        repaired = True
                except Exception:
                    safe_rollback(conn)
                # Çift gönderim riski: panel refresh’te web process’inden start YOK.
                # Sadece mikromail-worker resume eder.
            if repaired:
                try:
                    conn.commit()
                except Exception:
                    pass
        return jsonify({"campaigns": rows})

    @bp.route("/campaigns", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def create_campaign():
        try:
            return _create_campaign_inner()
        except Exception as exc:
            print(f"⚠️  create_campaign: {exc}")
            return jsonify({"error": f"Kampanya oluşturulamadı: {exc}"}), 500

    def _create_campaign_inner():
        data = request.get_json(silent=True) or {}
        name = (data.get("name") or "").strip()
        if not name:
            return jsonify({"error": "Kampanya adı gerekli."}), 400
        template_id = data.get("template_id")
        domain_raw = data.get("domain_id")
        auto_flag = data.get("auto_domain")
        if auto_flag is True or auto_flag == 1 or str(auto_flag).lower() in ("1", "true", "yes"):
            auto_domain = True
        elif auto_flag is False or auto_flag == 0 or str(auto_flag).lower() in ("0", "false", "no"):
            auto_domain = False
        else:
            # domain boş / "auto" → otomatik rotasyon
            auto_domain = domain_raw in (None, "", 0, "0", "auto")
        domain_id = None if auto_domain else domain_raw
        if not template_id:
            return jsonify({"error": "Şablon seçin."}), 400
        if not auto_domain and not domain_id:
            return jsonify({"error": "Domain seçin veya otomatik bırakın."}), 400
        now = iso(utcnow())
        # tag_filters[] öncelikli; yoksa tag_filter (tek / virgüllü)
        if data.get("tag_filters") is not None:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filters"))
        else:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filter"))
        max_recipients = data.get("max_recipients")
        try:
            max_recipients = int(max_recipients)
            if max_recipients <= 0:
                max_recipients = None
        except (TypeError, ValueError):
            max_recipients = None
        exclude_sent = data.get("exclude_previously_sent")
        exclude_sent = True if exclude_sent is None else bool(exclude_sent)
        recipient_mode = (data.get("recipient_mode") or "tag").strip().lower()
        if recipient_mode not in ("tag", "selected", "manual"):
            recipient_mode = "tag"
        raw_ids = data.get("contact_ids") or []
        contact_ids = []
        if isinstance(raw_ids, list):
            for x in raw_ids:
                try:
                    contact_ids.append(int(x))
                except (TypeError, ValueError):
                    continue
        manual_raw = data.get("manual_emails") or data.get("emails") or ""
        if isinstance(manual_raw, list):
            emails = [str(x) for x in manual_raw]
        else:
            emails = re.split(r"[\s,;]+", str(manual_raw or ""))
        emails = [e.strip() for e in emails if e and e.strip()]
        if recipient_mode == "selected" and not contact_ids:
            return jsonify({"error": "Seçili kontak yok. Mail Rehber’den kişi işaretle."}), 400
        if recipient_mode == "manual" and not emails:
            return jsonify({"error": "Elle e-posta listesi boş."}), 400
        if recipient_mode == "tag":
            # contact_ids kalabilir → tam etiketler ∪ manuel seçim (karışık)
            emails = []
        elif recipient_mode == "selected":
            emails = []
            tag_filter = ""
        elif recipient_mode == "manual":
            contact_ids = []
            tag_filter = ""
        try:
            rate = int(data.get("rate_per_minute") or 120)
        except (TypeError, ValueError):
            rate = 120
        rate = max(1, min(rate, 6000))
        scheduled_raw = (data.get("scheduled_at") or "").strip() or None
        # datetime-local (TR) → timezone-aware ISO
        if scheduled_raw:
            try:
                from mail_campaign_worker import normalize_scheduled_at
                scheduled_raw = normalize_scheduled_at(scheduled_raw) or scheduled_raw
            except Exception:
                if "T" in scheduled_raw and len(scheduled_raw) == 16:
                    scheduled_raw = scheduled_raw + ":00"
        with closing(get_db()) as conn:
            from mail_scrub import scrub_settings as _scrub_settings
            scrub_cfg = _scrub_settings(conn)
            only_verified = data.get("only_verified")
            if only_verified is None:
                only_verified = bool(scrub_cfg.get("scrub_campaign_only_valid"))
            else:
                only_verified = bool(only_verified)
            _tid = None
            try:
                from mail_tenant import current_tenant_id as _ctid
                _tid = _ctid()
            except Exception:
                _tid = None
            if not fetchone(conn, "SELECT id FROM mail_templates WHERE id = ?", (template_id,)):
                return jsonify({"error": "Şablon bulunamadı."}), 404
            if _tid:
                tpl_own = fetchone(
                    conn,
                    "SELECT id, tenant_id FROM mail_templates WHERE id = ?",
                    (template_id,),
                )
                if tpl_own and tpl_own.get("tenant_id") and int(tpl_own["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Şablon başka firmaya ait."}), 403
            from database import _table_columns, migrate_mail_campaigns_pro
            from mail_domain_pick import (
                ensure_auto_domain_column,
                pick_tenant_domain,
                tenant_domain_capacity_snapshot,
            )

            # Mikromail DB'de pro kolonlar eksik kalmış olabilir — insert öncesi garanti et
            try:
                migrate_mail_campaigns_pro(conn)
                ensure_auto_domain_column(conn)
            except Exception as mig_exc:
                print(f"⚠️  campaign migrate before insert: {mig_exc}")
                try:
                    conn.rollback()
                except Exception:
                    pass

            try:
                from mail_tenant import assert_tenant_domain, tenant_send_allowed
                if _tid:
                    ok_send, send_err = tenant_send_allowed(conn, int(_tid))
                    if not ok_send:
                        return jsonify({"error": send_err}), 400
            except Exception:
                pass

            if auto_domain:
                cap = tenant_domain_capacity_snapshot(conn, int(_tid) if _tid else None)
                if int(cap.get("sendable_count") or 0) <= 0:
                    return jsonify({
                        "error": "Gönderilebilir domain yok (paused/burned veya günlük cap dolu). "
                                 "Isıtma / Platform domain durumunu kontrol et.",
                        "domain_capacity": cap,
                    }), 400
                # Telemetri: ilk seçilen domain (auto_domain=1 iken worker rotasyon yapar)
                domain_id = pick_tenant_domain(conn, int(_tid) if _tid else None)
                if not domain_id:
                    return jsonify({"error": "Otomatik domain seçilemedi.", "domain_capacity": cap}), 400
            else:
                try:
                    domain_id = int(domain_id)
                except (TypeError, ValueError):
                    return jsonify({"error": "Geçersiz domain."}), 400
                if not fetchone(conn, "SELECT id FROM mail_domains WHERE id = ?", (domain_id,)):
                    return jsonify({"error": "Domain bulunamadı."}), 404
                try:
                    from mail_tenant import assert_tenant_domain
                    if _tid:
                        assert_tenant_domain(conn, int(domain_id), int(_tid))
                except PermissionError as p_exc:
                    return jsonify({"error": str(p_exc)}), 403
                except Exception:
                    pass

            notes_extra = (data.get("notes") or "").strip()
            if auto_domain:
                auto_note = "[auto-domain]"
                notes_extra = f"{auto_note} {notes_extra}".strip() if notes_extra else auto_note
            if recipient_mode == "tag" and contact_ids and tag_filter:
                mix_note = f"[mixed: {len(contact_ids)} seçili + etiket]"
                notes_extra = f"{mix_note} {notes_extra}".strip()
            elif recipient_mode != "tag":
                if notes_extra:
                    notes_extra = f"[{recipient_mode}] {notes_extra}"
                else:
                    notes_extra = f"[{recipient_mode}]"

            camp_cols = _table_columns(conn, "mail_campaigns") or set()
            required_pro = (
                "scheduled_at", "rate_per_minute", "max_recipients",
                "exclude_previously_sent", "total_count", "sent_count",
                "failed_count", "skipped_count", "error",
            )
            missing = [c for c in required_pro if c not in camp_cols]
            if missing:
                return jsonify({
                    "error": "Veritabanı güncel değil (eksik kolon: "
                             + ", ".join(missing)
                             + "). Render’da mikromail servisini Restart et."
                }), 500

            has_auto_col = "auto_domain" in camp_cols
            auto_val = 1 if auto_domain else 0
            insert_params = (
                name, template_id, domain_id, tag_filter, notes_extra,
                scheduled_raw, rate, max_recipients, 1 if exclude_sent else 0,
                now, now,
            )
            if _tid and "tenant_id" in camp_cols and has_auto_col:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_campaigns
                    (name, campaign_type, template_id, domain_id, status, tag_filter, notes,
                     scheduled_at, rate_per_minute, max_recipients, exclude_previously_sent,
                     total_count, sent_count, failed_count, skipped_count, error,
                     created_at, updated_at, tenant_id, auto_domain)
                    VALUES (?, 'bulk', ?, ?, 'draft', ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, '', ?, ?, ?, ?)
                    """,
                    insert_params + (int(_tid), auto_val),
                )
            elif _tid and "tenant_id" in camp_cols:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_campaigns
                    (name, campaign_type, template_id, domain_id, status, tag_filter, notes,
                     scheduled_at, rate_per_minute, max_recipients, exclude_previously_sent,
                     total_count, sent_count, failed_count, skipped_count, error,
                     created_at, updated_at, tenant_id)
                    VALUES (?, 'bulk', ?, ?, 'draft', ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, '', ?, ?, ?)
                    """,
                    insert_params + (int(_tid),),
                )
            elif has_auto_col:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_campaigns
                    (name, campaign_type, template_id, domain_id, status, tag_filter, notes,
                     scheduled_at, rate_per_minute, max_recipients, exclude_previously_sent,
                     total_count, sent_count, failed_count, skipped_count, error,
                     created_at, updated_at, auto_domain)
                    VALUES (?, 'bulk', ?, ?, 'draft', ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, '', ?, ?, ?)
                    """,
                    insert_params + (auto_val,),
                )
            else:
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_campaigns
                    (name, campaign_type, template_id, domain_id, status, tag_filter, notes,
                     scheduled_at, rate_per_minute, max_recipients, exclude_previously_sent,
                     total_count, sent_count, failed_count, skipped_count, error,
                     created_at, updated_at)
                    VALUES (?, 'bulk', ?, ?, 'draft', ?, ?, ?, ?, ?, ?, 0, 0, 0, 0, '', ?, ?)
                    """,
                    insert_params,
                )
            if not cid:
                return jsonify({"error": "Kampanya kaydı yazılamadı (DB)."}), 500
            try:
                attach_ids = None
                if recipient_mode == "selected":
                    attach_ids = contact_ids
                elif recipient_mode == "tag" and contact_ids:
                    attach_ids = contact_ids  # birleşim: etiket + seçili
                elif recipient_mode == "manual":
                    attach_ids = contact_ids or None
                attached = _attach_campaign_recipients(
                    conn, cid, tag_filter=tag_filter if recipient_mode == "tag" else "",
                    max_recipients=max_recipients,
                    exclude_previously_sent=exclude_sent, now=now, only_verified=only_verified,
                    contact_ids=attach_ids,
                    emails=emails if recipient_mode == "manual" else None,
                    tenant_id=_tid,
                )
            except Exception as attach_exc:
                print(f"⚠️  campaign attach: {attach_exc}")
                try:
                    conn.rollback()
                except Exception:
                    pass
                return jsonify({"error": f"Alıcılar eklenemedi: {attach_exc}"}), 500
            if recipient_mode == "manual" and int(attached or 0) == 0:
                execute(conn, "DELETE FROM mail_campaigns WHERE id = ?", (cid,))
                conn.commit()
                return jsonify({
                    "error": "Elle listedeki e-postalar eklenemedi. Adresleri kontrol et; "
                             "«Önceden mail atılmışları hariç tut» kutusunu kapatıp tekrar dene."
                }), 400
            execute(
                conn,
                "UPDATE mail_campaigns SET total_count = ?, updated_at = ? WHERE id = ?",
                (attached, now, cid),
            )
            tag_breakdown = []
            if recipient_mode == "tag" and tag_filter:
                tag_breakdown = _tag_breakdown_for_campaign(
                    conn, tag_filter,
                    exclude_previously_sent=exclude_sent,
                    only_verified=only_verified,
                    tenant_id=_tid,
                )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (cid,))
            out = _row(row)
            out["recipient_count"] = attached
            out["tag_breakdown"] = tag_breakdown
            out["tag_filters"] = _parse_tag_filter_list(tag_filter)
        return jsonify({"campaign": out}), 201

    @bp.route("/campaigns/select-preview", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def preview_campaign_selection():
        """Kampanya oluşturmadan önce filtreye kaç kişinin denk geldiğini gösterir."""
        data = request.get_json(silent=True) or {}
        if data.get("tag_filters") is not None:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filters"))
        else:
            tag_filter = _normalize_tag_filter_storage(data.get("tag_filter"))
        tags_list = _parse_tag_filter_list(tag_filter)
        exclude_sent = data.get("exclude_previously_sent")
        exclude_sent = True if exclude_sent is None else bool(exclude_sent)
        max_recipients = data.get("max_recipients")
        recipient_mode = (data.get("recipient_mode") or "tag").strip().lower()
        raw_ids = data.get("contact_ids") or []
        contact_ids = []
        if isinstance(raw_ids, list):
            for x in raw_ids:
                try:
                    contact_ids.append(int(x))
                except (TypeError, ValueError):
                    continue
        manual_raw = data.get("manual_emails") or data.get("emails") or ""
        if isinstance(manual_raw, list):
            emails = [str(x) for x in manual_raw]
        else:
            emails = re.split(r"[\s,;]+", str(manual_raw or ""))
        emails = [e.strip().lower() for e in emails if e and EMAIL_RE.match(e.strip().lower())]
        try:
            max_recipients = int(max_recipients)
            if max_recipients <= 0:
                max_recipients = None
        except (TypeError, ValueError):
            max_recipients = None
        tag_breakdown = []
        custom_exempt = []
        with closing(get_db()) as conn:
            from mail_scrub import scrub_settings as _scrub_settings
            scrub_cfg = _scrub_settings(conn)
            only_verified = data.get("only_verified")
            if only_verified is None:
                only_verified = bool(scrub_cfg.get("scrub_campaign_only_valid"))
            else:
                only_verified = bool(only_verified)
            approx = True
            mixed_selected = 0
            custom_exempt = _load_exclude_sent_exempt_custom(conn)
            _tid = None
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if recipient_mode == "manual":
                total = len(emails)
                approx = False
            elif recipient_mode == "selected":
                filtered = _filter_sendable_contact_ids(
                    conn, contact_ids,
                    exclude_previously_sent=exclude_sent,
                    only_verified=only_verified,
                    custom_exempt=custom_exempt,
                    tenant_id=_tid,
                )
                total = len(filtered)
                approx = False
            else:
                # Etiket ∪ manuel seçim (karışık)
                filtered = []
                if contact_ids:
                    filtered = _filter_sendable_contact_ids(
                        conn, contact_ids,
                        exclude_previously_sent=exclude_sent,
                        only_verified=only_verified,
                        custom_exempt=custom_exempt,
                        tenant_id=_tid,
                    )
                    mixed_selected = len(filtered)
                where_sql, params = _campaign_selection_where(
                    tag_filter, exclude_sent, only_verified=only_verified,
                    custom_exempt=custom_exempt, tenant_id=_tid,
                )
                if tags_list and filtered:
                    ph = ",".join(["?"] * len(filtered))
                    try:
                        total = int(scalar(
                            conn,
                            f"SELECT COUNT(*) FROM mail_contacts WHERE ({where_sql}) OR id IN ({ph})",
                            tuple(params) + tuple(filtered),
                        ) or 0)
                        approx = False
                    except Exception:
                        total = mixed_selected
                        approx = True
                elif tags_list:
                    total = None
                    if len(tags_list) == 1 and not exclude_sent and not only_verified and not _tid:
                        total = _registry_tag_count(conn, tags_list[0])
                    if total is None:
                        try:
                            total = int(scalar(
                                conn,
                                f"SELECT COUNT(*) FROM mail_contacts WHERE {where_sql}",
                                tuple(params),
                            ) or 0)
                            approx = False
                        except Exception:
                            total = None
                    if total is None:
                        try:
                            total, approx = _approx_contact_total(conn)
                        except Exception:
                            total = 0
                            approx = True
                    total = int(total or 0)
                elif filtered:
                    total = len(filtered)
                    approx = False
                else:
                    try:
                        total, approx = _approx_contact_total(conn)
                    except Exception:
                        total = 0
                        approx = True
                    total = int(total or 0)
                if tags_list:
                    tag_breakdown = _tag_breakdown_for_campaign(
                        conn, tags_list,
                        exclude_previously_sent=exclude_sent,
                        only_verified=only_verified,
                        custom_exempt=custom_exempt,
                        tenant_id=_tid,
                    )
        will_attach = min(total, max_recipients) if max_recipients else total
        exempt_selected = [
            t for t in tags_list if _tag_is_exclude_sent_exempt(t, custom_exempt)
        ]
        # Muaf/test etiketlerinin garanti edilecek tahmini sayısı (breakdown)
        priority_count = 0
        for row in tag_breakdown:
            if row.get("exclude_sent_exempt"):
                try:
                    priority_count += int(row.get("count") or 0)
                except (TypeError, ValueError):
                    pass
        priority_truncated = bool(
            max_recipients and priority_count and priority_count > int(max_recipients)
        )
        return jsonify({
            "matching_count": total,
            "will_attach": will_attach,
            "max_recipients": max_recipients,
            "approx": approx,
            "recipient_mode": recipient_mode,
            "tag_filters": tags_list,
            "tag_breakdown": tag_breakdown,
            "exclude_sent_exempt_tags": exempt_selected,
            "priority_guaranteed": min(priority_count, will_attach) if exempt_selected else 0,
            "priority_truncated": priority_truncated,
            "manual_selected": mixed_selected if recipient_mode == "tag" else (
                len(contact_ids) if recipient_mode == "selected" else 0
            ),
        })

    @bp.route("/campaigns/<int:campaign_id>", methods=["PATCH"])
    @mail_perm(*MAIL_CAMP)
    def update_campaign(campaign_id):
        data = request.get_json(silent=True) or {}
        now = iso(utcnow())
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not row:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if row["status"] not in ("draft", "scheduled"):
                return jsonify({"error": "Sadece taslak / zamanlanmış kampanyalar düzenlenebilir."}), 400
            scheduled_raw = data.get("scheduled_at") if "scheduled_at" in data else row.get("scheduled_at")
            if isinstance(scheduled_raw, str):
                scheduled_raw = scheduled_raw.strip() or None
                if scheduled_raw:
                    try:
                        from mail_campaign_worker import normalize_scheduled_at
                        scheduled_raw = normalize_scheduled_at(scheduled_raw) or scheduled_raw
                    except Exception:
                        if "T" in scheduled_raw and len(scheduled_raw) == 16:
                            scheduled_raw = scheduled_raw + ":00"
            try:
                rate = int(data.get("rate_per_minute") if "rate_per_minute" in data else (row.get("rate_per_minute") or 120))
            except (TypeError, ValueError):
                rate = 120
            rate = max(1, min(rate, 6000))
            execute(
                conn,
                """
                UPDATE mail_campaigns SET name = ?, template_id = ?, domain_id = ?,
                    tag_filter = ?, notes = ?, scheduled_at = ?, rate_per_minute = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    (data.get("name") if "name" in data else row["name"]).strip(),
                    data.get("template_id") if "template_id" in data else row["template_id"],
                    data.get("domain_id") if "domain_id" in data else row["domain_id"],
                    (data.get("tag_filter") if "tag_filter" in data else row["tag_filter"] or "").strip(),
                    (data.get("notes") if "notes" in data else row["notes"] or "").strip(),
                    scheduled_raw,
                    rate,
                    now,
                    campaign_id,
                ),
            )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
        return jsonify({"campaign": _row(row)})

    @bp.route("/campaigns/<int:campaign_id>", methods=["DELETE"])
    @mail_perm(*MAIL_CAMP)
    def delete_campaign(campaign_id):
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT status, tenant_id FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not row:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            row = _row(row)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and row.get("tenant_id") and int(row["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if row["status"] in ("sending", "queued"):
                return jsonify({"error": "Gönderimdeki kampanya silinemez — önce iptal edin."}), 400
            execute(conn, "DELETE FROM mail_campaign_recipients WHERE campaign_id = ?", (campaign_id,))
            execute(conn, "DELETE FROM mail_campaigns WHERE id = ?", (campaign_id,))
            conn.commit()
        return jsonify({"ok": True})

    @bp.route("/campaigns/<int:campaign_id>/queue", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def queue_campaign(campaign_id):
        """Kampanyayı hemen veya zamanlanmış olarak kuyruğa alır (arka plan worker)."""
        from mail_campaign_worker import start_campaign_send

        data = request.get_json(silent=True) or {}
        now_dt = utcnow()
        now = iso(now_dt)
        force_now = bool(data.get("send_now"))
        with closing(get_db()) as conn:
            camp = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid_guard = current_tenant_id()
                if _tid_guard and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid_guard):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if camp["status"] not in ("draft", "scheduled", "queued"):
                return jsonify({"error": f"Kampanya durumu uygun değil: {camp['status']}"}), 400
            pending = scalar(
                conn,
                "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ? AND status = 'pending'",
                (campaign_id,),
            ) or 0
            if pending <= 0:
                return jsonify({"error": "Gönderilecek alıcı yok. Kampanyayı yeniden oluşturun."}), 400
            if not fetchone(conn, "SELECT id FROM mail_templates WHERE id = ?", (camp["template_id"],)):
                return jsonify({"error": "Şablon bulunamadı."}), 400

            # Alibaba hesap günlük kotası — yetersizse kuyruğa alma
            try:
                from mail_account_quota import can_queue
                ok_q, q_err, q_snap = can_queue(conn, int(pending))
                if not ok_q:
                    return jsonify({
                        "error": q_err,
                        "quota": q_snap,
                    }), 400
            except Exception as q_exc:
                print(f"⚠️  queue quota check: {q_exc}")

            # Prepaid mail kredisi (global + firma tahsisi)
            try:
                from mail_credit import can_consume
                tid = camp.get("tenant_id")
                ok_c, c_err, c_snap = can_consume(
                    conn, int(pending), tenant_id=int(tid) if tid else None
                )
                if not ok_c:
                    return jsonify({"error": c_err, "credit": c_snap}), 400
            except Exception as c_exc:
                print(f"⚠️  queue credit check: {c_exc}")

            # Domain kapasitesi — auto: tahsisli havuz; pinned: tek domain
            try:
                from mail_domain_pick import (
                    campaign_is_auto,
                    ensure_auto_domain_column,
                    tenant_domain_capacity_snapshot,
                )
                from mail_domain_health import domain_is_send_blocked

                ensure_auto_domain_column(conn)
                if campaign_is_auto(camp):
                    tid = camp.get("tenant_id")
                    cap = tenant_domain_capacity_snapshot(
                        conn, int(tid) if tid else None
                    )
                    if int(cap.get("remaining_today") or 0) <= 0:
                        return jsonify({
                            "error": "Bugün gönderilebilir domain kapasitesi yok "
                                     "(cap dolu veya domainler paused).",
                            "domain_capacity": cap,
                        }), 400
                else:
                    blocked, block_reason = domain_is_send_blocked(conn, camp.get("domain_id"))
                    if blocked:
                        return jsonify({"error": f"Domain engeli: {block_reason}"}), 400
            except Exception as d_exc:
                print(f"⚠️  queue domain capacity: {d_exc}")

            scheduled_at = camp.get("scheduled_at")
            start_immediately = force_now
            if not start_immediately and scheduled_at:
                from mail_campaign_worker import _parse_iso, schedule_is_future
                sched = _parse_iso(scheduled_at)
                if not sched:
                    return jsonify({
                        "error": "Zamanlama saati okunamadı — kampanya başlatılmadı. Saati tekrar seç."
                    }), 400
                if schedule_is_future(scheduled_at, now=now_dt):
                    execute(
                        conn,
                        """
                        UPDATE mail_campaigns
                        SET status = 'scheduled', total_count = COALESCE(NULLIF(total_count, 0), ?),
                            queued_at = NULL, updated_at = ?, error = ''
                        WHERE id = ?
                        """,
                        (pending, now, campaign_id),
                    )
                    conn.commit()
                    mode = (get_mail_setting(conn, "provider_mode", "stub") or "stub").strip().lower()
                    return jsonify({
                        "ok": True,
                        "status": "scheduled",
                        "scheduled_at": scheduled_at,
                        "pending": pending,
                        "mode": mode,
                        "message": f"Kampanya zamanlandı · {pending} alıcı · {scheduled_at} (saat gelince başlar)",
                    })
                # Saat geçmiş → kullanıcı bilinçli gönderiyor sayılır, aşağıda kuyruk

            execute(
                conn,
                """
                UPDATE mail_campaigns
                SET status = 'queued', queued_at = ?, total_count = COALESCE(NULLIF(total_count, 0), ?),
                    updated_at = ?, error = ''
                WHERE id = ?
                """,
                (now, pending, now, campaign_id),
            )
            conn.commit()
            mode = (get_mail_setting(conn, "provider_mode", "stub") or "stub").strip().lower()

        start_campaign_send(campaign_id)
        return jsonify({
            "ok": True,
            "status": "queued",
            "pending": pending,
            "mode": mode,
            "message": (
                f"{pending} alıcı kuyruğa alındı — arka planda gönderiliyor."
                + (" (stub simülasyon)" if mode != "smtp" else " (SMTP)")
            ),
        })

    @bp.route("/campaigns/<int:campaign_id>/cancel", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def cancel_campaign(campaign_id):
        now = iso(utcnow())
        with closing(get_db()) as conn:
            camp = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid_guard = current_tenant_id()
                if _tid_guard and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid_guard):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if camp["status"] not in ("scheduled", "queued", "sending", "cancelling"):
                return jsonify({"error": f"İptal edilemez: {camp['status']}"}), 400
            new_status = "cancelled" if camp["status"] == "scheduled" else "cancelling"
            execute(
                conn,
                "UPDATE mail_campaigns SET status = ?, updated_at = ? WHERE id = ?",
                (new_status, now, campaign_id),
            )
            if new_status == "cancelled":
                execute(
                    conn,
                    "UPDATE mail_campaigns SET finished_at = ? WHERE id = ?",
                    (now, campaign_id),
                )
            conn.commit()
        return jsonify({"ok": True, "status": new_status})

    @bp.route("/campaigns/<int:campaign_id>/pause", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def pause_campaign(campaign_id):
        now = iso(utcnow())
        with closing(get_db()) as conn:
            camp = fetchone(conn, "SELECT status, tenant_id FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid_guard = current_tenant_id()
                if _tid_guard and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid_guard):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if camp["status"] not in ("sending", "queued", "scheduled"):
                return jsonify({"error": f"Duraklatılamaz: {camp['status']}"}), 400
            execute(
                conn,
                "UPDATE mail_campaigns SET status = 'paused', updated_at = ? WHERE id = ?",
                (now, campaign_id),
            )
            conn.commit()
            try:
                from mail_ops import audit
                audit(conn, request.headers.get("X-Admin-User") or "admin", "campaign_pause", f"id={campaign_id}")
                conn.commit()
            except Exception:
                pass
        return jsonify({"ok": True, "status": "paused"})

    @bp.route("/campaigns/<int:campaign_id>/resume", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def resume_campaign(campaign_id):
        from mail_campaign_worker import start_campaign_send
        now = iso(utcnow())
        with closing(get_db()) as conn:
            camp = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid_guard = current_tenant_id()
                if _tid_guard and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid_guard):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if camp["status"] != "paused":
                return jsonify({"error": f"Devam ettirilemez: {camp['status']}"}), 400
            pending = scalar(
                conn,
                "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ? AND status = 'pending'",
                (campaign_id,),
            ) or 0
            try:
                from mail_account_quota import can_queue
                ok_q, q_err, q_snap = can_queue(conn, int(pending))
                if not ok_q:
                    return jsonify({"error": q_err, "quota": q_snap}), 400
            except Exception as q_exc:
                print(f"⚠️  resume quota check: {q_exc}")
            try:
                from mail_credit import can_consume
                tid = camp.get("tenant_id")
                ok_c, c_err, c_snap = can_consume(
                    conn, int(pending), tenant_id=int(tid) if tid else None
                )
                if not ok_c:
                    return jsonify({"error": c_err, "credit": c_snap}), 400
            except Exception as c_exc:
                print(f"⚠️  resume credit check: {c_exc}")
            try:
                from mail_domain_pick import campaign_is_auto, tenant_domain_capacity_snapshot
                from mail_domain_health import domain_is_send_blocked
                camp_d = camp
                if campaign_is_auto(camp_d):
                    tid = camp_d.get("tenant_id")
                    cap = tenant_domain_capacity_snapshot(conn, int(tid) if tid else None)
                    if int(cap.get("remaining_today") or 0) <= 0:
                        return jsonify({
                            "error": "Bugün domain kapasitesi yok — devam ettirilemez.",
                            "domain_capacity": cap,
                        }), 400
                else:
                    blocked, block_reason = domain_is_send_blocked(conn, camp_d.get("domain_id"))
                    if blocked:
                        return jsonify({"error": f"Domain engeli: {block_reason}"}), 400
            except Exception as d_exc:
                print(f"⚠️  resume domain capacity: {d_exc}")
            execute(
                conn,
                "UPDATE mail_campaigns SET status = 'queued', updated_at = ?, error = '' WHERE id = ?",
                (now, campaign_id),
            )
            conn.commit()
        start_campaign_send(campaign_id)
        return jsonify({"ok": True, "status": "queued"})

    @bp.route("/campaigns/<int:campaign_id>/retry-failed", methods=["POST"])
    @mail_perm(*MAIL_CAMP)
    def retry_failed_campaign(campaign_id):
        from mail_campaign_worker import start_campaign_send
        now = iso(utcnow())
        with closing(get_db()) as conn:
            camp = fetchone(conn, "SELECT * FROM mail_campaigns WHERE id = ?", (campaign_id,))
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid_guard = current_tenant_id()
                if _tid_guard and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid_guard):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            if camp["status"] not in ("done", "error", "cancelled", "paused"):
                return jsonify({"error": f"Retry için uygun değil: {camp['status']}"}), 400
            failed_n = scalar(
                conn,
                "SELECT COUNT(*) FROM mail_campaign_recipients WHERE campaign_id = ? AND status = 'failed'",
                (campaign_id,),
            ) or 0
            try:
                from mail_account_quota import can_queue
                ok_q, q_err, q_snap = can_queue(conn, int(failed_n))
                if not ok_q:
                    return jsonify({"error": q_err, "quota": q_snap}), 400
            except Exception as q_exc:
                print(f"⚠️  retry-failed quota check: {q_exc}")
            try:
                from mail_credit import can_consume
                tid = camp.get("tenant_id")
                ok_c, c_err, c_snap = can_consume(
                    conn, int(failed_n), tenant_id=int(tid) if tid else None
                )
                if not ok_c:
                    return jsonify({"error": c_err, "credit": c_snap}), 400
            except Exception as c_exc:
                print(f"⚠️  retry-failed credit check: {c_exc}")
            try:
                from mail_domain_pick import campaign_is_auto, tenant_domain_capacity_snapshot
                from mail_domain_health import domain_is_send_blocked
                if campaign_is_auto(camp):
                    cap = tenant_domain_capacity_snapshot(conn, int(tid) if tid else None)
                    if int(cap.get("remaining_today") or 0) <= 0:
                        return jsonify({
                            "error": "Bugün domain kapasitesi yok — retry yapılamadı.",
                            "domain_capacity": cap,
                        }), 400
                else:
                    blocked, block_reason = domain_is_send_blocked(conn, camp.get("domain_id"))
                    if blocked:
                        return jsonify({"error": f"Domain engeli: {block_reason}"}), 400
            except Exception as d_exc:
                print(f"⚠️  retry-failed domain capacity: {d_exc}")
            n = execute(
                conn,
                """
                UPDATE mail_campaign_recipients SET status = 'pending', send_id = NULL
                WHERE campaign_id = ? AND status = 'failed'
                """,
                (campaign_id,),
            )
            reset = getattr(n, "rowcount", None)
            execute(
                conn,
                """
                UPDATE mail_campaigns
                SET status = 'queued', finished_at = NULL, error = '', updated_at = ?
                WHERE id = ?
                """,
                (now, campaign_id),
            )
            conn.commit()
        start_campaign_send(campaign_id)
        return jsonify({"ok": True, "status": "queued", "reset_failed": reset})

    @bp.route("/reports/campaigns", methods=["GET"])
    @mail_perm(*MAIL_REP)
    def report_campaigns():
        from mail_ops import campaign_analytics
        from mail_tenant import current_tenant_id
        cid = request.args.get("campaign_id")
        _tid = current_tenant_id()
        with closing(get_db()) as conn:
            rows = campaign_analytics(conn, int(cid) if cid else None, tenant_id=_tid)
        return jsonify({"campaigns": rows})

    @bp.route("/reports/engagement-timeline", methods=["GET"])
    @mail_perm(*MAIL_REP)
    def report_engagement_timeline():
        """Son N gün için günlük açılma / tıklama sayıları (grafik)."""
        try:
            days = min(max(int(request.args.get("days") or 14), 1), 90)
        except (TypeError, ValueError):
            days = 14
        from datetime import datetime, timedelta, timezone

        since = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%Y-%m-%d")
        from mail_tenant import current_tenant_id
        _tid = current_tenant_id()
        # tenant_id verilmişse (tenant login VEYA superadmin impersonate) SADECE o
        # firmanın gönderimleri sayılır — önceden filtre yoktu, her firma platform
        # genelindeki açılma/tıklama grafiğini görebiliyordu.
        tid_clause = " AND tenant_id = ?" if _tid else ""
        tid_params = (int(_tid),) if _tid else ()
        with closing(get_db()) as conn:
            opens = fetchall(
                conn,
                f"""
                SELECT substr(CAST(opened_at AS TEXT), 1, 10) AS d, COUNT(*) AS n
                FROM mail_sends
                WHERE opened_at IS NOT NULL AND CAST(opened_at AS TEXT) >= ?{tid_clause}
                GROUP BY 1 ORDER BY 1
                """,
                (since,) + tid_params,
            )
            clicks = fetchall(
                conn,
                f"""
                SELECT substr(CAST(clicked_at AS TEXT), 1, 10) AS d, COUNT(*) AS n
                FROM mail_sends
                WHERE clicked_at IS NOT NULL AND CAST(clicked_at AS TEXT) >= ?{tid_clause}
                GROUP BY 1 ORDER BY 1
                """,
                (since,) + tid_params,
            )
        open_map = {str(r["d"]): int(r["n"] or 0) for r in (opens or [])}
        click_map = {str(r["d"]): int(r["n"] or 0) for r in (clicks or [])}
        end = datetime.now(timezone.utc).date()
        labels = []
        open_series = []
        click_series = []
        for i in range(days - 1, -1, -1):
            d = (end - timedelta(days=i)).isoformat()
            labels.append(d)
            open_series.append(open_map.get(d, 0))
            click_series.append(click_map.get(d, 0))
        return jsonify({
            "days": days,
            "labels": labels,
            "opens": open_series,
            "clicks": click_series,
        })

    @bp.route("/campaigns/<int:campaign_id>/recipients", methods=["GET"])
    @mail_perm(*MAIL_REP, *MAIL_CAMP)
    def campaign_recipients_detail(campaign_id):
        """Kampanya alıcıları: e-posta + etiketler + gönderim durumu."""
        try:
            limit = min(int(request.args.get("limit") or 500), 5000)
        except (TypeError, ValueError):
            limit = 500
        status_filter = (request.args.get("status") or "").strip().lower()
        with closing(get_db()) as conn:
            camp = fetchone(
                conn,
                "SELECT id, name, tag_filter, status, total_count, tenant_id FROM mail_campaigns WHERE id = ?",
                (campaign_id,),
            )
            if not camp:
                return jsonify({"error": "Kampanya bulunamadı."}), 404
            camp = _row(camp)
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
                if _tid and camp.get("tenant_id") and int(camp["tenant_id"]) != int(_tid):
                    return jsonify({"error": "Bu kampanya başka firmaya ait."}), 403
            except Exception:
                pass
            where = ["r.campaign_id = ?"]
            params = [campaign_id]
            if status_filter:
                where.append("r.status = ?")
                params.append(status_filter)
            where_sql = " AND ".join(where)
            total = int(scalar(
                conn,
                f"SELECT COUNT(*) FROM mail_campaign_recipients r WHERE {where_sql}",
                tuple(params),
            ) or 0)
            rows = fetchall(
                conn,
                f"""
                SELECT
                    r.id AS recipient_id,
                    r.status AS recipient_status,
                    r.created_at,
                    c.id AS contact_id,
                    c.email,
                    c.name,
                    c.tags,
                    COALESCE(s.status, s2.status) AS send_status,
                    COALESCE(s.opened_at, s2.opened_at) AS opened_at,
                    COALESCE(s.clicked_at, s2.clicked_at) AS clicked_at,
                    COALESCE(NULLIF(TRIM(s.error), ''), NULLIF(TRIM(s2.error), '')) AS send_error
                FROM mail_campaign_recipients r
                JOIN mail_contacts c ON c.id = r.contact_id
                LEFT JOIN mail_sends s ON s.id = r.send_id
                LEFT JOIN mail_sends s2 ON s2.id = (
                    SELECT s3.id FROM mail_sends s3
                    WHERE s3.campaign_id = r.campaign_id AND s3.contact_id = r.contact_id
                    ORDER BY s3.id DESC LIMIT 1
                )
                WHERE {where_sql}
                ORDER BY r.id ASC
                LIMIT ?
                """,
                tuple(params + [limit]),
            )
            out = []
            for r in rows or []:
                d = dict(r)
                d["tags"] = _parse_tags(d.get("tags"))
                out.append(d)
            camp_d = _row(camp)
        return jsonify({
            "campaign": camp_d,
            "recipients": out,
            "total": total,
            "limit": limit,
            "truncated": total > len(out),
        })

    @bp.route("/contacts/export", methods=["GET"])
    @mail_perm(*MAIL_CRM)
    def export_contacts():
        tag = (request.args.get("tag") or "").strip()
        try:
            limit = min(int(request.args.get("limit") or 50000), 200000)
        except (TypeError, ValueError):
            limit = 50000
        with closing(get_db()) as conn:
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if tag:
                clause, params = _tag_match_clause(tag)
                params = list(params)
                if _tid:
                    clause += " AND tenant_id = ?"
                    params.append(int(_tid))
                rows = fetchall(
                    conn,
                    f"SELECT email, name, phone, tags, source, unsubscribed, created_at FROM mail_contacts WHERE {clause} ORDER BY id DESC LIMIT ?",
                    tuple(params) + (limit,),
                )
            elif _tid:
                rows = fetchall(
                    conn,
                    "SELECT email, name, phone, tags, source, unsubscribed, created_at FROM mail_contacts WHERE tenant_id = ? ORDER BY id DESC LIMIT ?",
                    (int(_tid), limit),
                )
            else:
                rows = fetchall(
                    conn,
                    "SELECT email, name, phone, tags, source, unsubscribed, created_at FROM mail_contacts ORDER BY id DESC LIMIT ?",
                    (limit,),
                )
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["email", "name", "phone", "tags", "source", "unsubscribed", "created_at"])
        for r in rows or []:
            r = dict(r)
            w.writerow([
                r.get("email") or "",
                r.get("name") or "",
                r.get("phone") or "",
                r.get("tags") or "",
                r.get("source") or "",
                r.get("unsubscribed") or 0,
                r.get("created_at") or "",
            ])
        from flask import Response
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=mail-contacts.csv"},
        )

    @bp.route("/webhooks/bounce", methods=["POST"])
    def bounce_webhook():
        """DirectMail / generic bounce-complaint webhook.
        Auth: X-Mailing-Webhook-Secret veya ?secret=
        Body örnek: { "email": "...", "type": "bounce"|"complaint", "reason": "..." }
        veya Alibaba tarzı alanlar: rcpt, bounceType, status
        """
        from mail_ops import audit, suppress_email
        data = request.get_json(silent=True) or {}
        secret = (
            request.headers.get("X-Mailing-Webhook-Secret")
            or request.args.get("secret")
            or ""
        ).strip()
        with closing(get_db()) as conn:
            expected = (get_mail_setting(conn, "webhook_secret", "") or "").strip()
            if not expected or secret != expected:
                return jsonify({"error": "Unauthorized"}), 401
            email = str(
                data.get("email")
                or data.get("rcpt")
                or data.get("recipient")
                or data.get("Destination")
                or ""
            ).strip().lower()
            # nested
            if not email and isinstance(data.get("mail"), dict):
                dests = data["mail"].get("destination") or []
                if dests:
                    email = str(dests[0]).strip().lower()
            btype = str(data.get("type") or data.get("bounceType") or data.get("event") or "bounce").strip().lower()
            reason = str(data.get("reason") or data.get("diagnosticCode") or btype)[:200]
            if not email or not EMAIL_RE.match(email):
                return jsonify({"error": "email gerekli"}), 400
            if "complaint" in btype or "spam" in btype:
                suppress_email(conn, email, reason="complaint", source="bounce_webhook")
                kind = "complaint"
            else:
                suppress_email(conn, email, reason="bounce", source="bounce_webhook")
                kind = "bounce"
            execute(
                conn,
                """
                UPDATE mail_sends SET status = 'bounced', error = ?
                WHERE LOWER(to_email) = ? AND status IN ('sent','simulated','queued')
                """,
                (f"{kind}: {reason}", email),
            )
            try:
                from mail_ops import tag_send_outcome
                crow = fetchone(
                    conn,
                    "SELECT id FROM mail_contacts WHERE LOWER(email) = ? ORDER BY id DESC LIMIT 1",
                    (email,),
                )
                if crow and crow.get("id"):
                    tag_send_outcome(conn, crow["id"], "bounced", iso(utcnow()))
            except Exception:
                pass
            audit(conn, "webhook", kind, email)
            # Metrik spike → domain auto-pause
            try:
                from mail_domain_health import evaluate_and_maybe_pause
                dom = fetchone(
                    conn,
                    """
                    SELECT domain_id FROM mail_sends
                    WHERE LOWER(to_email) = ? AND domain_id IS NOT NULL
                    ORDER BY id DESC LIMIT 1
                    """,
                    (email,),
                )
                if dom and dom.get("domain_id"):
                    evaluate_and_maybe_pause(conn, int(dom["domain_id"]))
            except Exception as hexc:
                print(f"⚠️  bounce→domain health: {hexc}")
            conn.commit()
        return jsonify({"ok": True, "email": email, "type": kind})

    # ── Sends / Reports ────────────────────────────────────────
    @bp.route("/sends", methods=["GET"])
    @mail_perm(*MAIL_REP)
    def list_sends():
        status = (request.args.get("status") or "").strip()
        channel = (request.args.get("channel") or "").strip()
        try:
            limit = min(int(request.args.get("limit") or 200), 1000)
        except (TypeError, ValueError):
            limit = 200
        with closing(get_db()) as conn:
            sql = "SELECT * FROM mail_sends WHERE 1=1"
            params = []
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            if _tid:
                sql += (
                    " AND (campaign_id IN (SELECT id FROM mail_campaigns WHERE tenant_id = ?)"
                    " OR contact_id IN (SELECT id FROM mail_contacts WHERE tenant_id = ?))"
                )
                params.extend([int(_tid), int(_tid)])
            if status:
                sql += " AND status = ?"
                params.append(status)
            if channel:
                sql += " AND channel = ?"
                params.append(channel)
            sql += " ORDER BY id DESC LIMIT ?"
            params.append(limit)
            rows = _rows(fetchall(conn, sql, tuple(params)))
        return jsonify({"sends": rows, "count": len(rows)})

    @bp.route("/reports/summary", methods=["GET"])
    @mail_perm(*MAIL_REP)
    def reports_summary():
        with closing(get_db()) as conn:
            try:
                from mail_tenant import current_tenant_id
                _tid = current_tenant_id()
            except Exception:
                _tid = None
            tenant_clause = ""
            tenant_params = ()
            if _tid:
                tenant_clause = (
                    " WHERE (campaign_id IN (SELECT id FROM mail_campaigns WHERE tenant_id = ?)"
                    " OR contact_id IN (SELECT id FROM mail_contacts WHERE tenant_id = ?))"
                )
                tenant_params = (int(_tid), int(_tid))
            by_status = _rows(fetchall(
                conn,
                f"SELECT status, COUNT(*) AS cnt FROM mail_sends{tenant_clause} GROUP BY status ORDER BY cnt DESC",
                tenant_params,
            ))
            by_channel = _rows(fetchall(
                conn,
                f"SELECT channel, COUNT(*) AS cnt FROM mail_sends{tenant_clause} GROUP BY channel ORDER BY cnt DESC",
                tenant_params,
            ))
            recent = _rows(fetchall(
                conn,
                f"SELECT * FROM mail_sends{tenant_clause} ORDER BY id DESC LIMIT 20",
                tenant_params,
            ))
        return jsonify({
            "by_status": by_status,
            "by_channel": by_channel,
            "recent": recent,
        })

    # ── Domains / Settings ─────────────────────────────────────
    def _domain_public(row):
        try:
            from mail_tenant import enrich_domain_public

            return enrich_domain_public(row)
        except Exception:
            d = _row(row) if row else None
            if not d:
                return None
            pw = (d.get("smtp_password_enc") or d.get("smtp_password") or "").strip()
            d["smtp_password_set"] = bool(pw)
            d["smtp_password"] = ""
            d.pop("smtp_password_enc", None)
            return d

    @bp.route("/domains", methods=["GET"])
    @mail_perm(*MAIL_SET)
    def list_domains():
        from flask import session as _sess
        from mail_tenant import current_tenant_id, heal_ready_domains, list_allocated_domains

        with closing(get_db()) as conn:
            try:
                heal_ready_domains(conn)
            except Exception:
                pass
            tid = current_tenant_id()
            if _sess.get("mail_is_superadmin") and not tid:
                rows = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC")
            elif tid:
                rows = list_allocated_domains(conn, int(tid))
            else:
                rows = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC")
        return jsonify({"domains": [_domain_public(r) for r in (rows or [])]})

    @bp.route("/account-quota", methods=["GET"])
    @mail_perm(*MAIL_CAMP)
    def account_quota_get():
        """Alibaba günlük kota + prepaid mail kredisi."""
        from mail_account_quota import quota_snapshot
        from mail_credit import credit_snapshot, tenant_credit_snapshot

        with closing(get_db()) as conn:
            snap = quota_snapshot(conn)
            credit = credit_snapshot(conn)
            tenant_credit = None
            try:
                from mail_tenant import current_tenant_id
                tid = current_tenant_id()
                if tid:
                    tenant_credit = tenant_credit_snapshot(conn, int(tid))
            except Exception:
                pass
        return jsonify({"quota": snap, "credit": credit, "tenant_credit": tenant_credit})

    @bp.route("/domain-capacity", methods=["GET"])
    @mail_perm(*MAIL_CAMP)
    def domain_capacity_get():
        """Tenant tahsisli domainlerin bugünkü kalan kapasitesi (otomatik rotasyon UI)."""
        from mail_domain_pick import ensure_auto_domain_column, tenant_domain_capacity_snapshot
        from mail_tenant import current_tenant_id

        with closing(get_db()) as conn:
            try:
                ensure_auto_domain_column(conn)
            except Exception:
                pass
            try:
                tid = current_tenant_id()
                snap = tenant_domain_capacity_snapshot(conn, int(tid) if tid else None)
            except Exception as exc:
                print(f"⚠️  domain_capacity_get: {exc}")
                snap = {
                    "allocated_count": 0,
                    "sendable_count": 0,
                    "remaining_today": 0,
                    "domains": [],
                    "auto_default": True,
                    "note": "Kapasite hesaplanamadı.",
                }
            try:
                conn.commit()
            except Exception:
                pass
        return jsonify(snap)

    @bp.route("/domains/<int:domain_id>", methods=["PATCH"])
    @mail_perm(*MAIL_SET)
    def update_domain(domain_id):
        data = request.get_json(silent=True) or {}
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_domains WHERE id = ?", (domain_id,))
            if not row:
                return jsonify({"error": "Domain bulunamadı."}), 404
            row = dict(row)
            try:
                from flask import session as _sess
                from mail_tenant import assert_tenant_domain, current_tenant_id
                if not _sess.get("mail_is_superadmin"):
                    _tid = current_tenant_id()
                    assert_tenant_domain(conn, int(domain_id), int(_tid) if _tid else None)
            except PermissionError as p_exc:
                return jsonify({"error": str(p_exc)}), 403
            except Exception:
                pass
            from mail_delivery import normalize_from_local, normalize_mail_domain

            from_name = (data.get("from_name") if "from_name" in data else row.get("from_name") or "").strip()
            from_local = normalize_from_local(
                data.get("from_local") if "from_local" in data else row.get("from_local") or "noreply",
                default="noreply",
            )
            status = (data.get("status") if "status" in data else row.get("status") or "pending").strip()
            dns_status = (data.get("dns_status") if "dns_status" in data else row.get("dns_status") or "unconfigured").strip()
            notes = (data.get("notes") if "notes" in data else row.get("notes") or "").strip()
            if "@" in (row.get("domain") or ""):
                execute(
                    conn,
                    "UPDATE mail_domains SET domain = ? WHERE id = ?",
                    (normalize_mail_domain(row.get("domain")), domain_id),
                )
            smtp_plain = row.get("smtp_password") or ""
            if smtp_plain and str(smtp_plain).startswith("enc:v1:"):
                smtp_plain = ""
            smtp_enc = row.get("smtp_password_enc") or ""
            if "smtp_password" in data and data.get("smtp_password") not in (None, ""):
                from mail_tenant import encrypt_secret

                smtp_plain = str(data.get("smtp_password")).strip()
                try:
                    smtp_enc = encrypt_secret(smtp_plain)
                except Exception:
                    smtp_enc = ""
            try:
                execute(
                    conn,
                    """
                    UPDATE mail_domains SET
                        from_name = ?, from_local = ?, status = ?, dns_status = ?, notes = ?,
                        smtp_password_enc = ?, smtp_password = ?
                    WHERE id = ?
                    """,
                    (from_name, from_local, status, dns_status, notes, smtp_enc, smtp_plain, domain_id),
                )
            except Exception:
                try:
                    execute(
                        conn,
                        """
                        UPDATE mail_domains SET
                            from_name = ?, from_local = ?, status = ?, dns_status = ?, notes = ?,
                            smtp_password = ?
                        WHERE id = ?
                        """,
                        (from_name, from_local, status, dns_status, notes, smtp_plain or smtp_enc, domain_id),
                    )
                except Exception:
                    execute(
                        conn,
                        """
                        UPDATE mail_domains SET
                            from_name = ?, from_local = ?, status = ?, dns_status = ?, notes = ?
                        WHERE id = ?
                        """,
                        (from_name, from_local, status, dns_status, notes, domain_id),
                    )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_domains WHERE id = ?", (domain_id,))
        return jsonify({"domain": _domain_public(row)})

    @bp.route("/settings", methods=["GET"])
    @mail_perm(*MAIL_SET)
    def get_settings():
        keys = (
            "provider_mode", "smtp_host", "smtp_port", "smtp_user", "smtp_password",
            "webhook_secret", "default_domain_id",
            "smartico_affiliate_id", "smartico_subid_param",
            "scrub_smtp_verify", "scrub_rate_per_minute", "scrub_auto_suppress_invalid",
            "scrub_suppress_disposable", "scrub_suppress_role", "scrub_campaign_only_valid",
            "scrub_skip_hours", "scrub_mail_from",
        )
        with closing(get_db()) as conn:
            from mail_scrub import ensure_mail_scrub_schema, scrub_settings as _scrub_settings
            ensure_mail_scrub_schema(conn)
            settings = {k: get_mail_setting(conn, k, "") or "" for k in keys}
            # Mask password
            pw = settings.get("smtp_password") or ""
            settings["smtp_password_set"] = bool(pw)
            settings["smtp_password"] = ""
            settings["webhook_secret_masked"] = _mask_secret(settings.get("webhook_secret") or "")
            scrub = _scrub_settings(conn)
            settings["scrub"] = scrub
            try:
                sc_cfg = smartico_api.get_config(conn)
                settings["smartico_api_configured"] = bool(sc_cfg["api_key"])
                settings["smartico_api_host"] = sc_cfg["api_host"]
                settings["smartico_api_key_masked"] = (
                    smartico_api.mask_key(sc_cfg["api_key"]) if sc_cfg["api_key"] else ""
                )
            except Exception as sc_exc:
                settings["smartico_api_configured"] = False
                settings["smartico_api_host"] = ""
                settings["smartico_api_key_masked"] = ""
                settings["smartico_api_error"] = str(sc_exc)
            try:
                from mail_tenant import current_tenant_id, list_allocated_domains
                from flask import session as _sess
                _tid = current_tenant_id()
                if _sess.get("mail_is_superadmin") and not _tid:
                    dom_rows = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC") or []
                elif _tid:
                    dom_rows = list_allocated_domains(conn, int(_tid)) or []
                else:
                    dom_rows = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC") or []
            except Exception:
                dom_rows = fetchall(conn, "SELECT * FROM mail_domains ORDER BY id ASC") or []
            domains = [_domain_public(r) for r in dom_rows]
        return jsonify({"settings": settings, "domains": domains})

    @bp.route("/settings", methods=["PATCH"])
    @mail_perm(*MAIL_SET)
    def patch_settings():
        data = request.get_json(silent=True) or {}
        allowed = {
            "provider_mode", "smtp_host", "smtp_port", "smtp_user", "smtp_password",
            "webhook_secret", "default_domain_id",
            "smartico_affiliate_id", "smartico_subid_param",
            "scrub_smtp_verify", "scrub_rate_per_minute", "scrub_auto_suppress_invalid",
            "scrub_suppress_disposable", "scrub_suppress_role", "scrub_campaign_only_valid",
            "scrub_skip_hours", "scrub_mail_from",
        }
        with closing(get_db()) as conn:
            if data.get("rotate_webhook_secret"):
                upsert_mail_setting(conn, "webhook_secret", secrets.token_hex(24))
            if "smartico_api_key" in data or "smartico_api_host" in data:
                # Global Smartico rapor API key/host — mikromail'in kendi DB'sinde
                # (smartico_settings tablosu); ana panelden tamamen ayrı, superadmin-only.
                from flask import session as _sess
                if not _sess.get("mail_is_superadmin"):
                    return jsonify({"error": "Yalnızca süper admin değiştirebilir."}), 403
                existing = smartico_api.get_config(conn)
                new_key = (data.get("smartico_api_key") or "").strip() or existing["api_key"]
                new_host = (data.get("smartico_api_host") or "").strip() or existing["api_host"]
                if not new_key:
                    return jsonify({"error": "Smartico API anahtarı boş olamaz."}), 400
                smartico_api.save_config(conn, new_key, new_host)
            bool_keys = {
                "scrub_smtp_verify", "scrub_auto_suppress_invalid", "scrub_suppress_disposable",
                "scrub_suppress_role", "scrub_campaign_only_valid",
            }
            for key, val in data.items():
                if key not in allowed:
                    continue
                if key == "smtp_password" and (val is None or val == ""):
                    continue  # empty = keep existing
                if key == "smtp_password" and val is not None:
                    # Yapıştırma artığı / gizli karakter temizliği
                    val = (
                        str(val)
                        .replace("\u200b", "")
                        .replace("\ufeff", "")
                        .replace("\r", "")
                        .replace("\n", "")
                        .strip()
                    )
                if key in bool_keys:
                    if isinstance(val, bool):
                        val = "1" if val else "0"
                    else:
                        val = "1" if str(val).strip().lower() in ("1", "true", "yes", "on") else "0"
                upsert_mail_setting(conn, key, "" if val is None else str(val).strip())
            conn.commit()
            settings = {k: get_mail_setting(conn, k, "") or "" for k in allowed}
            pw = settings.get("smtp_password") or ""
            settings["smtp_password_set"] = bool(pw)
            settings["smtp_password"] = ""
            settings["webhook_secret_masked"] = _mask_secret(settings.get("webhook_secret") or "")
            from mail_scrub import scrub_settings as _scrub_settings
            settings["scrub"] = _scrub_settings(conn)
            try:
                sc_cfg = smartico_api.get_config(conn)
                settings["smartico_api_configured"] = bool(sc_cfg["api_key"])
                settings["smartico_api_host"] = sc_cfg["api_host"]
                settings["smartico_api_key_masked"] = (
                    smartico_api.mask_key(sc_cfg["api_key"]) if sc_cfg["api_key"] else ""
                )
            except Exception as sc_exc:
                settings["smartico_api_configured"] = False
                settings["smartico_api_host"] = ""
                settings["smartico_api_key_masked"] = ""
                settings["smartico_api_error"] = str(sc_exc)
        return jsonify({"settings": settings})

    @bp.route("/settings/test-smtp", methods=["POST"])
    @mail_perm(*MAIL_SET)
    def test_smtp_settings():
        """Kampanya olmadan DirectMail login testi — 535 teşhisi için."""
        data = request.get_json(silent=True) or {}
        domain_id = data.get("domain_id")
        try:
            domain_id = int(domain_id) if domain_id not in (None, "") else None
        except (TypeError, ValueError):
            domain_id = None
        from mail_delivery import smtp_login_test

        with closing(get_db()) as conn:
            if domain_id is None:
                raw = get_mail_setting(conn, "default_domain_id", "") or ""
                try:
                    domain_id = int(raw) if raw else None
                except (TypeError, ValueError):
                    domain_id = None
            # Formdaki değerlerle anlık test (henüz kaydetmeden)
            result = smtp_login_test(
                conn,
                domain_id=domain_id,
                override_password=(data.get("smtp_password") or None),
                override_user=(data.get("smtp_user") or None),
                override_host=(data.get("smtp_host") or None),
                override_port=(data.get("smtp_port") or None),
                probe_hosts=bool(data.get("probe_hosts", True)),
            )
            # Çalışan host/user bulunduysa Ayarlar’a yaz
            if result.get("ok") and data.get("save_working"):
                if result.get("host"):
                    upsert_mail_setting(conn, "smtp_host", result["host"])
                if result.get("user"):
                    upsert_mail_setting(conn, "smtp_user", result["user"])
                conn.commit()
            if result.get("ok"):
                try:
                    from mail_tenant import heal_ready_domains

                    heal_ready_domains(conn)
                except Exception:
                    pass
        status = 200 if result.get("ok") else 400
        return jsonify(result), status

    def _mask_secret(s):
        if not s:
            return ""
        if len(s) <= 8:
            return "•" * len(s)
        return s[:4] + "•" * (len(s) - 8) + s[-4:]

    # ── IVR ────────────────────────────────────────────────────
    @bp.route("/ivr/rules", methods=["GET"])
    @mail_perm(*MAIL_IVR)
    def get_ivr_rules():
        with closing(get_db()) as conn:
            rows = _rows(fetchall(conn, "SELECT * FROM mail_ivr_rules ORDER BY id ASC"))
        return jsonify({"rules": rows})

    @bp.route("/ivr/rules/<int:rule_id>", methods=["PATCH"])
    @mail_perm(*MAIL_IVR)
    def patch_ivr_rule(rule_id):
        data = request.get_json(silent=True) or {}
        with closing(get_db()) as conn:
            row = fetchone(conn, "SELECT * FROM mail_ivr_rules WHERE id = ?", (rule_id,))
            if not row:
                return jsonify({"error": "Kural bulunamadı."}), 404
            execute(
                conn,
                """
                UPDATE mail_ivr_rules SET name = ?, active = ?, template_id = ?,
                    domain_id = ?, delay_seconds = ?
                WHERE id = ?
                """,
                (
                    (data.get("name") if "name" in data else row["name"] or "").strip(),
                    1 if data.get("active", row["active"]) else 0,
                    data.get("template_id") if "template_id" in data else row["template_id"],
                    data.get("domain_id") if "domain_id" in data else row["domain_id"],
                    int(data.get("delay_seconds") if "delay_seconds" in data else row["delay_seconds"] or 0),
                    rule_id,
                ),
            )
            conn.commit()
            row = fetchone(conn, "SELECT * FROM mail_ivr_rules WHERE id = ?", (rule_id,))
        return jsonify({"rule": _row(row)})

    @bp.route("/ivr/events", methods=["GET"])
    @mail_perm(*MAIL_IVR)
    def list_ivr_events():
        try:
            limit = min(int(request.args.get("limit") or 100), 500)
        except (TypeError, ValueError):
            limit = 100
        with closing(get_db()) as conn:
            rows = _rows(fetchall(
                conn,
                "SELECT * FROM mail_ivr_events ORDER BY id DESC LIMIT ?",
                (limit,),
            ))
        return jsonify({"events": rows})

    @bp.route("/webhooks/ivr", methods=["POST"])
    def ivr_webhook():
        """Harici IVR santralinden çağrı cevabı bildirimi.
        Auth: X-Mailing-Webhook-Secret header veya ?secret= query.
        Body JSON: { phone, email?, answered_at?, name? }
        """
        data = request.get_json(silent=True) or {}
        secret = (
            request.headers.get("X-Mailing-Webhook-Secret")
            or request.args.get("secret")
            or ""
        ).strip()
        now = iso(utcnow())
        with closing(get_db()) as conn:
            expected = (get_mail_setting(conn, "webhook_secret", "") or "").strip()
            if not expected or secret != expected:
                return jsonify({"error": "Unauthorized"}), 401

            phone = str(data.get("phone") or data.get("tel") or "").strip()
            email = str(data.get("email") or "").strip().lower()
            answered_at = str(data.get("answered_at") or now).strip()
            name = str(data.get("name") or "").strip()

            event_id = insert_returning_id(
                conn,
                """
                INSERT INTO mail_ivr_events
                (phone, email, answered_at, contact_id, send_id, status, payload, error, created_at)
                VALUES (?, ?, ?, NULL, NULL, 'received', ?, '', ?)
                """,
                (phone, email, answered_at, json.dumps(data, ensure_ascii=False), now),
            )

            rule = fetchone(
                conn,
                "SELECT * FROM mail_ivr_rules WHERE active = 1 ORDER BY id ASC LIMIT 1",
            )
            if not rule:
                execute(
                    conn,
                    "UPDATE mail_ivr_events SET status = ?, error = ? WHERE id = ?",
                    ("skipped", "Aktif IVR kuralı yok", event_id),
                )
                conn.commit()
                return jsonify({"ok": True, "event_id": event_id, "status": "skipped", "reason": "no_active_rule"})

            if not rule["template_id"] or not rule["domain_id"]:
                execute(
                    conn,
                    "UPDATE mail_ivr_events SET status = ?, error = ? WHERE id = ?",
                    ("skipped", "IVR kuralında şablon/domain eksik", event_id),
                )
                conn.commit()
                return jsonify({"ok": True, "event_id": event_id, "status": "skipped", "reason": "rule_incomplete"})

            # Match contact by phone or email
            contact = None
            if phone:
                contact = fetchone(
                    conn,
                    "SELECT * FROM mail_contacts WHERE phone != '' AND phone = ? LIMIT 1",
                    (phone,),
                )
            if not contact and email:
                contact = fetchone(
                    conn,
                    "SELECT * FROM mail_contacts WHERE LOWER(email) = ? LIMIT 1",
                    (email,),
                )
            # Auto-create contact if email provided
            if not contact and email and EMAIL_RE.match(email):
                cid = insert_returning_id(
                    conn,
                    """
                    INSERT INTO mail_contacts
                    (email, phone, name, tags, source, unsubscribed, notes, created_at, updated_at)
                    VALUES (?, ?, ?, ?, 'ivr', 0, '', ?, ?)
                    """,
                    (email, phone, name, _tags_json(["ivr"]), now, now),
                )
                contact = fetchone(conn, "SELECT * FROM mail_contacts WHERE id = ?", (cid,))

            contact = _row(contact) if contact else None

            if not contact or not contact.get("email"):
                execute(
                    conn,
                    "UPDATE mail_ivr_events SET status = ?, error = ? WHERE id = ?",
                    ("no_contact", "Eşleşen kontak/e-posta yok", event_id),
                )
                conn.commit()
                return jsonify({"ok": True, "event_id": event_id, "status": "no_contact"})

            if contact.get("unsubscribed"):
                execute(
                    conn,
                    "UPDATE mail_ivr_events SET status = ?, contact_id = ?, error = ? WHERE id = ?",
                    ("unsubscribed", contact["id"], "Kontak abonelikten çıkmış", event_id),
                )
                conn.commit()
                return jsonify({"ok": True, "event_id": event_id, "status": "unsubscribed"})

            tpl = fetchone(conn, "SELECT * FROM mail_templates WHERE id = ?", (rule["template_id"],))
            if not tpl:
                execute(
                    conn,
                    "UPDATE mail_ivr_events SET status = ?, contact_id = ?, error = ? WHERE id = ?",
                    ("error", contact["id"], "Şablon bulunamadı", event_id),
                )
                conn.commit()
                return jsonify({"ok": False, "event_id": event_id, "status": "error"}), 400

            contact_d = _contact_out(contact)
            subject = _render_template(tpl["subject"], contact_d)
            html_body = _render_template(
                tpl.get("html_body") or _plain_to_html(tpl.get("text_body") or ""),
                contact_d,
            )
            text_body = _render_template(tpl.get("text_body") or "", contact_d)
            delay_sec = int(rule.get("delay_seconds") or 0)
            if delay_sec > 0:
                # Webhook'u uzun tutmamak için üst sınır
                time.sleep(min(delay_sec, 30))
            from mail_delivery import deliver_mail
            mode = (get_mail_setting(conn, "provider_mode", "stub") or "stub").strip().lower()
            send_id, status, err = deliver_mail(
                conn,
                channel="ivr",
                to_email=contact_d["email"],
                subject=subject,
                contact=contact_d,
                contact_id=contact_d["id"],
                template_id=rule["template_id"],
                domain_id=rule["domain_id"],
                to_phone=phone or contact_d.get("phone") or "",
                html_body=html_body,
                text_body=text_body,
                inject_tracking=_inject_tracking,
            )
            execute(
                conn,
                "UPDATE mail_ivr_events SET status = ?, contact_id = ?, send_id = ?, error = ? WHERE id = ?",
                (status, contact_d["id"], send_id, err or "", event_id),
            )
            conn.commit()
        return jsonify({
            "ok": True,
            "event_id": event_id,
            "send_id": send_id,
            "status": status,
            "mode": mode,
        })

    return bp
