"""Fatura Hesaplama — Excel şablonu oluşturma ve toplu yükleme."""

from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime
from io import BytesIO

from database import execute, fetchall, fetchone, insert_returning_id, iso, scalar, utcnow, uses_postgres

SECTION_LABELS_TR = {
    "sport": "Spor Bahisleri",
    "casino": "Casino Bahisleri",
    "special": "Özel Kalemler",
}

HEADER_DATE = frozenset({"tarih", "entry_date", "date", "gun"})
HEADER_PROVIDER = frozenset({"saglayici", "provider", "provider_name", "saglayici_adi", "provider adi"})
HEADER_GGR = frozenset({"ggr", "ggr_amount", "ggr_try", "toplam ggr", "ggr try", "ggr amount"})
HEADER_STAKE = frozenset({"stake", "stake_amount", "stake_try", "toplam stake", "stake try"})
HEADER_WINNING = frozenset({"winning", "winning_amount", "winning_try", "toplam winning", "winning try", "kazanc"})


def capture_invoice_calc_day_rates(conn, entry_date, rates=None):
    """Kayıt anındaki canlı kuru o güne kilitle — önceki günlere dokunmaz."""
    from accounting_fx import fetch_exchange_rates

    period = entry_date[:7]
    now = iso(utcnow())
    if rates is None:
        rates = fetch_exchange_rates(fresh=True)
    usd = float(rates.get("usd_try") or 0)
    eur = float(rates.get("eur_try") or 0)
    if usd <= 0 or eur <= 0:
        return False
    source = (rates.get("source") or "").strip()
    captured = rates.get("fetched_at")
    if hasattr(captured, "isoformat"):
        captured = captured.isoformat()
    else:
        captured = str(captured or now)
    exists = scalar(conn, "SELECT COUNT(*) FROM acc_invoice_calc_day_meta WHERE entry_date = ?", (entry_date,))
    if exists:
        execute(
            conn,
            """
            UPDATE acc_invoice_calc_day_meta
            SET usd_try = ?, eur_try = ?, rate_source = ?, captured_at = ?, updated_at = ?
            WHERE entry_date = ?
            """,
            (usd, eur, source, captured, now, entry_date),
        )
    else:
        # Bu tabloda id yok — insert_returning_id(RETURNING id) Postgres'te transaction'ı bozar.
        execute(
            conn,
            """
            INSERT INTO acc_invoice_calc_day_meta
            (entry_date, period, usd_try, eur_try, rate_source, captured_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (entry_date, period, usd, eur, source, captured, now),
        )
    return True


def _safe_capture_day_rates(conn, entry_date, rates=None):
    """Kur kaydı opsiyonel; hata upsert transaction'ını abort etmemeli."""
    sp = "invoice_calc_fx"
    if uses_postgres():
        try:
            execute(conn, f"SAVEPOINT {sp}")
            capture_invoice_calc_day_rates(conn, entry_date, rates=rates)
            execute(conn, f"RELEASE SAVEPOINT {sp}")
            return True
        except Exception:
            try:
                execute(conn, f"ROLLBACK TO SAVEPOINT {sp}")
            except Exception:
                try:
                    conn.rollback()
                except Exception:
                    pass
            return False
    try:
        capture_invoice_calc_day_rates(conn, entry_date, rates=rates)
        return True
    except Exception:
        return False


def load_invoice_calc_day_rates(conn, period):
    rows = fetchall(
        conn,
        """
        SELECT entry_date, usd_try, eur_try, rate_source, captured_at
        FROM acc_invoice_calc_day_meta
        WHERE period = ?
        """,
        (period,),
    )
    out = {}
    for r in rows:
        key = r["entry_date"]
        if hasattr(key, "isoformat") and not isinstance(key, str):
            key = key.isoformat()
        else:
            key = str(key or "").strip()[:10]
        out[key] = {
            "usd_try": round(float(r["usd_try"] or 0), 6),
            "eur_try": round(float(r["eur_try"] or 0), 6),
            "rate_source": r["rate_source"] or "",
            "captured_at": r["captured_at"] or "",
        }
    return out


def _apply_fx_amounts(try_amount, usd_try, eur_try):
    try_amount = float(try_amount or 0)
    if not usd_try or not eur_try:
        return None, None
    return round(try_amount / usd_try, 2), round(try_amount / eur_try, 2)


def enrich_invoice_calc_fx(daily_totals, day_rates):
    """Günlük toplamlara kilitli kur ile USD/EUR karşılıklarını ekle; ay grand total FX döner."""
    fx_grand = {
        "ggr_usd": 0.0, "ggr_eur": 0.0,
        "commission_usd": 0.0, "commission_eur": 0.0,
    }
    enriched = []
    for item in daily_totals:
        d = dict(item)
        rate = day_rates.get(d["entry_date"])
        if rate and rate["usd_try"] and rate["eur_try"]:
            d["usd_try"] = rate["usd_try"]
            d["eur_try"] = rate["eur_try"]
            for field in ("ggr", "commission"):
                usd, eur = _apply_fx_amounts(d.get(field + "_amount"), rate["usd_try"], rate["eur_try"])
                d[field + "_usd"] = usd
                d[field + "_eur"] = eur
                if usd is not None:
                    fx_grand[field + "_usd"] += usd
                if eur is not None:
                    fx_grand[field + "_eur"] += eur
        else:
            d["usd_try"] = d["eur_try"] = None
            for field in ("ggr", "commission"):
                d[field + "_usd"] = None
                d[field + "_eur"] = None
        enriched.append(d)
    for k in fx_grand:
        fx_grand[k] = round(fx_grand[k], 2)
    return enriched, fx_grand


def _norm_header(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.strip().lower()
    text = text.replace("ı", "i").replace("ğ", "g").replace("ü", "u").replace("ş", "s").replace("ö", "o").replace("ç", "c")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text.replace(" ", "_")


def _normalize_provider_name(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\u00a0", " ").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _load_workbook_bytes(raw):
    import openpyxl

    bio = BytesIO(raw)
    try:
        return openpyxl.load_workbook(bio, read_only=True, data_only=True)
    except Exception:
        bio.seek(0)
        return openpyxl.load_workbook(bio, data_only=True)


def _ensure_invoice_calc_schema(conn):
    cols = _table_columns(conn)
    if cols and "ggr_amount" not in cols:
        raise ValueError(
            "Veritabanı şeması güncel değil (ggr_amount eksik). Deploy sonrası sayfayı yenileyip tekrar deneyin."
        )


def _table_columns(conn):
    from database import _table_columns as db_table_columns

    return db_table_columns(conn, "acc_invoice_calc_daily")


def _parse_amount(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        amount = float(value)
    else:
        text = str(value).strip()
        negative = text.startswith("-")
        text = text.lstrip("-")
        text = text.replace(".", "").replace(",", ".") if re.search(r",\d{1,2}$", text) else text.replace(",", "")
        try:
            amount = float(text)
        except (TypeError, ValueError):
            return None
        if negative:
            amount = -amount
    return round(amount, 2)


def _parse_ggr(value):
    """GGR negatif olabilir — signed parse."""
    result = _parse_amount(value)
    return result if result is not None else None


def _parse_entry_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date().isoformat()
    except ValueError:
        return None


def _map_headers(header_row):
    mapping = {}
    for idx, raw in enumerate(header_row):
        key = _norm_header(raw)
        if key in HEADER_DATE or "tarih" in key:
            mapping["entry_date"] = idx
        elif key in HEADER_PROVIDER or key.startswith("saglayici"):
            mapping["provider_name"] = idx
        elif key in HEADER_GGR or key == "ggr" or key.startswith("ggr"):
            mapping["ggr_amount"] = idx
        elif key in HEADER_STAKE or key.startswith("stake"):
            mapping["stake_amount"] = idx
        elif key in HEADER_WINNING or key.startswith("winning") or key.startswith("kazanc"):
            mapping["winning_amount"] = idx
    return mapping


def upsert_invoice_calc_daily_batch(conn, entry_date, rows, who="", capture_fx=True):
    """rows: [{provider_id, ggr_amount}, ...]"""
    period = entry_date[:7]
    now = iso(utcnow())
    who = (who or "").strip()
    valid_ids = {r["id"] for r in fetchall(conn, "SELECT id FROM acc_invoice_calc_providers")}
    saved = 0
    deleted = 0
    skipped = 0
    for item in rows:
        try:
            pid = int(item.get("provider_id"))
        except (TypeError, ValueError):
            skipped += 1
            continue
        if pid not in valid_ids:
            skipped += 1
            continue
        ggr = _parse_ggr(item.get("ggr_amount"))
        if ggr is None:
            skipped += 1
            continue
        exists = scalar(
            conn,
            "SELECT COUNT(*) FROM acc_invoice_calc_daily WHERE entry_date = ? AND provider_id = ?",
            (entry_date, pid),
        )
        if ggr == 0:
            if exists:
                execute(
                    conn,
                    "DELETE FROM acc_invoice_calc_daily WHERE entry_date = ? AND provider_id = ?",
                    (entry_date, pid),
                )
                deleted += 1
            continue
        if exists:
            execute(
                conn,
                """
                UPDATE acc_invoice_calc_daily
                SET ggr_amount = ?, stake_amount = 0, winning_amount = 0, created_by = ?, updated_at = ?
                WHERE entry_date = ? AND provider_id = ?
                """,
                (ggr, who, now, entry_date, pid),
            )
        else:
            insert_returning_id(
                conn,
                """
                INSERT INTO acc_invoice_calc_daily
                (period, entry_date, provider_id, ggr_amount, stake_amount, winning_amount, created_by, updated_at)
                VALUES (?, ?, ?, ?, 0, 0, ?, ?)
                """,
                (period, entry_date, pid, ggr, who, now),
            )
        saved += 1
    if saved > 0 and capture_fx:
        has_rows = scalar(
            conn,
            "SELECT COUNT(*) FROM acc_invoice_calc_daily WHERE entry_date = ?",
            (entry_date,),
        )
        meta_exists = scalar(
            conn,
            "SELECT COUNT(*) FROM acc_invoice_calc_day_meta WHERE entry_date = ?",
            (entry_date,),
        )
        if has_rows and not meta_exists:
            _safe_capture_day_rates(conn, entry_date)
    return {"saved": saved, "deleted": deleted, "skipped": skipped}


def generate_template_bytes(conn, entry_date):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    entry_dt = datetime.strptime(entry_date, "%Y-%m-%d")
    fmt_date = "DD.MM.YYYY"
    fmt_ggr = '"₺" #.##0,00'
    fmt_rate = "0"

    providers = fetchall(
        conn,
        """
        SELECT id, section, name, commission_rate
        FROM acc_invoice_calc_providers
        WHERE active = 1
        ORDER BY CASE section WHEN 'sport' THEN 0 WHEN 'casino' THEN 1 WHEN 'special' THEN 2 ELSE 3 END,
                 sort_order, name
        """,
    )
    existing = {}
    for row in fetchall(
        conn,
        """
        SELECT provider_id, ggr_amount, stake_amount, winning_amount
        FROM acc_invoice_calc_daily
        WHERE entry_date = ?
        """,
        (entry_date,),
    ):
        existing[row["provider_id"]] = dict(row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gunluk Veri"

    headers = ["Tarih", "Sağlayıcı", "Bölüm", "Oran %", "GGR (TRY)"]
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    last_row = 1
    for p in providers:
        ex = existing.get(p["id"], {})
        c_date = ws.cell(row=row_idx, column=1, value=entry_dt)
        c_date.number_format = fmt_date
        ws.cell(row=row_idx, column=2, value=p["name"])
        ws.cell(row=row_idx, column=3, value=SECTION_LABELS_TR.get(p["section"], p["section"]))
        c_rate = ws.cell(row=row_idx, column=4, value=float(p["commission_rate"] or 0))
        c_rate.number_format = fmt_rate
        ggr_val = ex.get("ggr_amount")
        if ggr_val is None and ex:
            ggr_val = float(ex.get("stake_amount") or 0) - float(ex.get("winning_amount") or 0)
        c_ggr = ws.cell(row=row_idx, column=5)
        c_ggr.number_format = fmt_ggr
        if ggr_val not in (None, ""):
            c_ggr.value = float(ggr_val)
        last_row = row_idx
        row_idx += 1

    # Boş GGR hücreleri de sayı biçiminde kalsın — Genel biçime çevrilince import bozulmasın.
    if last_row >= 2:
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=1).number_format = fmt_date
            ws.cell(row=r, column=4).number_format = fmt_rate
            ws.cell(row=r, column=5).number_format = fmt_ggr

    widths = [12, 42, 18, 10, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Aciklama")
    info["A1"] = "Fatura Hesaplama — Excel Yükleme Kılavuzu"
    info["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "1. 'Gunluk Veri' sekmesindeki GGR (TRY) sütununu Pronet panelinden doldurun.",
        "2. Tarih sütunu tarih, GGR sütunu sayı biçimindedir — sütun biçimini Genel yapmayın.",
        "3. Tarih sütununu değiştirmeyin (tek gün) veya her satırda farklı gün kullanabilirsiniz.",
        "4. Sağlayıcı adlarını değiştirmeyin — paneldeki isimlerle birebir eşleşmelidir.",
        "5. Boş bırakılan veya 0 olan satırlar yüklenmez / mevcut kayıt silinir.",
        "6. GGR negatif olabilir (oyuncu kazancı fazlaysa).",
        "7. Panel → Muhasebe → Fatura Hesaplama → Excel Yükle ile dosyayı yükleyin.",
        "",
        f"Şablon tarihi: {entry_date}",
        f"Sağlayıcı sayısı: {len(providers)}",
    ]
    for i, line in enumerate(lines, 2):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 90

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def import_workbook(conn, file_storage, who=""):
    filename = (file_storage.filename or "").lower()
    if not filename.endswith((".xlsx", ".xlsm")):
        raise ValueError("Sadece .xlsx dosyası yükleyebilirsiniz.")

    raw = file_storage.read()
    if not raw:
        raise ValueError("Dosya boş.")
    if len(raw) > 15 * 1024 * 1024:
        raise ValueError("Dosya çok büyük (en fazla 15 MB).")

    _ensure_invoice_calc_schema(conn)

    wb = _load_workbook_bytes(raw)
    try:
        ws = wb["Gunluk Veri"] if "Gunluk Veri" in wb.sheetnames else wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel dosyası boş.")

        col_map = _map_headers(header)
        if "provider_name" not in col_map:
            raise ValueError("Excel'de 'Sağlayıcı' sütunu bulunamadı.")
        has_ggr = "ggr_amount" in col_map
        has_stake_win = "stake_amount" in col_map and "winning_amount" in col_map
        if not has_ggr and not has_stake_win:
            raise ValueError("Excel'de 'GGR (TRY)' sütunu gerekli.")

        provider_rows = fetchall(conn, "SELECT id, name FROM acc_invoice_calc_providers")
        name_to_id = {_normalize_provider_name(r["name"]).lower(): r["id"] for r in provider_rows}

        grouped = {}
        unknown = []
        skipped_rows = 0
        parsed_rows = 0

        for values in rows_iter:
            if not values or all(v is None or str(v).strip() == "" for v in values):
                continue
            parsed_rows += 1

            def cell(key):
                idx = col_map.get(key)
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            entry_date = _parse_entry_date(cell("entry_date"))
            if not entry_date and "entry_date" not in col_map:
                raise ValueError("Excel'de 'Tarih' sütunu bulunamadı.")
            if not entry_date:
                skipped_rows += 1
                continue

            pname = _normalize_provider_name(cell("provider_name"))
            if not pname:
                skipped_rows += 1
                continue
            pid = name_to_id.get(pname.lower())
            if not pid:
                unknown.append(pname)
                continue

            if has_ggr:
                ggr = _parse_ggr(cell("ggr_amount"))
            else:
                stake = _parse_amount(cell("stake_amount"))
                winning = _parse_amount(cell("winning_amount"))
                ggr = round(stake - winning, 2) if stake is not None and winning is not None else None
            if ggr is None:
                skipped_rows += 1
                continue

            grouped.setdefault(entry_date, []).append({
                "provider_id": pid,
                "ggr_amount": ggr,
            })

        if not grouped:
            if unknown:
                sample = ", ".join(sorted(set(unknown))[:5])
                raise ValueError(
                    f"Eşleşen sağlayıcı bulunamadı ({len(set(unknown))} farklı isim). "
                    f"Şablonu panelden yeniden indirin. Örnek: {sample}"
                )
            raise ValueError("İşlenecek satır bulunamadı. GGR değerlerini kontrol edin.")

        totals = {"saved": 0, "deleted": 0, "skipped": 0}
        dates = []
        for entry_date in sorted(grouped.keys()):
            result = upsert_invoice_calc_daily_batch(
                conn, entry_date, grouped[entry_date], who=who, capture_fx=False
            )
            totals["saved"] += result["saved"]
            totals["deleted"] += result["deleted"]
            totals["skipped"] += result["skipped"]
            dates.append(entry_date)

        bulk_rates = None
        if dates:
            try:
                from accounting_fx import fetch_exchange_rates

                bulk_rates = fetch_exchange_rates(fresh=True)
            except Exception:
                bulk_rates = None

        for entry_date in dates:
            has_rows = scalar(
                conn,
                "SELECT COUNT(*) FROM acc_invoice_calc_daily WHERE entry_date = ?",
                (entry_date,),
            )
            meta_exists = scalar(
                conn,
                "SELECT COUNT(*) FROM acc_invoice_calc_day_meta WHERE entry_date = ?",
                (entry_date,),
            )
            if has_rows and not meta_exists:
                _safe_capture_day_rates(conn, entry_date, rates=bulk_rates)

        conn.commit()
        unknown_unique = sorted(set(unknown))
        periods = sorted({d[:7] for d in dates})
        return {
            "ok": True,
            "dates": dates,
            "periods": periods,
            "parsed_rows": parsed_rows,
            "saved": totals["saved"],
            "deleted": totals["deleted"],
            "skipped_rows": skipped_rows + totals["skipped"],
            "unknown_providers": unknown_unique[:30],
            "unknown_count": len(unknown_unique),
        }
    finally:
        wb.close()
